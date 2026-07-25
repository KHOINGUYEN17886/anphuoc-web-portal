import openpyxl
import sqlite3
import os
import psycopg2
from datetime import datetime, timedelta

excel_path = r'C:\All_Report\1_Mapping\StaffList_Store_20.04.26.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Sheet1']

db_path = r'C:\All_Report\8_RETAIL_COMMANDER\tools\web_portal\operational_data.db'
conn_sqlite = sqlite3.connect(db_path)
cur_sqlite = conn_sqlite.cursor()

# Get store map from SQLite
cur_sqlite.execute('SELECT store_code, store_name FROM tb_stores')
stores = cur_sqlite.fetchall()
store_map = {}
for code, name in stores:
    store_map[name.strip().lower()] = code
    clean_name = name.replace('CH', '').replace('Cửa hàng', '').strip().lower()
    store_map[clean_name] = code

def parse_excel_date(val):
    if not val:
        return datetime.today().strftime('%Y-%m-%d')
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    try:
        s = str(val).strip()
        if s.isdigit() and len(s) == 5:
            d = datetime(1899, 12, 30) + timedelta(days=int(s))
            return d.strftime('%Y-%m-%d')
        elif s.replace('.', '', 1).isdigit() and 40000 <= float(s) <= 60000:
            d = datetime(1899, 12, 30) + timedelta(days=int(float(s)))
            return d.strftime('%Y-%m-%d')
        elif '/' in s:
            parts = s.split('/')
            if len(parts) == 3:
                return f"{parts[2].zfill(4)}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        elif len(s) >= 10 and '-' in s:
            return s[:10]
    except Exception:
        pass
    return datetime.today().strftime('%Y-%m-%d')

emp_records = []
for r in range(3, ws.max_row + 1):
    emp_code = ws.cell(row=r, column=1).value
    emp_name = ws.cell(row=r, column=2).value
    pos = ws.cell(row=r, column=4).value
    hire_date_raw = ws.cell(row=r, column=5).value
    dob_raw = ws.cell(row=r, column=6).value
    gender = ws.cell(row=r, column=7).value
    workplace = ws.cell(row=r, column=9).value
    
    if not emp_code or not emp_name:
        continue
        
    emp_code = str(emp_code).strip()
    emp_name = str(emp_name).strip()
    if emp_name.lower() == 'x' or len(emp_name) < 2:
        continue
        
    pos = str(pos).strip() if pos else 'Nhân viên bán hàng'
    if pos == 'CH Trưởng':
        pos = 'Cửa hàng trưởng'
    elif pos == 'CH Phó':
        pos = 'Cửa hàng phó'
        
    gender = str(gender).strip() if gender else 'Nữ'
    hire_date = parse_excel_date(hire_date_raw)
    dob = str(dob_raw).strip() if dob_raw else ''
    workplace_clean = str(workplace).strip().lower() if workplace else ''
    
    st_code = store_map.get(workplace_clean)
    if not st_code:
        for sname, scode in store_map.items():
            if sname in workplace_clean or workplace_clean in sname:
                st_code = scode
                break
                
    if not st_code:
        st_code = 'STORE_GENERAL'
        
    avatar_url = f"https://ui-avatars.com/api/?name={emp_name.replace(' ', '+')}&background=1B2A4A&color=fff"
    emp_records.append((emp_code, st_code, emp_name, pos, gender, dob, hire_date, avatar_url))

print(f"Parsed {len(emp_records)} valid employee records.")

# Upsert into local SQLite
cur_sqlite.execute("DELETE FROM tb_store_employees")
for rec in emp_records:
    cur_sqlite.execute("""
        INSERT INTO tb_store_employees (employee_code, store_code, full_name, position, gender, dob, appointment_date, avatar_url, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVE', ?)
    """, (rec[0], rec[1], rec[2], rec[3], rec[4], rec[5], rec[6], rec[7], rec[6]))

conn_sqlite.commit()
print("Updated local SQLite operational_data.db successfully!")

# Upsert into Neon Postgres if available
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    try:
        conn_pg = psycopg2.connect(DATABASE_URL)
        cur_pg = conn_pg.cursor()
        cur_pg.execute("DELETE FROM tb_store_employees")
        for rec in emp_records:
            cur_pg.execute("""
                INSERT INTO tb_store_employees (employee_code, store_code, full_name, position, gender, dob, appointment_date, avatar_url, status, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE', %s)
            """, (rec[0], rec[1], rec[2], rec[3], rec[4], rec[5], rec[6], rec[7], rec[6]))
        conn_pg.commit()
        print("Updated Neon PostgreSQL database successfully!")
    except Exception as e:
        print("Postgres sync notice:", e)
