import pandas as pd
import sqlite3
import os
import random

DATABASE_URL = os.environ.get('DATABASE_URL')
stores_info_path = r"C:\All_Report\1_Mapping\StoresInfo.xlsx"
passcodes_path = r"C:\All_Report\1_Mapping\Store_Passcodes.xlsx"

# 1. Load existing PINs or create new mapping
existing_pins = {}
if os.path.exists(passcodes_path):
    try:
        df_pins = pd.read_excel(passcodes_path)
        df_pins['store_code'] = df_pins['store_code'].astype(str).str.strip()
        df_pins['passcode'] = df_pins['passcode'].astype(str).str.strip()
        existing_pins = dict(zip(df_pins['store_code'], df_pins['passcode']))
        print(f"Loaded {len(existing_pins)} existing PINs from Store_Passcodes.xlsx")
    except Exception as e:
        print(f"Warning reading Store_Passcodes.xlsx: {e}")

is_postgres = False
if DATABASE_URL:
    try:
        import psycopg2
        print("Connecting to Neon PostgreSQL Cloud Database...")
        conn = psycopg2.connect(DATABASE_URL)
        is_postgres = True
    except ImportError:
        print("psycopg2 not installed. Falling back to local SQLite...")
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'operational_data.db')
        conn = sqlite3.connect(db_path)
else:
    print("DATABASE_URL env var not found. Connecting to local SQLite database...")
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'operational_data.db')
    conn = sqlite3.connect(db_path)

cursor = conn.cursor()

# Create tables
id_type = "SERIAL PRIMARY KEY" if is_postgres else "INTEGER PRIMARY KEY AUTOINCREMENT"

cursor.execute("""
CREATE TABLE IF NOT EXISTS tb_stores (
    store_code TEXT PRIMARY KEY,
    store_name TEXT NOT NULL,
    brand TEXT NOT NULL,
    region TEXT,
    asm_name TEXT,
    passcode TEXT DEFAULT '1234'
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS tb_asms (
    asm_name TEXT PRIMARY KEY,
    passcode TEXT DEFAULT '9999',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")

# Note: tb_traffic is now DAILY. report_date is renamed/treated as traffic_date.
# data_source: 'store_actual' (cửa hàng tự nhập, mặc định) vs 'synthetic_backfill'
# (nhồi tự động cho khoảng trống lịch sử — xem tools/qlkd_operational/backfill_traffic_synthetic.py).
cursor.execute("""
CREATE TABLE IF NOT EXISTS tb_traffic (
    store_code TEXT NOT NULL,
    traffic_date TEXT NOT NULL,
    traffic_val INTEGER NOT NULL,
    bills_val INTEGER DEFAULT 0,
    company_online_bills INTEGER DEFAULT 0,
    store_online_bills INTEGER DEFAULT 0,
    data_source TEXT DEFAULT 'store_actual',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY (store_code, traffic_date),
    FOREIGN KEY (store_code) REFERENCES tb_stores(store_code)
)
""")

# Alter table to add new columns to tb_traffic if they don't exist
# LƯU Ý (Postgres): mỗi ALTER phải commit() ngay khi thành công — nếu để chung
# 1 transaction, một ALTER lỗi sau đó (cột đã tồn tại) sẽ rollback() xoá luôn
# các ALTER thành công TRƯỚC ĐÓ nhưng chưa commit (bug đã phát hiện 2026-07-22:
# cột data_source bị mất dấu vết dù log in "Added column..." vì ALTER
# contract_number ngay sau đó fail rồi rollback cả transaction).
try:
    cursor.execute("ALTER TABLE tb_traffic ADD COLUMN company_online_bills INTEGER DEFAULT 0")
    conn.commit()
    print("Added column company_online_bills to tb_traffic")
except Exception as e:
    conn.rollback()

try:
    cursor.execute("ALTER TABLE tb_traffic ADD COLUMN store_online_bills INTEGER DEFAULT 0")
    conn.commit()
    print("Added column store_online_bills to tb_traffic")
except Exception as e:
    conn.rollback()

try:
    cursor.execute("ALTER TABLE tb_traffic ADD COLUMN data_source TEXT DEFAULT 'store_actual'")
    conn.commit()
    print("Added column data_source to tb_traffic")
except Exception as e:
    conn.rollback()

try:
    cursor.execute("ALTER TABLE tb_contracts ADD COLUMN contract_number TEXT DEFAULT 'Đang GD'")
    conn.commit()
    print("Added column contract_number to tb_contracts")
except Exception as e:
    conn.rollback()

cursor.execute(f"""
CREATE TABLE IF NOT EXISTS tb_contracts (
    id {id_type},
    store_code TEXT NOT NULL,
    report_date TEXT NOT NULL,
    contract_number TEXT DEFAULT 'Đang GD',
    contract_value REAL NOT NULL,
    product_category TEXT NOT NULL,
    quantity INTEGER NOT NULL,
    deposit_paid REAL DEFAULT 0.0,
    product_category TEXT NOT NULL,
    quantity INTEGER NOT NULL,
    deposit_paid REAL DEFAULT 0.0,
    installment_2 REAL DEFAULT 0.0,
    status TEXT NOT NULL,
    reason TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (store_code) REFERENCES tb_stores(store_code)
)
""")

cursor.execute(f"""
CREATE TABLE IF NOT EXISTS tb_unsigned_contracts (
    id {id_type},
    store_code TEXT NOT NULL,
    report_date TEXT NOT NULL,
    prev_year_value REAL NOT NULL,
    expected_signing_time TEXT NOT NULL,
    product_category TEXT NOT NULL,
    quantity INTEGER NOT NULL,
    status TEXT NOT NULL,
    reason TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (store_code) REFERENCES tb_stores(store_code)
)
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS tb_operational_details (
    store_code TEXT NOT NULL,
    report_date TEXT NOT NULL,
    op_open_close_status TEXT DEFAULT 'Đạt',
    op_open_close_note TEXT,
    op_uniform_status TEXT DEFAULT 'Đạt',
    op_uniform_note TEXT,
    op_greet_status TEXT DEFAULT 'Đạt',
    op_greet_note TEXT,
    op_feedback_status TEXT DEFAULT 'Đạt',
    op_feedback_note TEXT,
    op_other_status TEXT DEFAULT 'Đạt',
    op_other_note TEXT,
    hr_target INTEGER DEFAULT 0,
    hr_actual INTEGER DEFAULT 0,
    hr_guard INTEGER DEFAULT 0,
    hr_resigned_note TEXT,
    hr_leave_note TEXT,
    hr_absent_note TEXT,
    inv_stock_status TEXT,
    inv_info_goods TEXT,
    inv_return_warehouse TEXT,
    inv_proposal TEXT,
    market_product_feedback TEXT,
    market_missing_products TEXT,
    market_competitors TEXT,
    market_other_feedback TEXT,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY (store_code, report_date),
    FOREIGN KEY (store_code) REFERENCES tb_stores(store_code)
)
""")

cursor.execute(f"""
CREATE TABLE IF NOT EXISTS tb_support_requests (
    id {id_type},
    store_code TEXT NOT NULL,
    report_date TEXT NOT NULL,
    category TEXT NOT NULL,
    priority TEXT DEFAULT 'Trung bình',
    issue_item TEXT NOT NULL,
    deadline TEXT,
    person_in_charge TEXT DEFAULT 'QLKD / ASM',
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    FOREIGN KEY (store_code) REFERENCES tb_stores(store_code)
)
""")

conn.commit()

# 2. Parse StoresInfo.xlsx
if os.path.exists(stores_info_path):
    print(f"Reading StoresInfo.xlsx from {stores_info_path}...")
    df = pd.read_excel(stores_info_path, sheet_name='Danh bạ CH')
    
    # Columns expected: CODE, TÊN CỬA HÀNG - CHUẨN, VÙNG, ASM
    code_col = 'CODE'
    name_col = 'TÊN CỬA HÀNG - CHUẨN'
    region_col = 'VÙNG'
    asm_col = 'ASM'
    
    # We will do update/upsert on stores so we don't accidentally lose any PINs
    new_pins_list = []
    inserted = 0
    
    for _, row in df.iterrows():
        code = str(row.get(code_col, '')).strip()
        name = str(row.get(name_col, '')).strip()
        if not code or code == 'nan' or not name or name == 'nan':
            continue
        
        reg = str(row.get(region_col, '')).strip() if pd.notna(row.get(region_col)) else 'Khác'
        asm = str(row.get(asm_col, '')).strip() if pd.notna(row.get(asm_col)) else 'Chưa gán'
        br = 'AP'
        
        # PIN handling
        pin = existing_pins.get(code)
        if not pin:
            pin = f"{random.randint(1000, 9999)}"
            existing_pins[code] = pin
            
        new_pins_list.append({
            'store_code': code,
            'store_name': name,
            'asm_name': asm,
            'passcode': pin
        })
        
        try:
            if is_postgres:
                cursor.execute("""
                INSERT INTO tb_stores (store_code, store_name, brand, region, asm_name, passcode)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (store_code) DO UPDATE SET 
                    store_name = EXCLUDED.store_name,
                    brand = EXCLUDED.brand,
                    region = EXCLUDED.region,
                    asm_name = EXCLUDED.asm_name
                """, (code, name, br, reg, asm, pin))
            else:
                cursor.execute("""
                INSERT OR REPLACE INTO tb_stores (store_code, store_name, brand, region, asm_name, passcode)
                VALUES (?, ?, ?, ?, ?, ?)
                """, (code, name, br, reg, asm, pin))
            inserted += 1
        except Exception as e:
            print(f"Error inserting {code}: {e}")
            
    conn.commit()
    
    # Extract unique ASMs and seed tb_asms
    unique_asms = set(item['asm_name'] for item in new_pins_list if item['asm_name'])
    for asm_name in unique_asms:
        if asm_name in ('Chưa gán', 'Khác', 'nan'):
            continue
        try:
            if is_postgres:
                cursor.execute("""
                INSERT INTO tb_asms (asm_name, passcode)
                VALUES (%s, %s)
                ON CONFLICT (asm_name) DO NOTHING
                """, (asm_name, '9999'))
            else:
                cursor.execute("""
                INSERT OR IGNORE INTO tb_asms (asm_name, passcode)
                VALUES (?, ?)
                """, (asm_name, '9999'))
        except Exception as e:
            print(f"Error seeding ASM {asm_name}: {e}")
    conn.commit()
    
    # Save PINs spreadsheet
    df_out_pins = pd.DataFrame(new_pins_list)
    df_out_pins.to_excel(passcodes_path, index=False)
    print(f"Saved store PINs list to {passcodes_path}")
    
    db_type_str = "Neon PostgreSQL Cloud" if is_postgres else "SQLite Local"
    print(f"Successfully loaded {inserted} active stores into {db_type_str} database from StoresInfo.xlsx!")
else:
    print(f"Error: StoresInfo.xlsx not found at {stores_info_path}")

# 3. Parse StaffList_Store_20.04.26.xlsx to seed real employee data
staff_list_path = r"C:\All_Report\1_Mapping\StaffList_Store_20.04.26.xlsx"
if os.path.exists(staff_list_path):
    print(f"Seeding tb_store_employees from {staff_list_path}...")
    import openpyxl
    from datetime import datetime, timedelta, date
    
    wb_staff = openpyxl.load_workbook(staff_list_path, data_only=True)
    ws_staff = wb_staff['Sheet1']
    
    cursor.execute("SELECT store_code, store_name FROM tb_stores")
    st_rows = cursor.fetchall()
    st_map = {}
    for scode, sname in st_rows:
        st_map[sname.strip().lower()] = scode
        clean_name = sname.replace('CH', '').replace('Cửa hàng', '').strip().lower()
        st_map[clean_name] = scode
        
    def parse_excel_date_py(val):
        if not val:
            return datetime.today().strftime('%Y-%m-%d')
        if isinstance(val, (datetime, date)):
            return val.strftime('%Y-%m-%d')
        try:
            s = str(val).strip()
            if s.isdigit() and len(s) == 5:
                d = datetime(1899, 12, 30) + timedelta(days=int(s))
                return d.strftime('%Y-%m-%d')
            elif s.replace('.', '', 1).isdigit() and 40000 <= float(s) <= 60000:
                d = datetime(1899, 12, 30) + timedelta(days=int(float(s)))
                return d.strftime('%Y-%m-%d')
            elif '/' in s:
                parts = s.split('/')
                if len(parts) == 3:
                    return f"{parts[2].zfill(4)}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
            elif len(s) >= 10 and '-' in s:
                return s[:10]
        except Exception:
            pass
        return datetime.today().strftime('%Y-%m-%d')

    emp_count = 0
    for r in range(3, ws_staff.max_row + 1):
        emp_code = ws_staff.cell(row=r, column=1).value
        emp_name = ws_staff.cell(row=r, column=2).value
        pos = ws_staff.cell(row=r, column=4).value
        hire_date_raw = ws_staff.cell(row=r, column=5).value
        dob_raw = ws_staff.cell(row=r, column=6).value
        gender = ws_staff.cell(row=r, column=7).value
        workplace = ws_staff.cell(row=r, column=9).value
        
        if not emp_code or not emp_name:
            continue
        emp_code = str(emp_code).strip()
        emp_name = str(emp_name).strip()
        if emp_name.lower() == 'x' or len(emp_name) < 2:
            continue
            
        pos = str(pos).strip() if pos else 'Nhân viên bán hàng'
        if pos == 'CH Trưởng': pos = 'Cửa hàng trưởng'
        elif pos == 'CH Phó': pos = 'Cửa hàng phó'
        
        gender = str(gender).strip() if gender else 'Nữ'
        hire_date = parse_excel_date_py(hire_date_raw)
        dob = str(dob_raw).strip() if dob_raw else ''
        workplace_clean = str(workplace).strip().lower() if workplace else ''
        
        st_code = st_map.get(workplace_clean)
        if not st_code:
            for sname, scode in st_map.items():
                if sname in workplace_clean or workplace_clean in sname:
                    st_code = scode
                    break
        if not st_code: st_code = 'STORE_GENERAL'
        
        avatar_url = f"https://ui-avatars.com/api/?name={emp_name.replace(' ', '+')}&background=1B2A4A&color=fff"
        
        try:
            if is_postgres:
                cursor.execute("""
                    INSERT INTO tb_store_employees (employee_code, store_code, full_name, position, gender, dob, appointment_date, avatar_url, status, created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE', %s)
                    ON CONFLICT (employee_code) DO UPDATE SET
                        store_code = EXCLUDED.store_code,
                        full_name = EXCLUDED.full_name,
                        position = EXCLUDED.position,
                        gender = EXCLUDED.gender,
                        appointment_date = EXCLUDED.appointment_date
                """, (emp_code, st_code, emp_name, pos, gender, dob, hire_date, avatar_url, hire_date))
            else:
                cursor.execute("""
                    INSERT INTO tb_store_employees (employee_code, store_code, full_name, position, gender, dob, appointment_date, avatar_url, status, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVE', ?)
                    ON CONFLICT(employee_code) DO UPDATE SET
                        store_code = excluded.store_code,
                        full_name = excluded.full_name,
                        position = excluded.position,
                        gender = excluded.gender,
                        appointment_date = excluded.appointment_date
                """, (emp_code, st_code, emp_name, pos, gender, dob, hire_date, avatar_url, hire_date))
            emp_count += 1
        except Exception as e:
            print(f"Error seeding employee {emp_code}: {e}")
            
    conn.commit()
    print(f"Successfully seeded {emp_count} real employees into database from StaffList_Store_20.04.26.xlsx!")

conn.close()
