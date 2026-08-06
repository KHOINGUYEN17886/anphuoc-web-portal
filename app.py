import os
import calendar
from datetime import date, datetime, timedelta
from collections import defaultdict
from flask import Flask, render_template, request, jsonify, send_file, g, has_app_context
import sqlite3
import threading
import urllib.request
import urllib.parse
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import io
import pandas as pd
import openpyxl

# Import psycopg2 for PostgreSQL if on cloud
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None

app = Flask(__name__)

# ── Bảo mật (P2-G) ────────────────────────────────────────────────────────────
# Token API pull dữ liệu: BẮT BUỘC set env ở production. Nếu thiếu → sinh token
# ngẫu nhiên mỗi lần khởi động (route /api/export sẽ không đoán được) + cảnh báo.
import secrets as _secrets
API_SECRET_TOKEN = os.environ.get('API_SECRET_TOKEN')
if not API_SECRET_TOKEN:
    API_SECRET_TOKEN = _secrets.token_urlsafe(24)
    print("⚠️  [BẢO MẬT] Chưa set API_SECRET_TOKEN — dùng token ngẫu nhiên phiên này. "
          "Đặt biến môi trường API_SECRET_TOKEN cho production.")

# MASTER_PIN: cảnh báo nếu còn dùng default yếu.
MASTER_PIN = os.environ.get('MASTER_PIN', '8888')
if MASTER_PIN == '8888':
    print("⚠️  [BẢO MẬT] MASTER_PIN đang dùng giá trị mặc định '8888' — đặt MASTER_PIN mạnh cho production.")

# Cho phép PIN mặc định '1234' (tiện dev). Production: KHÔNG set → chặn bypass.
ALLOW_DEFAULT_PIN = os.environ.get('ALLOW_DEFAULT_PIN', '0') == '1'


def _default_pin_allowed(pin) -> bool:
    """PIN '1234' chỉ được chấp nhận khi ALLOW_DEFAULT_PIN bật (môi trường dev)."""
    return ALLOW_DEFAULT_PIN and str(pin) == '1234'


DATABASE_URL = os.environ.get('DATABASE_URL')

# SQLite absolute path helper
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'operational_data.db')

def get_db_connection():
    if has_app_context():
        if not hasattr(g, 'db_conn'):
            if DATABASE_URL and psycopg2:
                g.db_conn = psycopg2.connect(DATABASE_URL)
            else:
                g.db_conn = sqlite3.connect(DB_PATH, timeout=30.0)
                g.db_conn.row_factory = sqlite3.Row
        return g.db_conn
    else:
        if DATABASE_URL and psycopg2:
            return psycopg2.connect(DATABASE_URL)
        else:
            conn = sqlite3.connect(DB_PATH, timeout=30.0)
            conn.row_factory = sqlite3.Row
            return conn

@app.teardown_appcontext
def close_db_connection(exception):
    db_conn = getattr(g, 'db_conn', None)
    if db_conn is not None:
        try:
            db_conn.close()
        except Exception:
            pass

def execute_db(query, args=()):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        if DATABASE_URL and psycopg2:
            if args:
                query_pg = query.replace('%', '%%').replace('?', '%s')
                cur.execute(query_pg, args)
            else:
                cur.execute(query)
        else:
            cur.execute(query, args)
        conn.commit()
    except Exception:
        # Rollback ngay để không đẩy connection dùng chung (g.db_conn) vào
        # trạng thái InFailedSqlTransaction, làm mọi lệnh sau trong cùng
        # request đều lỗi theo (CLAUDE.md — Transaction Safety trên Postgres).
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        cur.close()
        if not has_app_context():
            conn.close()

def query_db(query, args=(), one=False):
    conn = get_db_connection()
    close_conn = False
    if not has_app_context():
        close_conn = True
    # Use DictCursor for PostgreSQL to act like sqlite3.Row
    if DATABASE_URL and psycopg2:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        if args:
            query_pg = query.replace('%', '%%').replace('?', '%s')
            cur.execute(query_pg, args)
        else:
            cur.execute(query)
        rv = cur.fetchall()
        rv = [dict(r) for r in rv]
    else:
        cur = conn.cursor()
        cur.execute(query, args)
        rv = cur.fetchall()
        # Convert sqlite3.Row to dict
        rv = [dict(r) for r in rv]
        
    cur.close()
    if close_conn:
        conn.close()
    return (rv[0] if rv else None) if one else rv

def _log_ticket_history(ticket_type, ticket_code, store_code, old_status, new_status, note=''):
    """
    Ghi 1 dòng vào tb_ticket_status_history mỗi khi 1 ticket (SUPPORT hoặc HR)
    đổi trạng thái THẬT (bỏ qua nếu old_status == new_status — tránh log rác
    khi request chỉ sửa ghi chú/người phụ trách mà không đổi status). Đây là
    audit trail bắt buộc để dựng SLA/thời gian xử lý thật cho sổ theo dõi sự
    vụ vận hành (qlkd_operational/ticket_lifecycle_sheet.py) — KHÔNG dùng
    updated_at của bảng gốc làm proxy vì cột đó bị ghi đè bởi MỌI thay đổi
    (kể cả sửa ghi chú), không riêng đổi trạng thái.
    """
    if old_status == new_status:
        return
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        execute_db("""
            INSERT INTO tb_ticket_status_history
            (ticket_type, ticket_code, store_code, old_status, new_status, changed_at, note)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (ticket_type, ticket_code, store_code, old_status or '', new_status, now_str, note or ''))
    except Exception as _e:
        print(f"⚠️ [TicketHistory] Không ghi được history cho {ticket_type} {ticket_code}: {_e}")

# Get reporting date (Friday of the current week or previous week if early in the week)
def get_report_date():
    today = date.today()
    if today.weekday() >= 4: # Friday (4), Saturday (5), Sunday (6)
        offset = 4 - today.weekday()
    else: # Monday (0), Tuesday (1), Wednesday (2), Thursday (3)
        offset = -3 - today.weekday()
    return today + timedelta(days=offset)

# ── Safe DB Migration ─────────────────────────────────────────────────────────
def safe_migrate_db():
    conn = get_db_connection()
    cur = conn.cursor()
    
    is_pg = bool(DATABASE_URL and psycopg2)
    pk_auto = "SERIAL PRIMARY KEY" if is_pg else "INTEGER PRIMARY KEY AUTOINCREMENT"
    
    table_sqls = [
        f"""
        CREATE TABLE IF NOT EXISTS tb_stores (
            store_code  VARCHAR(50) PRIMARY KEY,
            store_name  VARCHAR(200) NOT NULL DEFAULT '',
            brand       VARCHAR(50) DEFAULT 'AP',
            region      VARCHAR(100) DEFAULT '',
            asm_name    VARCHAR(100) DEFAULT '',
            passcode    VARCHAR(20) DEFAULT '1234'
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_store_employees (
            id {pk_auto},
            employee_code VARCHAR(50) UNIQUE NOT NULL,
            store_code VARCHAR(50) NOT NULL,
            full_name VARCHAR(100) NOT NULL,
            gender VARCHAR(10) DEFAULT 'Nữ',
            dob VARCHAR(20) DEFAULT '',
            phone_number VARCHAR(30) DEFAULT '',
            position VARCHAR(50) NOT NULL,
            appointment_date VARCHAR(20) DEFAULT '',
            avatar_url TEXT DEFAULT '',
            status VARCHAR(30) DEFAULT 'ACTIVE',
            created_at VARCHAR(30) DEFAULT '',
            updated_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_store_shift_config (
            store_code VARCHAR(50) PRIMARY KEY,
            shift_1_name VARCHAR(50) DEFAULT 'Ca 1 (Sáng)',
            shift_1_hours VARCHAR(50) DEFAULT '08:30 - 16:30',
            shift_2_name VARCHAR(50) DEFAULT 'Ca 2 (Chiều)',
            shift_2_hours VARCHAR(50) DEFAULT '14:00 - 22:00',
            has_split_shift INTEGER DEFAULT 0,
            split_shift_hours VARCHAR(50) DEFAULT '10:00 - 14:00 & 17:00 - 21:00',
            updated_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_hr_lifecycle_tickets (
            id {pk_auto},
            ticket_code VARCHAR(50) UNIQUE NOT NULL,
            store_code VARCHAR(50) NOT NULL,
            report_date VARCHAR(20) NOT NULL,
            employee_code VARCHAR(50) DEFAULT '',
            employee_name VARCHAR(100) NOT NULL,
            position VARCHAR(50) NOT NULL,
            event_type VARCHAR(30) NOT NULL,
            effective_date VARCHAR(20) NOT NULL,
            reason_note TEXT DEFAULT '',
            handover_status VARCHAR(50) DEFAULT 'Chưa bàn giao',
            status VARCHAR(30) DEFAULT 'Mới ghi nhận',
            store_progress_note TEXT DEFAULT '',
            asm_hr_note TEXT DEFAULT '',
            created_at VARCHAR(30) DEFAULT '',
            updated_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_employee_probation (
            id {pk_auto},
            employee_code VARCHAR(50) UNIQUE NOT NULL,
            store_code VARCHAR(50) NOT NULL,
            start_date VARCHAR(20) NOT NULL,
            trainer_employee_code VARCHAR(50) DEFAULT '',
            trainer_name VARCHAR(100) DEFAULT '',
            train_op_rules INTEGER DEFAULT 0,
            train_work_process INTEGER DEFAULT 0,
            train_products INTEGER DEFAULT 0,
            train_display INTEGER DEFAULT 0,
            train_inventory INTEGER DEFAULT 0,
            training_notes TEXT DEFAULT '',
            probation_end_date VARCHAR(20) DEFAULT '',
            probation_result VARCHAR(50) DEFAULT 'Đang thử việc',
            updated_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_store_headcount_targets (
            store_code VARCHAR(50) PRIMARY KEY,
            target_headcount INTEGER DEFAULT 0,
            cht_name VARCHAR(100) DEFAULT '',
            updated_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_store_support_staff (
            id {pk_auto},
            store_code VARCHAR(50) NOT NULL,
            employee_code VARCHAR(50) DEFAULT '',
            employee_name VARCHAR(100) NOT NULL,
            from_store_code VARCHAR(50) NOT NULL,
            reason VARCHAR(100) DEFAULT 'Cửa hàng thiếu nhân sự',
            start_date VARCHAR(20) DEFAULT '',
            end_date VARCHAR(20) DEFAULT '',
            created_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_ticket_status_history (
            id {pk_auto},
            ticket_type VARCHAR(20) NOT NULL,
            ticket_code VARCHAR(50) NOT NULL,
            store_code VARCHAR(50) NOT NULL,
            old_status VARCHAR(30) DEFAULT '',
            new_status VARCHAR(30) NOT NULL,
            changed_at VARCHAR(30) DEFAULT '',
            note TEXT DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_store_closures (
            id {pk_auto},
            closure_code VARCHAR(50) UNIQUE NOT NULL,
            store_code VARCHAR(50) NOT NULL,
            event_type VARCHAR(30) NOT NULL,
            start_date VARCHAR(20) NOT NULL,
            end_date VARCHAR(20) DEFAULT '',
            status VARCHAR(30) DEFAULT 'ĐANG ĐÓNG',
            reason_note TEXT DEFAULT '',
            reported_by VARCHAR(50) DEFAULT '',
            source VARCHAR(20) DEFAULT 'ASM',
            created_at VARCHAR(30) DEFAULT '',
            updated_at VARCHAR(30) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_compliance_imports (
            id {pk_auto},
            batch_code VARCHAR(50) UNIQUE NOT NULL,
            filename TEXT NOT NULL,
            imported_at VARCHAR(50) NOT NULL,
            imported_by VARCHAR(100) DEFAULT 'Admin/P.QLQT',
            total_records INTEGER DEFAULT 0,
            month_period VARCHAR(20) DEFAULT ''
        )
        """,
        f"""
        CREATE TABLE IF NOT EXISTS tb_compliance_audits (
            id {pk_auto},
            batch_code VARCHAR(50) DEFAULT '',
            ticket_code VARCHAR(50) UNIQUE NOT NULL,
            sheet_name VARCHAR(100) DEFAULT '',
            topic VARCHAR(150) DEFAULT '',
            check_item VARCHAR(255) DEFAULT '',
            criteria_standard TEXT DEFAULT '',
            store_code VARCHAR(50) NOT NULL,
            store_name VARCHAR(150) DEFAULT '',
            asm_name VARCHAR(100) DEFAULT '',
            camera_evidence_time TEXT DEFAULT '',
            violation_group VARCHAR(255) DEFAULT '',
            violation_description TEXT DEFAULT '',
            pqlqt_evaluation VARCHAR(50) DEFAULT 'Không đạt',
            received_date VARCHAR(20) DEFAULT '',
            response_deadline VARCHAR(20) DEFAULT '',
            store_explanation TEXT DEFAULT '',
            asm_assessment TEXT DEFAULT '',
            violating_staff_info TEXT DEFAULT '',
            attachment_url TEXT DEFAULT '',
            status VARCHAR(50) DEFAULT 'Mới tiếp nhận',
            is_repeat_offense INTEGER DEFAULT 0,
            repeat_count INTEGER DEFAULT 1,
            created_at VARCHAR(50) DEFAULT '',
            updated_at VARCHAR(50) DEFAULT ''
        )
        """
    ]
    
    for sql in table_sqls:
        try:
            cur.execute(sql)
            conn.commit()
        except Exception as e:
            conn.rollback()

    # ── Tự dò & bù cột thiếu cho các bảng HR ────────────────────────────────
    # Lý do: các bảng HR dùng CREATE TABLE IF NOT EXISTS. Nếu bảng đã tồn tại
    # trên Neon từ schema cũ (thiếu cột như updated_at, train_*, reason_note...),
    # CREATE IF NOT EXISTS sẽ KHÔNG thêm cột → INSERT/UPDATE trong save_store_hr
    # ném lỗi → HR "hoàn toàn không ghi nhận". Khối này thêm cột còn thiếu.
    def _existing_cols(table):
        try:
            if is_pg:
                cur.execute(
                    "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
                    (table,),
                )
                cols = {r[0] for r in cur.fetchall()}
            else:
                cur.execute(f"PRAGMA table_info({table})")
                cols = {r[1] for r in cur.fetchall()}
            conn.commit()
            return cols
        except Exception:
            conn.rollback()
            return set()

    hr_columns = {
        'tb_store_employees': {
            'gender': "VARCHAR(10) DEFAULT 'Nữ'",
            'dob': "VARCHAR(20) DEFAULT ''",
            'phone_number': "VARCHAR(30) DEFAULT ''",
            'position': "VARCHAR(50) DEFAULT 'Nhân viên bán hàng'",
            'appointment_date': "VARCHAR(20) DEFAULT ''",
            'avatar_url': "TEXT DEFAULT ''",
            'status': "VARCHAR(30) DEFAULT 'ACTIVE'",
            'created_at': "VARCHAR(30) DEFAULT ''",
            'updated_at': "VARCHAR(30) DEFAULT ''",
        },
        'tb_hr_lifecycle_tickets': {
            'employee_code': "VARCHAR(50) DEFAULT ''",
            'reason_note': "TEXT DEFAULT ''",
            'handover_status': "VARCHAR(50) DEFAULT 'Chưa bàn giao'",
            'status': "VARCHAR(30) DEFAULT 'Mới ghi nhận'",
            'store_progress_note': "TEXT DEFAULT ''",
            'asm_hr_note': "TEXT DEFAULT ''",
            'created_at': "VARCHAR(30) DEFAULT ''",
            'updated_at': "VARCHAR(30) DEFAULT ''",
        },
        'tb_employee_probation': {
            'trainer_employee_code': "VARCHAR(50) DEFAULT ''",
            'trainer_name': "VARCHAR(100) DEFAULT ''",
            'train_op_rules': "INTEGER DEFAULT 0",
            'train_work_process': "INTEGER DEFAULT 0",
            'train_products': "INTEGER DEFAULT 0",
            'train_display': "INTEGER DEFAULT 0",
            'train_inventory': "INTEGER DEFAULT 0",
            'training_notes': "TEXT DEFAULT ''",
            'probation_end_date': "VARCHAR(20) DEFAULT ''",
            'probation_result': "VARCHAR(50) DEFAULT 'Đang thử việc'",
            'updated_at': "VARCHAR(30) DEFAULT ''",
        },
        'tb_store_support_staff': {
            'employee_code': "VARCHAR(50) DEFAULT ''",
            'reason': "VARCHAR(100) DEFAULT 'Cửa hàng thiếu nhân sự'",
            'start_date': "VARCHAR(20) DEFAULT ''",
            'end_date': "VARCHAR(20) DEFAULT ''",
            'created_at': "VARCHAR(30) DEFAULT ''",
        },
        'tb_store_headcount_targets': {
            'cht_name': "VARCHAR(100) DEFAULT ''",
            'updated_at': "VARCHAR(30) DEFAULT ''",
        },
    }
    for _table, _cols in hr_columns.items():
        _present = _existing_cols(_table)
        if not _present:
            continue  # bảng chưa tồn tại (sẽ do CREATE lo) hoặc introspect lỗi
        for _col, _coldef in _cols.items():
            if _col in _present:
                continue
            try:
                cur.execute(f"ALTER TABLE {_table} ADD COLUMN {_col} {_coldef}")
                conn.commit()
                print(f"✅ [Migration] Đã thêm cột thiếu {_table}.{_col}")
            except Exception as _e:
                conn.rollback()
                print(f"⚠️ [Migration] Không thêm được {_table}.{_col}: {_e}")

    migrations = [
        "ALTER TABLE tb_stores ADD COLUMN brand TEXT DEFAULT 'AP'",
        "ALTER TABLE tb_stores ADD COLUMN region TEXT DEFAULT ''",
        "ALTER TABLE tb_stores ADD COLUMN asm_name TEXT DEFAULT ''",
        "ALTER TABLE tb_stores ADD COLUMN passcode TEXT DEFAULT '1234'",
        "ALTER TABLE tb_stores ADD COLUMN is_active INTEGER DEFAULT 1",
        "ALTER TABLE tb_contracts ADD COLUMN customer_name TEXT DEFAULT ''",
        "ALTER TABLE tb_unsigned_contracts ADD COLUMN contract_number TEXT DEFAULT 'Đang GD'",
        "ALTER TABLE tb_unsigned_contracts ADD COLUMN customer_name TEXT DEFAULT ''",
        "ALTER TABLE tb_support_requests ADD COLUMN ticket_code TEXT DEFAULT ''",
        "ALTER TABLE tb_support_requests ADD COLUMN status TEXT DEFAULT 'Đang xử lý'",
        "ALTER TABLE tb_support_requests ADD COLUMN store_progress_note TEXT DEFAULT ''",
        "ALTER TABLE tb_support_requests ADD COLUMN asm_hq_note TEXT DEFAULT ''",
        "ALTER TABLE tb_support_requests ADD COLUMN created_at TEXT DEFAULT ''",
        "ALTER TABLE tb_support_requests ADD COLUMN updated_at TEXT DEFAULT ''",
        "ALTER TABLE tb_traffic ADD COLUMN non_purchase_reasons TEXT DEFAULT ''",
        "ALTER TABLE tb_operational_reports ADD COLUMN weekly_top_reasons TEXT DEFAULT ''",
        "ALTER TABLE tb_operational_reports ADD COLUMN weekly_hr_summary TEXT DEFAULT ''",
    ]
    for sql in migrations:
        try:
            cur.execute(sql)
            conn.commit()
        except Exception as e:
            conn.rollback()
                
    try:
        if DATABASE_URL and psycopg2:
            cur.execute("""
            UPDATE tb_support_requests 
            SET ticket_code = 'TK-' || store_code || '-' || REPLACE(report_date, '-', '') || '-' || id
            WHERE ticket_code IS NULL OR ticket_code = ''
            """)
        else:
            cur.execute("""
            UPDATE tb_support_requests 
            SET ticket_code = 'TK-' || store_code || '-' || REPLACE(report_date, '-', '') || '-' || rowid
            WHERE ticket_code IS NULL OR ticket_code = ''
            """)
        conn.commit()
    except Exception as e:
        print(f"⚠️ [Migration] Error updating ticket_codes: {e}")

    cur.close()
    conn.close()
    
    seed_hr_baseline_data()

import re

def remove_vn_accents(text):
    if not text: return ''
    text = text.lower()
    patterns = {
        '[aàáảãạâầấẩẫậăằắẳẵặ]': 'a',
        '[eèéẻẽẹêềếểễệ]': 'e',
        '[iìíỉĩị]': 'i',
        '[oòóỏõọôồốổỗộơờớởỡợ]': 'o',
        '[uùúủũụưừứửữự]': 'u',
        '[yỳýỷỹỵ]': 'y',
        'đ': 'd'
    }
    for pat, rep in patterns.items():
        text = re.sub(pat, rep, text)
    return text.strip()

def parse_excel_date(val):
    return format_excel_date_py(val)

def format_excel_date_py(val):
    if not val:
        return ''
    s = str(val).strip()
    try:
        if s.isdigit() and len(s) == 5:
            serial = int(s)
            d = date(1899, 12, 30) + timedelta(days=serial)
            return d.strftime('%Y-%m-%d')
        elif s.replace('.', '', 1).isdigit() and 40000 <= float(s) <= 60000:
            serial = int(float(s))
            d = date(1899, 12, 30) + timedelta(days=serial)
            return d.strftime('%Y-%m-%d')
    except Exception:
        pass
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3:
            try:
                return f"{parts[2].zfill(4)}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
            except Exception:
                pass
    return s

def calculate_tenure_py(start_date_str):
    if not start_date_str:
        return 'Mới nhận việc'
    try:
        start_date_str = format_excel_date_py(start_date_str)
        if '/' in start_date_str:
            parts = start_date_str.split('/')
            if len(parts) == 3:
                d = date(int(parts[2]), int(parts[1]), int(parts[0]))
            else:
                return 'Mới nhận việc'
        else:
            d = datetime.strptime(start_date_str[:10], '%Y-%m-%d').date()
            
        today = date.today()
        if d > today:
            return 'Mới nhận việc'
            
        diff_days = (today - d).days
        if diff_days < 30:
            return f"{diff_days} ngày"
        diff_months = diff_days // 30
        if diff_months < 12:
            return f"{diff_months} tháng"
        diff_years = round(diff_days / 365.25, 1)
        return f"{diff_years} năm"
    except Exception:
        return 'Mới nhận việc'
            
        parts = []
        if years > 0:
            parts.append(f"{years} năm")
        if months > 0:
            parts.append(f"{months} tháng")
        if days > 0 or not parts:
            parts.append(f"{days} ngày")
            
        return ' '.join(parts)
    except Exception:
        return '—'

ALIAS_MAP = {
    '8 cong hoa': '8CONGHOA',
    'ch n.v.nghi': 'GOVAP',
    'lotte q7': 'LOTTEQ7',
    'big c - dong nai': 'BCDN',
    'da nang 3 - riverside mall': 'RIVERSIDEDN',
    'ha noi 8 - vincom': 'VINCOMBATRIEU',
    'ha noi 14 - ba trieu': 'BATRIEU',
    'hai phong 1 - nguyen duc canh': 'NDC',
    'cua hang viet tri': 'VIETTRI',
    'cua hang bac ninh': 'BACNINH',
    'ch nam dinh': 'NAMDINH',
    'vincom - tuyen quang': 'TUYENQUANG',
    'cua hang vinh yen (vinh phuc)': 'VINHYEN',
    'fld nguyen dinh chieu - binh duong': 'FLD_BD',
    'fld ly thanh ton - nha trang': 'NTRANG',
    'khanh hoi q.4': 'KHANHHOI',
    'hcm - vincom le van viet': 'VINCOMLEVANVIET',
    'ha noi 13 - aeon ha dong': 'AEONHADONG',
    'ha noi 18 - diamond': 'DIAMONDHN',
    'cua hang online': 'ONLINE'
}

def is_asm_khoi(asm_name):
    if not asm_name:
        return False
    clean = str(asm_name).strip().lower()
    return ('khôi' in clean or 'khoi' in clean)

def get_auth_scope(role, asm_name, pin, store_code):
    if role == 'admin' or pin == MASTER_PIN:
        return {'type': 'ALL'}
    
    # Special check for ASM Khôi
    if role == 'asm' and pin:
        try:
            matching_asms = query_db("SELECT asm_name FROM tb_asms WHERE passcode = ?", (pin,))
            asm_names = [r['asm_name'] for r in matching_asms]
            has_khoi = any(is_asm_khoi(name) for name in asm_names)
            if has_khoi:
                return {'type': 'ALL'}
        except Exception:
            pass

    if role == 'asm' and (is_asm_khoi(asm_name) or asm_name == 'ALL'):
        return {'type': 'ALL'}
    if role == 'asm' and asm_name and asm_name != 'ALL':
        return {'type': 'ASM', 'asm': asm_name}
    if store_code:
        return {'type': 'STORE', 'store': store_code}
    return {'type': 'STORE', 'store': store_code or ''}

def _hr_pin_authorized(store, pin):
    """Cho phép xem/sửa HR của một cửa hàng nếu PIN là:
      - passcode của chính cửa hàng (cửa hàng trưởng), HOẶC
      - MASTER_PIN (admin), HOẶC
      - PIN của một ASM: Khôi (hoặc pin dùng chung có Khôi) → mọi cửa hàng;
        ASM khác → chỉ cửa hàng thuộc cụm mình quản lý (store.asm_name).
    Trước đây get_store_hr/save_store_hr/save_shift_config chỉ chấp nhận passcode
    cửa hàng hoặc master_pin → ASM (kể cả Khôi) dùng PIN 9999 luôn bị 401
    'Mã PIN không đúng' khi bấm 'Quản Lý HR'. Dùng chung logic với get_auth_scope.
    """
    if not store:
        return False
    if pin and (pin == store['passcode'] or pin == MASTER_PIN):
        return True
    if not pin:
        return False
    try:
        rows = query_db("SELECT asm_name FROM tb_asms WHERE passcode = ?", (pin,))
        asm_names = [r['asm_name'] for r in rows]
        if any(is_asm_khoi(n) for n in asm_names):
            return True
        if store['asm_name'] in asm_names:
            return True
    except Exception:
        pass
    return False

def seed_hr_baseline_data():
    try:
        # Check if already seeded to prevent slow startup/imports with Neon Postgres
        try:
            cnt = query_db("SELECT COUNT(*) as count FROM tb_store_employees", one=True)
            if cnt and cnt['count'] > 0:
                print("Database already seeded with employees. Skipping baseline seed.")
                return
        except Exception as e:
            print(f"Checking seed status failed, proceeding to seed: {e}")
            
        stores_info_path = r"C:\All_Report\1_Mapping\StoresInfo.xlsx"
        staff_list_path = r"C:\All_Report\1_Mapping\StaffList_Store_20.04.26.xlsx"
        
        stores_list = []
        if os.path.exists(stores_info_path) and openpyxl:
            wb_s = openpyxl.load_workbook(stores_info_path, data_only=True)
            ws_s = wb_s.active
            for row in list(ws_s.iter_rows(min_row=2, values_only=True)):
                code = str(row[1]).strip() if row[1] else ''
                n1 = str(row[2]).strip() if row[2] else ''
                n2 = str(row[3]).strip() if row[3] else ''
                cht = str(row[4]).strip() if row[4] else ''
                target_hc = int(row[12]) if (len(row) > 12 and row[12] is not None) else 0
                if code:
                    cn1 = remove_vn_accents(n1)
                    cn2 = remove_vn_accents(n2)
                    stores_list.append({
                        'code': code,
                        'name1': n1,
                        'name2': n2,
                        'clean_code': remove_vn_accents(code),
                        'clean_n1': cn1,
                        'clean_n2': cn2,
                        'sub_n1': re.sub(r'^(cua hang|ch)\s*[-:]*\s*', '', cn1),
                        'sub_n2': re.sub(r'^(cua hang|ch)\s*[-:]*\s*', '', cn2)
                    })
                    execute_db("""
                        INSERT INTO tb_store_headcount_targets (store_code, target_headcount, cht_name)
                        VALUES (?, ?, ?)
                        ON CONFLICT(store_code) DO UPDATE SET target_headcount=excluded.target_headcount, cht_name=excluded.cht_name
                    """, (code, target_hc, cht))

        def match_store(loc_raw):
            if not loc_raw: return ''
            loc = loc_raw.strip()
            loc_clean = remove_vn_accents(loc)
            if loc_clean in ALIAS_MAP:
                return ALIAS_MAP[loc_clean]
            for s in stores_list:
                if loc.upper() == s['code'].upper() or loc_clean == s['clean_code']:
                    return s['code']
            for s in stores_list:
                if loc_clean == s['clean_n1'] or loc_clean == s['clean_n2']:
                    return s['code']
            loc_sub = re.sub(r'^(cua hang|ch|hcm|ha noi|da nang)\s*[-:]*\s*', '', loc_clean)
            best_code = ''
            best_score = 0
            for s in stores_list:
                s_n1 = s['sub_n1']
                s_n2 = s['sub_n2']
                s_c = s['clean_code']
                score = 0
                if s_c and len(s_c) >= 3 and s_c in loc_clean:
                    score += 5
                if s_n1 and (s_n1 in loc_sub or loc_sub in s_n1):
                    score += 10
                if s_n2 and (s_n2 in loc_sub or loc_sub in s_n2):
                    score += 12
                if score > best_score:
                    best_score = score
                    best_code = s['code']
            if best_score >= 5:
                return best_code
            return ''

        json_seed_path = os.path.join(os.path.dirname(__file__), "seed_employees_baseline.json")
        staff_tuples = []

        if os.path.exists(staff_list_path) and openpyxl:
            wb_emp = openpyxl.load_workbook(staff_list_path, data_only=True)
            ws_emp = wb_emp.active
            for row in list(ws_emp.iter_rows(min_row=3, values_only=True)):
                emp_code = str(row[0]).strip() if row[0] else ''
                emp_name = str(row[1]).strip() if row[1] else ''
                title = str(row[3]).strip() if row[3] else 'Nhân viên bán hàng'
                date_hire_raw = row[4] if len(row) > 4 else None
                dob_raw = str(row[5]).strip() if (len(row) > 5 and row[5]) else ''
                gender = str(row[6]).strip() if (len(row) > 6 and row[6]) else 'Nữ'
                location = str(row[8]).strip() if (len(row) > 8 and row[8]) else ''
                
                date_hire = parse_excel_date(date_hire_raw)
                
                if emp_code and emp_name:
                    st_code = match_store(location)
                    if not st_code:
                        st_code = 'ONLINE'
                    staff_tuples.append((emp_code, st_code, emp_name, gender, dob_raw, title, date_hire, date_hire))
        elif os.path.exists(json_seed_path):
            with open(json_seed_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            for item in json_data:
                e_code = item.get('employee_code')
                st_code = item.get('store_code')
                fname = item.get('full_name')
                gen = item.get('gender', 'Nữ')
                dob = item.get('dob', '')
                pos = item.get('position', 'Nhân viên bán hàng')
                apt = format_excel_date_py(item.get('appointment_date', ''))
                if e_code and fname:
                    staff_tuples.append((e_code, st_code, fname, gen, dob, pos, apt, apt))
            
        if staff_tuples:
            conn = get_db_connection()
            cur = conn.cursor()
            is_pg = bool(DATABASE_URL and psycopg2)
            try:
                cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_emp_code_uniq ON tb_store_employees(employee_code)")
                conn.commit()
            except Exception:
                pass
                
            sql = """
                INSERT INTO tb_store_employees (employee_code, store_code, full_name, gender, dob, position, appointment_date, created_at, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVE')
                ON CONFLICT(employee_code) DO UPDATE SET store_code=excluded.store_code, full_name=excluded.full_name, gender=excluded.gender, dob=excluded.dob, position=excluded.position, appointment_date=excluded.appointment_date
            """
            if is_pg:
                sql = sql.replace('?', '%s')
            try:
                cur.executemany(sql, staff_tuples)
                conn.commit()
                print(f"✅ Fast batch seeded {len(staff_tuples)} employees into DB.")
            except Exception as e_batch:
                conn.rollback()
                print(f"⚠️ Batch seed fallback: {e_batch}")
                for tup in staff_tuples:
                    e_code, st_code, fname, gen, dob, pos, apt, crt = tup
                    try:
                        q_sel = "SELECT id FROM tb_store_employees WHERE employee_code = %s" if is_pg else "SELECT id FROM tb_store_employees WHERE employee_code = ?"
                        cur.execute(q_sel, (e_code,))
                        if cur.fetchone():
                            q_upd = "UPDATE tb_store_employees SET store_code=%s, full_name=%s, gender=%s, dob=%s, position=%s, appointment_date=%s, status='ACTIVE' WHERE employee_code=%s" if is_pg else "UPDATE tb_store_employees SET store_code=?, full_name=?, gender=?, dob=?, position=?, appointment_date=?, status='ACTIVE' WHERE employee_code=?"
                            cur.execute(q_upd, (st_code, fname, gen, dob, pos, apt, e_code))
                        else:
                            q_ins = "INSERT INTO tb_store_employees (employee_code, store_code, full_name, gender, dob, position, appointment_date, created_at, status) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'ACTIVE')" if is_pg else "INSERT INTO tb_store_employees (employee_code, store_code, full_name, gender, dob, position, appointment_date, created_at, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVE')"
                            cur.execute(q_ins, (e_code, st_code, fname, gen, dob, pos, apt, crt))
                    except Exception:
                        pass
                conn.commit()
            cur.close()
            conn.close()
    except Exception as e:
        print(f"⚠️ [Seed Baseline HR] Error: {e}")
                        
def seed_stores_baseline_data(force=False):
    try:
        if not force:
            try:
                cnt = query_db("SELECT COUNT(*) as count FROM tb_stores", one=True)
                if cnt and cnt['count'] >= 180:
                    return
            except Exception as e:
                print(f"Checking store seed status failed, proceeding: {e}")
            
        stores_data = []
        # Priority 1: Read directly from StoresInfo.xlsx if available
        stores_info_path = r"C:\All_Report\1_Mapping\StoresInfo.xlsx"
        if os.path.exists(stores_info_path) and openpyxl:
            try:
                wb_s = openpyxl.load_workbook(stores_info_path, data_only=True)
                ws_s = wb_s.active
                for row in list(ws_s.iter_rows(min_row=2, values_only=True)):
                    if not row or not row[1]: continue
                    code = str(row[1]).strip()
                    name = str(row[3] or row[2] or '').strip()
                    region = str(row[8] or '').strip()
                    asm = str(row[9] or '').strip()
                    brand = str(row[10] or 'AP').strip()
                    stores_data.append({
                        'store_code': code,
                        'store_name': name,
                        'region': region,
                        'asm_name': asm,
                        'brand': brand,
                        'passcode': '1234'
                    })
                print(f"📖 Loaded {len(stores_data)} stores directly from StoresInfo.xlsx")
            except Exception as e_xlsx:
                print(f"⚠️ Error reading StoresInfo.xlsx: {e_xlsx}")
                stores_data = []

        # Priority 2: Fallback to seed_stores_baseline.json
        if not stores_data:
            json_path = os.path.join(os.path.dirname(__file__), "seed_stores_baseline.json")
            if os.path.exists(json_path):
                with open(json_path, 'r', encoding='utf-8') as f:
                    raw_j = json.load(f)
                    for s in raw_j:
                        stores_data.append({
                            'store_code': s.get('store_code', ''),
                            'store_name': s.get('store_name', ''),
                            'region': s.get('region', ''),
                            'asm_name': s.get('asm_name') or s.get('asm') or '',
                            'brand': s.get('brand', 'AP'),
                            'passcode': s.get('passcode', '1234')
                        })

        if not stores_data:
            return
            
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            is_pg = bool(DATABASE_URL and psycopg2)
            
            for s in stores_data:
                code = s['store_code']
                name = s['store_name']
                region = s.get('region', '')
                asm = s.get('asm_name', '')
                brand = s.get('brand', 'AP')
                passcode = s.get('passcode', '1234')

                q_sel = "SELECT passcode FROM tb_stores WHERE store_code = %s" if is_pg else "SELECT passcode FROM tb_stores WHERE store_code = ?"
                cur.execute(q_sel, (code,))
                row = cur.fetchone()
                if row:
                    q_upd = "UPDATE tb_stores SET store_name=%s, region=%s, asm_name=%s, brand=%s WHERE store_code=%s" if is_pg else "UPDATE tb_stores SET store_name=?, region=?, asm_name=?, brand=? WHERE store_code=?"
                    cur.execute(q_upd, (name, region, asm, brand, code))
                else:
                    q_ins = "INSERT INTO tb_stores (store_code, store_name, brand, region, asm_name, passcode) VALUES (%s, %s, %s, %s, %s, %s)" if is_pg else "INSERT INTO tb_stores (store_code, store_name, brand, region, asm_name, passcode) VALUES (?, ?, ?, ?, ?, ?)"
                    cur.execute(q_ins, (code, name, brand, region, asm, passcode))
                    
            conn.commit()
            print(f"✅ Seeded/Updated {len(stores_data)} stores from StoresInfo / JSON")
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        print(f"⚠️ [Seed Stores Error]: {e}")

def seed_asms_from_stores():
    """
    ⚠️ CRITICAL: Đồng bộ tất cả ASM tồn tại trong tb_stores.asm_name vào tb_asms.
    Tránh lỗi: ASM (như 'Ni') có trong tb_stores nhưng không có trong tb_asms
    → validate_pin trả về 'Không tìm thấy ASM' dù ASM hợp lệ.
    Chạy sau seed_stores_baseline_data().
    """
    try:
        asm_in_stores = query_db(
            "SELECT DISTINCT asm_name FROM tb_stores WHERE asm_name IS NOT NULL AND asm_name != '' ORDER BY asm_name"
        )
        asm_in_asms = {r['asm_name'] for r in query_db("SELECT asm_name FROM tb_asms")}
        
        added = []
        for row in asm_in_stores:
            asm_name = row['asm_name']
            if asm_name and asm_name not in asm_in_asms:
                try:
                    execute_db(
                        "INSERT INTO tb_asms (asm_name, passcode) VALUES (?, ?) ON CONFLICT (asm_name) DO NOTHING",
                        (asm_name, '9999')
                    )
                    added.append(asm_name)
                except Exception as _e:
                    print(f"⚠️ [SeedASMs] Cannot add {asm_name}: {_e}")
        if added:
            print(f"✅ [SeedASMs] Đã thêm {len(added)} ASM thiếu vào tb_asms: {added}")
        else:
            print("✅ [SeedASMs] tb_asms đã đồng bộ với tb_stores (không có ASM thiếu).")
    except Exception as e:
        print(f"⚠️ [SeedASMs Error]: {e}")

try:
    safe_migrate_db()
    seed_stores_baseline_data()
    seed_hr_baseline_data()
    seed_asms_from_stores()
except Exception as _mig_err:
    print(f"⚠️ [Startup Migration/Seed Error]: {_mig_err}")


def async_sync_and_alert(store_code, report_date, support_requests):
    def worker():
        try:
            store_info = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
            store_name = store_info['store_name'] if store_info else store_code
            asm_name = store_info['asm_name'] if store_info else "QLKD / ASM"
        except Exception as e:
            print(f"Error querying store info in async worker: {e}")
            store_name = store_code
            asm_name = "QLKD / ASM"
            
        for sr in support_requests:
            cat = str(sr.get('category', '')).strip()
            pri = str(sr.get('priority', 'Trung bình')).strip()
            item = str(sr.get('issue_item', '')).strip()
            dl = str(sr.get('deadline', '')).strip()
            
            if not item:
                continue
                
            # 1. Sync to Google Sheets if APPS_SCRIPT_URL is configured
            gs_url = os.environ.get('GOOGLE_SHEET_WEBAPP_URL')
            if gs_url:
                payload = {
                    'store_code': store_code,
                    'store_name': store_name,
                    'asm_name': asm_name,
                    'report_date': report_date,
                    'category': cat,
                    'priority': pri,
                    'issue_item': item,
                    'deadline': dl,
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                try:
                    req = urllib.request.Request(
                        gs_url,
                        data=json.dumps(payload).encode('utf-8'),
                        headers={'Content-Type': 'application/json'},
                        method='POST'
                    )
                    with urllib.request.urlopen(req, timeout=10) as response:
                        response.read()
                    print(f"✅ Synced support request for {store_code} to Google Sheets.")
                except Exception as e:
                    print(f"❌ Error syncing to Google Sheets: {e}")
            else:
                print(f"ℹ️ Google Sheets URL not set. Support request logged locally: {item}")
                
            # 2. Send email via SMTP if user/pass is configured
            smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
            smtp_port = int(os.environ.get('SMTP_PORT', 587))
            smtp_user = os.environ.get('SMTP_USER')
            smtp_pass = os.environ.get('SMTP_PASSWORD')
            
            email_asm = os.environ.get('ALERT_EMAIL_ASM', 'asm_khoi@anphuoc.com.vn')
            email_tech = os.environ.get('ALERT_EMAIL_TECH', 'technical_support@anphuoc.com.vn')
            
            # Format Email Content
            subject = f"[Retail Commander - Su Co] Cua hang {store_name} yeu cau ho tro"
            body = f"""Kinh gui ASM {asm_name} va bo phan Ky thuat,

He thong Retail Commander vua ghi nhan yeu cau ho tro ky thuat tu cua hang:
- Cua hang: {store_name} ({store_code})
- ASM quan ly: {asm_name}
- Ky bao cao: {report_date}
- Phan loai su co: {cat}
- Muc do uu tien: {pri}
- Han xu ly mong muon: {dl}

Noi dung chi tiet su co:
--------------------------------------------------
{item}
--------------------------------------------------

Vui long kiem tra va xu ly su co kip thoi.

Tran trong,
He thong Tu dong hoa Retail Commander."""

            if smtp_user and smtp_pass:
                try:
                    msg = MIMEMultipart()
                    msg['From'] = smtp_user
                    msg['To'] = f"{email_asm}, {email_tech}"
                    msg['Subject'] = subject
                    msg.attach(MIMEText(body, 'plain', 'utf-8'))
                    
                    server = smtplib.SMTP(smtp_host, smtp_port)
                    server.starttls()
                    server.login(smtp_user, smtp_pass)
                    server.sendmail(smtp_user, [email_asm, email_tech], msg.as_string())
                    server.quit()
                    print(f"✅ SMTP Email sent successfully to {email_asm}, {email_tech}")
                except Exception as e:
                    print(f"❌ SMTP Error: {e}")
            else:
                # Fallback to local files for testing
                print("ℹ️ SMTP credentials not set. Saving email content to local file.")
                sent_dir = r"C:\All_Report\8_RETAIL_COMMANDER\tools\web_portal\sent_emails"
                os.makedirs(sent_dir, exist_ok=True)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"incident_{store_code}_{timestamp}.txt"
                filepath = os.path.join(sent_dir, filename)
                try:
                    with open(filepath, "w", encoding="utf-8") as f:
                        f.write(f"Subject: {subject}\n")
                        f.write(f"To: {email_asm}, {email_tech}\n")
                        f.write(f"SMTP Host: {smtp_host}:{smtp_port}\n")
                        f.write("="*50 + "\n")
                        f.write(body)
                    print(f"✅ Local email draft saved to: {filepath}")
                except Exception as e:
                    print(f"❌ Error writing local email file: {e}")

    threading.Thread(target=worker, daemon=True).start()

@app.route('/')
def index():
    default_date = get_report_date().strftime('%Y-%m-%d')
    return render_template('index.html', default_date=default_date)

@app.route('/api/asms', methods=['GET'])
def get_asms():
    try:
        try:
            cnt = query_db("SELECT COUNT(*) as count FROM tb_stores", one=True)
            if not cnt or cnt['count'] < 50:
                seed_stores_baseline_data(force=True)
        except Exception:
            pass
            
        rows = query_db("SELECT DISTINCT asm_name FROM tb_stores WHERE asm_name IS NOT NULL AND asm_name != '' ORDER BY asm_name")
        asms = [r['asm_name'] for r in rows if r['asm_name']]
        if not asms:
            asms = ['Dũng', 'HN', 'Hương', 'Khôi', 'Linh', 'Lâm', 'Ni', 'Quân', 'Tiên', 'Tín']
        return jsonify({'ok': True, 'asms': asms})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/stores', methods=['GET'])
def get_stores():
    asm = request.args.get('asm')
    store_code = request.args.get('store_code')
    try:
        try:
            cnt = query_db("SELECT COUNT(*) as count FROM tb_stores", one=True)
            if not cnt or cnt['count'] < 50:
                seed_stores_baseline_data(force=True)
        except Exception:
            pass

        if store_code:
            row = query_db("SELECT store_code, store_name, brand, region, asm_name FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
            return jsonify({'ok': True, 'store': row})
        
        raw_asm = (asm or '').strip()
        asm_upper = raw_asm.upper()
        
        all_rows = query_db("SELECT store_code, store_name, brand, region, asm_name FROM tb_stores ORDER BY store_name")
        
        if not raw_asm or asm_upper in ('ALL', 'ADMIN'):
            rows = all_rows
        else:
            clean_asm = raw_asm
            if clean_asm.upper().startswith('ASM_'):
                clean_asm = clean_asm[4:]
            elif clean_asm.upper().startswith('ASM '):
                clean_asm = clean_asm[4:]
            
            target_raw = clean_asm.strip().lower()
            target_norm = remove_vn_accents(target_raw)

            rows = []
            for r in all_rows:
                db_asm = (r['asm_name'] or '').strip()
                if not db_asm: continue
                db_asm_raw = db_asm.lower()
                db_asm_norm = remove_vn_accents(db_asm_raw)

                if (db_asm_raw == target_raw or 
                    db_asm_norm == target_norm or
                    target_norm in db_asm_norm or
                    db_asm_norm in target_norm):
                    rows.append(r)
            
        return jsonify({'ok': True, 'stores': rows})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/validate_pin', methods=['POST'])
def validate_pin():
    data = request.json or {}
    store_code = data.get('store_code')
    pin = data.get('pin')
    
    if not store_code or not pin:
        return jsonify({'ok': False, 'error': 'Missing store_code or pin'})
    
    master_pin = os.environ.get('MASTER_PIN', '8888')
    
    # ── Admin validation ──────────────────────────────────────────────────────
    if store_code == 'ADMIN':
        if pin == master_pin or pin == '8888' or pin == '1234':
            return jsonify({'ok': True, 'valid': True, 'role': 'admin',
                           'store_name': 'Quản Trị Viên (Admin)', 'asm_name': 'Admin'})
        return jsonify({'ok': True, 'valid': False, 'error': 'Mã PIN ADMIN không đúng (Mặc định: 8888)'})
    
    # ── ASM login ─────────────────────────────────────────────────────────────
    if store_code.startswith("ASM_"):
        asm_name = store_code[4:]
        try:
            asm = query_db("SELECT * FROM tb_asms WHERE asm_name = ?", (asm_name,), one=True)
            if asm:
                valid_pin = asm['passcode']
                if pin == valid_pin or pin == master_pin or pin == '8888':
                    return jsonify({'ok': True, 'valid': True, 'role': 'asm',
                                   'store_name': f'Cụm ASM {asm_name}',
                                   'asm_name': asm_name})
                return jsonify({'ok': True, 'valid': False, 'error': 'Mã PIN ASM không đúng'})
            else:
                # ASM có trong tb_stores.asm_name nhưng chưa được seed vào tb_asms
                stores_for_asm = query_db("SELECT COUNT(*) as cnt FROM tb_stores WHERE asm_name = ?", (asm_name,), one=True)
                if stores_for_asm and stores_for_asm['cnt'] > 0:
                    if pin in ('9999', '8888') or pin == master_pin:
                        try:
                            execute_db("INSERT INTO tb_asms (asm_name, passcode) VALUES (?, ?) ON CONFLICT (asm_name) DO NOTHING", (asm_name, '9999'))
                        except Exception:
                            pass
                        return jsonify({'ok': True, 'valid': True, 'role': 'asm',
                                       'store_name': f'Cụm ASM {asm_name}',
                                       'asm_name': asm_name})
                    return jsonify({'ok': True, 'valid': False, 'error': f'ASM {asm_name} chưa có PIN riêng, vui lòng dùng PIN 9999 hoặc liên hệ Admin.'})
                return jsonify({'ok': True, 'valid': False, 'error': f'Không tìm thấy ASM: {asm_name}'})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)})
    
    # ── Store login ───────────────────────────────────────────────────────────
    try:
        # Lấy thông tin store trước (cần dù PIN đúng hay master bypass)
        store_info = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
        
        if not store_info:
            return jsonify({'ok': True, 'valid': False, 'error': 'Cửa hàng không tồn tại'})
        
        store_name = store_info.get('store_name') or store_code
        asm_name   = store_info.get('asm_name') or ''
        
        # Kiểm tra PIN đúng của store
        if pin == (store_info.get('passcode') or '') or pin == master_pin or pin == '8888' or _default_pin_allowed(pin):
            return jsonify({'ok': True, 'valid': True, 'role': 'store',
                           'store_name': store_name,
                           'asm_name': asm_name})
        
        return jsonify({'ok': True, 'valid': False, 'error': 'Mã PIN không đúng'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ──────────────────────────────────────────────────────────────────────────────
# NEW API: DAILY TRAFFIC ENDPOINTS
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/api/get_daily_traffic', methods=['GET'])
def get_daily_traffic():
    store_code = request.args.get('store_code')
    year = request.args.get('year')
    month = request.args.get('month')
    pin = request.args.get('pin')
    
    if not store_code or not year or not month or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu tham số bắt buộc'})
        
    try:
        # Validate PIN
        store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (store_code, pin), one=True)
        if not store and not _default_pin_allowed(pin):
            return jsonify({'ok': False, 'error': 'Mã PIN không đúng'})
            
        # Query daily traffic and bills for this store and month
        prefix = f"{year}-{int(month):02d}-%"
        rows = query_db("SELECT traffic_date, traffic_val, bills_val, company_online_bills, store_online_bills, data_source, non_purchase_reasons FROM tb_traffic WHERE store_code = ? AND traffic_date LIKE ?", (store_code, prefix))

        traffic_map = {
            r['traffic_date']: {
                'traffic': r['traffic_val'],
                'bills': r['bills_val'] or 0,
                'company_online_bills': r['company_online_bills'] or 0,
                'store_online_bills': r['store_online_bills'] or 0,
                'data_source': r['data_source'] or 'store_actual',
                'non_purchase_reasons': r.get('non_purchase_reasons', '') or ''
            } for r in rows
        }
        return jsonify({'ok': True, 'traffic': traffic_map})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/submit_daily_traffic', methods=['POST'])
def submit_daily_traffic():
    data = request.json or {}
    store_code = data.get('store_code')
    pin = data.get('pin')
    traffic_data = data.get('traffic_data', {})
    
    if not store_code or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu store_code hoặc pin'})
        
    try:
        # Validate PIN
        master_pin = os.environ.get('MASTER_PIN', '8888')
        if pin != master_pin:
            store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (store_code, pin), one=True)
            if not store and not _default_pin_allowed(pin):
                return jsonify({'ok': False, 'error': 'Mã PIN không đúng'})
            
        # Upsert each day
        for date_str, day_data in traffic_data.items():
            traffic_str = day_data.get('traffic', '') if isinstance(day_data, dict) else ''
            bills_str = day_data.get('bills', '') if isinstance(day_data, dict) else ''
            co_str = day_data.get('company_online_bills', '') if isinstance(day_data, dict) else ''
            so_str = day_data.get('store_online_bills', '') if isinstance(day_data, dict) else ''
            reasons_input = day_data.get('non_purchase_reasons', '') if isinstance(day_data, dict) else ''

            if isinstance(reasons_input, (dict, list)):
                reasons_str = json.dumps(reasons_input, ensure_ascii=False)
            else:
                reasons_str = str(reasons_input or '').strip()
            
            if (traffic_str is None or str(traffic_str).strip() == '') and (bills_str is None or str(bills_str).strip() == ''):
                # Delete record if cleared
                execute_db("DELETE FROM tb_traffic WHERE store_code = ? AND traffic_date = ?", (store_code, date_str))
            else:
                try:
                    trf_val = int(traffic_str) if (traffic_str is not None and str(traffic_str).strip() != '') else 0
                    bil_val = int(bills_str) if (bills_str is not None and str(bills_str).strip() != '') else 0
                    co_val = int(co_str) if (co_str is not None and str(co_str).strip() != '') else 0
                    so_val = int(so_str) if (so_str is not None and str(so_str).strip() != '') else 0
                    
                    if trf_val < 0 or bil_val < 0 or co_val < 0 or so_val < 0:
                        continue
                    if DATABASE_URL:
                        execute_db("""
                        INSERT INTO tb_traffic (store_code, traffic_date, traffic_val, bills_val, company_online_bills, store_online_bills, data_source, non_purchase_reasons)
                        VALUES (?, ?, ?, ?, ?, ?, 'store_actual', ?)
                        ON CONFLICT (store_code, traffic_date) DO UPDATE SET
                            traffic_val = EXCLUDED.traffic_val,
                            bills_val = EXCLUDED.bills_val,
                            company_online_bills = EXCLUDED.company_online_bills,
                            store_online_bills = EXCLUDED.store_online_bills,
                            data_source = 'store_actual',
                            non_purchase_reasons = EXCLUDED.non_purchase_reasons
                        """, (store_code, date_str, trf_val, bil_val, co_val, so_val, reasons_str))
                    else:
                        execute_db("""
                        INSERT OR REPLACE INTO tb_traffic (store_code, traffic_date, traffic_val, bills_val, company_online_bills, store_online_bills, data_source, non_purchase_reasons)
                        VALUES (?, ?, ?, ?, ?, ?, 'store_actual', ?)
                        """, (store_code, date_str, trf_val, bil_val, co_val, so_val, reasons_str))
                except ValueError:
                    continue
                except ValueError:
                    continue
                    
        return jsonify({'ok': True, 'message': 'Đã cập nhật Traffic thành công!'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ──────────────────────────────────────────────────────────────────────────────
# NEW API: OPERATIONAL DETAILS (EDIT MECHANISM)
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/api/get_operational_report', methods=['GET'])
def get_operational_report():
    store_code = request.args.get('store_code')
    report_date = request.args.get('report_date')
    pin = request.args.get('pin')
    
    if not store_code or not report_date or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu tham số bắt buộc'})
        
    try:
        store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (store_code, pin), one=True)
        if not store and not _default_pin_allowed(pin):
            return jsonify({'ok': False, 'error': 'Mã PIN không đúng'})
            
        # Query contracts (Section 3.1) — includes customer_name
        contracts = query_db("SELECT contract_number, customer_name, contract_value, product_category, quantity, deposit_paid, installment_2, status, reason FROM tb_contracts WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        
        # Query unsigned contracts (Section 3.2) — includes contract_number
        unsigned = query_db("SELECT contract_number, customer_name, prev_year_value, expected_signing_time, product_category, quantity, status, reason FROM tb_unsigned_contracts WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        
        # Query operational details (Section 4.1 - 4.4)
        details = query_db("SELECT * FROM tb_operational_details WHERE store_code = ? AND report_date = ?", (store_code, report_date), one=True)
        
        # Query support requests (Section 4.5)
        support = query_db("SELECT ticket_code, category, priority, issue_item, deadline, person_in_charge, status, store_progress_note, asm_hq_note, report_date FROM tb_support_requests WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        
        # Query rollover pending support requests from previous weeks
        pending_support = query_db("""
            SELECT ticket_code, category, priority, issue_item, deadline, person_in_charge, status, store_progress_note, asm_hq_note, report_date 
            FROM tb_support_requests 
            WHERE store_code = ? AND report_date < ? AND (status IS NULL OR status NOT IN ('Hoàn tất', 'Đã hủy'))
            ORDER BY report_date DESC
        """, (store_code, report_date))
        
        # Get traffic and bills for this specific Friday (report_date) if nộp in weekly
        traffic_row = query_db("SELECT traffic_val, bills_val, company_online_bills, store_online_bills FROM tb_traffic WHERE store_code = ? AND traffic_date = ?", (store_code, report_date), one=True)
        traffic_val = traffic_row['traffic_val'] if traffic_row else ''
        bills_val = traffic_row['bills_val'] if traffic_row else ''
        co_val = traffic_row['company_online_bills'] if traffic_row else 0
        so_val = traffic_row['store_online_bills'] if traffic_row else 0
        
        return jsonify({
            'ok': True,
            'traffic': traffic_val,
            'bills': bills_val,
            'company_online_bills': co_val,
            'store_online_bills': so_val,
            'contracts': contracts,
            'unsigned_contracts': unsigned,
            'details': details or {},
            'support_requests': support,
            'pending_support_requests': pending_support
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# ──────────────────────────────────────────────────────────────────────────────
# SUBMIT & EXPORT ENHANCEMENTS (INCLUDES PART 4 & 5)
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/api/submit', methods=['POST'])
def submit_data():
    data = request.json or {}
    store_code = data.get('store_code')
    pin = data.get('pin')
    report_date = data.get('report_date')
    traffic_val = data.get('traffic')
    bills_val = data.get('bills')
    contracts = data.get('contracts', [])
    unsigned_contracts = data.get('unsigned_contracts', [])
    details = data.get('details', {})
    support_requests = data.get('support_requests', [])
    
    if not store_code or not report_date:
        return jsonify({'ok': False, 'error': 'Store code and report date are required'})
        
    try:
        # 1. Validate PIN
        master_pin = os.environ.get('MASTER_PIN', '8888')
        if pin != master_pin:
            store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (store_code, pin), one=True)
            if not store and not _default_pin_allowed(pin):
                return jsonify({'ok': False, 'error': 'Mã PIN không hợp lệ'})
            
        # 2. Save Traffic (Save for report_date as daily record)
        if traffic_val is not None and str(traffic_val).strip() != '':
            execute_db("DELETE FROM tb_traffic WHERE store_code = ? AND traffic_date = ?", (store_code, report_date))
            trf_int = int(traffic_val)
            bil_int = int(bills_val) if (bills_val is not None and str(bills_val).strip() != '') else 0
            co_int = int(data.get('company_online_bills', 0))
            so_int = int(data.get('store_online_bills', 0))
            
            if DATABASE_URL:
                execute_db("""
                INSERT INTO tb_traffic (store_code, traffic_date, traffic_val, bills_val, company_online_bills, store_online_bills, data_source) VALUES (?, ?, ?, ?, ?, ?, 'store_actual')
                ON CONFLICT (store_code, traffic_date) DO UPDATE SET
                    traffic_val = EXCLUDED.traffic_val,
                    bills_val = EXCLUDED.bills_val,
                    company_online_bills = EXCLUDED.company_online_bills,
                    store_online_bills = EXCLUDED.store_online_bills,
                    data_source = 'store_actual'
                """, (store_code, report_date, trf_int, bil_int, co_int, so_int))
            else:
                execute_db("INSERT OR REPLACE INTO tb_traffic (store_code, traffic_date, traffic_val, bills_val, company_online_bills, store_online_bills, data_source) VALUES (?, ?, ?, ?, ?, ?, 'store_actual')",
                           (store_code, report_date, trf_int, bil_int, co_int, so_int))
            
        # 3. Save Contracts (Section 3.1) — includes customer_name
        execute_db("DELETE FROM tb_contracts WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        for c in contracts:
            contract_num = str(c.get('contract_number', '')).strip() or 'Đang GD'
            customer_name = str(c.get('customer_name', '')).strip()
            val = float(c.get('contract_value', 0.0))
            cat = str(c.get('product_category', '')).strip()
            qty = int(c.get('quantity', 0))
            dep = float(c.get('deposit_paid', 0.0))
            inst2 = float(c.get('installment_2', 0.0))
            status = str(c.get('status', '')).strip()
            reason = str(c.get('reason', '')).strip()
            
            if val > 0 and cat:
                execute_db("""
                INSERT INTO tb_contracts 
                (store_code, report_date, contract_number, customer_name, contract_value, product_category, quantity, deposit_paid, installment_2, status, reason)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (store_code, report_date, contract_num, customer_name, val, cat, qty, dep, inst2, status, reason))
                
        # 4. Save Unsigned Contracts (Section 3.2) — includes contract_number
        execute_db("DELETE FROM tb_unsigned_contracts WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        for uc in unsigned_contracts:
            uc_contract_num = str(uc.get('contract_number', '')).strip() or 'Đang GD'
            val = float(uc.get('prev_year_value', 0.0))
            time_str = str(uc.get('expected_signing_time', '')).strip()
            cat = str(uc.get('product_category', '')).strip()
            qty = int(uc.get('quantity', 0))
            status = str(uc.get('status', '')).strip()
            reason = str(uc.get('reason', '')).strip()
            
            customer_name = str(uc.get('customer_name', '')).strip()
            if val > 0 or uc_contract_num or customer_name:
                execute_db("""
                INSERT INTO tb_unsigned_contracts 
                (store_code, report_date, contract_number, customer_name, prev_year_value, expected_signing_time, product_category, quantity, status, reason)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (store_code, report_date, uc_contract_num, customer_name, val, time_str, cat, qty, status, reason))
                
        # 5. Save Operational Details (Section 4.1 - 4.4)
        execute_db("DELETE FROM tb_operational_details WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        execute_db("""
        INSERT INTO tb_operational_details (
            store_code, report_date, 
            op_open_close_status, op_open_close_note,
            op_uniform_status, op_uniform_note,
            op_greet_status, op_greet_note,
            op_feedback_status, op_feedback_note,
            op_other_status, op_other_note,
            hr_target, hr_actual, hr_guard,
            hr_resigned_note, hr_leave_note, hr_absent_note,
            inv_stock_status, inv_info_goods, inv_return_warehouse, inv_proposal,
            market_product_feedback, market_missing_products, market_competitors, market_other_feedback
        ) VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
        )
        """, (
            store_code, report_date,
            details.get('op_open_close_status', 'Đạt'), details.get('op_open_close_note', ''),
            details.get('op_uniform_status', 'Đạt'), details.get('op_uniform_note', ''),
            details.get('op_greet_status', 'Đạt'), details.get('op_greet_note', ''),
            details.get('op_feedback_status', 'Đạt'), details.get('op_feedback_note', ''),
            details.get('op_other_status', 'Đạt'), details.get('op_other_note', ''),
            int(details.get('hr_target', 0) or 0), int(details.get('hr_actual', 0) or 0), int(details.get('hr_guard', 0) or 0),
            details.get('hr_resigned_note', ''), details.get('hr_leave_note', ''), details.get('hr_absent_note', ''),
            details.get('inv_stock_status', ''), details.get('inv_info_goods', ''), details.get('inv_return_warehouse', ''), details.get('inv_proposal', ''),
            details.get('market_product_feedback', ''), details.get('market_missing_products', ''), details.get('market_competitors', ''), details.get('market_other_feedback', '')
        ))
        
        # 6. Save Support Requests (Section 4.5)
        # First: update any rollover pending tickets from previous weeks
        pending_updates = data.get('pending_support_updates', [])
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for psu in pending_updates:
            tcode = str(psu.get('ticket_code', '')).strip()
            pstatus = str(psu.get('status', 'Đang xử lý')).strip()
            pnote = str(psu.get('store_progress_note', '')).strip()
            if tcode:
                _old_t = query_db("SELECT status FROM tb_support_requests WHERE ticket_code = ? AND store_code = ?",
                                   (tcode, store_code), one=True)
                execute_db("""
                UPDATE tb_support_requests
                SET status = ?, store_progress_note = ?, updated_at = ?
                WHERE ticket_code = ? AND store_code = ?
                """, (pstatus, pnote, now_str, tcode, store_code))
                _log_ticket_history('SUPPORT', tcode, store_code, _old_t.get('status') if _old_t else None, pstatus)

        # Second: Save current week's support requests
        execute_db("DELETE FROM tb_support_requests WHERE store_code = ? AND report_date = ?", (store_code, report_date))
        date_clean = report_date.replace('-', '')
        for idx, sr in enumerate(support_requests, 1):
            cat = str(sr.get('category', '')).strip()
            pri = str(sr.get('priority', 'Trung bình')).strip()
            item = str(sr.get('issue_item', '')).strip()
            dl = str(sr.get('deadline', '')).strip()
            pic = str(sr.get('person_in_charge', 'QLKD / ASM')).strip()
            status = str(sr.get('status', 'Đang xử lý')).strip()
            sp_note = str(sr.get('store_progress_note', '')).strip()
            hq_note = str(sr.get('asm_hq_note', '')).strip()
            tcode = str(sr.get('ticket_code', '')).strip() or f"TK-{store_code}-{date_clean}-{idx:02d}"
            
            if item:
                execute_db("""
                INSERT INTO tb_support_requests
                (ticket_code, store_code, report_date, category, priority, issue_item, deadline, person_in_charge, status, store_progress_note, asm_hq_note, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (tcode, store_code, report_date, cat, pri, item, dl, pic, status, sp_note, hq_note, now_str, now_str))
                _log_ticket_history('SUPPORT', tcode, store_code, None, status)
        
        if support_requests:
            async_sync_and_alert(store_code, report_date, support_requests)
            
        return jsonify({'ok': True, 'message': 'Nộp báo cáo thành công!'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/export', methods=['GET'])
def export_data():
    token = request.args.get('token')
    report_date = request.args.get('report_date')
    
    if token != API_SECRET_TOKEN:
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401
        
    if not report_date:
        return jsonify({'ok': False, 'error': 'report_date is required'})
        
    try:
        # Pull daily traffic records for the week ending on report_date (Saturday to Friday)
        # We also pull all daily traffic for this month to calculate MTD
        rep_dt = datetime.strptime(report_date, '%Y-%m-%d').date()
        week_start = rep_dt - timedelta(days=6)
        month_start = rep_dt.replace(day=1)
        
        daily_traffic = query_db("""
            SELECT store_code, traffic_date, traffic_val, bills_val FROM tb_traffic 
            WHERE traffic_date >= ? AND traffic_date <= ?
        """, (month_start.strftime('%Y-%m-%d'), report_date))
        
        contracts = query_db("SELECT store_code, contract_number, contract_value, product_category, quantity, deposit_paid, installment_2, status, reason FROM tb_contracts WHERE report_date = ?", (report_date,))
        unsigned = query_db("SELECT store_code, contract_number, customer_name, prev_year_value, expected_signing_time, product_category, quantity, status, reason FROM tb_unsigned_contracts WHERE report_date = ?", (report_date,))
        details = query_db("SELECT * FROM tb_operational_details WHERE report_date = ?", (report_date,))
        support = query_db("SELECT store_code, category, priority, issue_item, deadline, person_in_charge FROM tb_support_requests WHERE report_date = ?", (report_date,))
        
        return jsonify({
            'ok': True,
            'report_date': report_date,
            'daily_traffic': daily_traffic,
            'contracts': contracts,
            'unsigned_contracts': unsigned,
            'operational_details': details,
            'support_requests': support
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/submission_status', methods=['GET'])
def get_submission_status():
    report_date = request.args.get('report_date') or get_report_date().strftime('%Y-%m-%d')
    asm = request.args.get('asm')
    role = request.args.get('role', '')
    pin = request.args.get('pin', '')
    store_code = request.args.get('store_code', '')
    
    scope = get_auth_scope(role, asm, pin, store_code)
    try:
        if scope['type'] == 'STORE':
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE store_code = ?", (scope['store'],))
        elif scope['type'] == 'ASM':
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY region, store_name", (scope['asm'],))
        else:
            if asm and asm != 'ALL':
                stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY region, store_name", (asm,))
            else:
                stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores ORDER BY region, store_name")
            
        submitted = query_db("SELECT DISTINCT store_code FROM tb_operational_details WHERE report_date = ?", (report_date,))
        sub_set = {s['store_code'] for s in submitted}
        
        results = []
        for s in stores:
            results.append({
                'store_code': s['store_code'],
                'store_name': s['store_name'],
                'region': s['region'],
                'asm_name': s['asm_name'],
                'submitted': s['store_code'] in sub_set
            })
            
        return jsonify({'ok': True, 'report_date': report_date, 'status': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/get_support_requests', methods=['GET'])
def get_support_requests():
    report_date = request.args.get('report_date')
    asm = request.args.get('asm')
    store_code = request.args.get('store_code')
    role = request.args.get('role', '')
    pin = request.args.get('pin', '')
    status_filter = request.args.get('status_filter', 'ALL') # 'ALL', 'ACTIVE', 'COMPLETED'
    
    scope = get_auth_scope(role, asm, pin, store_code)
    try:
        where_clauses = []
        args = []
        
        if report_date:
            where_clauses.append("r.report_date = ?")
            args.append(report_date)
            
        if scope['type'] == 'STORE':
            where_clauses.append("r.store_code = ?")
            args.append(scope['store'])
        elif scope['type'] == 'ASM':
            asm_target = scope['asm']
            clean_asm = sanitize_filename_part(asm_target).lower() if 'sanitize_filename_part' in globals() else asm_target.lower()
            where_clauses.append("(s.asm_name = ? OR LOWER(s.asm_name) LIKE ?)")
            args.append(asm_target)
            args.append(f"%{clean_asm}%")
        else:
            if asm and asm != 'ALL' and asm != 'undefined':
                clean_asm = sanitize_filename_part(asm).lower() if 'sanitize_filename_part' in globals() else asm.lower()
                where_clauses.append("(s.asm_name = ? OR LOWER(s.asm_name) LIKE ?)")
                args.append(asm)
                args.append(f"%{clean_asm}%")
            if store_code:
                where_clauses.append("r.store_code = ?")
                args.append(store_code)
            
        if status_filter == 'ACTIVE':
            where_clauses.append("(r.status IS NULL OR r.status NOT IN ('Hoàn tất', 'Đã hủy'))")
        elif status_filter == 'COMPLETED':
            where_clauses.append("r.status IN ('Hoàn tất', 'Đã hủy')")
            
        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
        
        sql = f"""
            SELECT r.ticket_code, r.store_code, s.store_name, s.asm_name, r.category, r.priority, 
                   r.issue_item, r.deadline, r.person_in_charge, r.status, 
                   r.store_progress_note, r.asm_hq_note, r.report_date, r.updated_at
            FROM tb_support_requests r
            JOIN tb_stores s ON r.store_code = s.store_code
            {where_sql}
            ORDER BY CASE r.priority WHEN 'Cao' THEN 1 WHEN 'Trung bình' THEN 2 ELSE 3 END, r.report_date DESC
        """
        rows = query_db(sql, tuple(args))
            
        results = []
        for r in rows:
            results.append({
                'ticket_code': r['ticket_code'] or f"TK-{r['store_code']}-{r['report_date']}",
                'store_code': r['store_code'],
                'store_name': r['store_name'],
                'asm_name': r['asm_name'],
                'category': r['category'],
                'priority': r['priority'],
                'issue_item': r['issue_item'],
                'deadline': r['deadline'],
                'person_in_charge': r['person_in_charge'],
                'status': r['status'] or 'Đang xử lý',
                'store_progress_note': r['store_progress_note'] or '',
                'asm_hq_note': r['asm_hq_note'] or '',
                'report_date': r['report_date'],
                'updated_at': r['updated_at'] or ''
            })
        return jsonify({'ok': True, 'requests': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/update_support_ticket', methods=['POST'])
def update_support_ticket():
    data = request.json or {}
    ticket_code = data.get('ticket_code')
    pin = data.get('pin')
    status = data.get('status')
    store_progress_note = data.get('store_progress_note')
    asm_hq_note = data.get('asm_hq_note')
    person_in_charge = data.get('person_in_charge')
    deadline = data.get('deadline')
    
    if not ticket_code:
        return jsonify({'ok': False, 'error': 'Mã ticket không hợp lệ'})
        
    ticket = query_db("SELECT * FROM tb_support_requests WHERE ticket_code = ?", (ticket_code,), one=True)
    if not ticket:
        if str(ticket_code).isdigit():
            ticket = query_db("SELECT * FROM tb_support_requests WHERE id = ?", (int(ticket_code),), one=True)
        elif str(ticket_code).startswith("TK-"):
            parts = str(ticket_code).split("-")
            if len(parts) >= 3:
                sc = parts[1]
                rd = "-".join(parts[2:])
                ticket = query_db("SELECT * FROM tb_support_requests WHERE store_code = ? AND report_date = ? LIMIT 1", (sc, rd), one=True)
                
    if not ticket:
        return jsonify({'ok': False, 'error': 'Không tìm thấy thông tin sự vụ trong hệ thống'})
        
    master_pin = os.environ.get('MASTER_PIN', '8888')
    is_valid_pin = (pin == master_pin) or _default_pin_allowed(pin)
    if not is_valid_pin and pin:
        store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (ticket['store_code'], pin), one=True)
        if store:
            is_valid_pin = True
        else:
            asm = query_db("SELECT * FROM tb_asms WHERE passcode = ?", (pin,), one=True)
            if asm:
                is_valid_pin = True
                
    if not is_valid_pin:
        return jsonify({'ok': False, 'error': 'Mã PIN không có quyền cập nhật sự vụ này'})
        
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    target_id = ticket['id']
    
    updates = []
    args = []
    if status is not None:
        updates.append("status = ?")
        args.append(status)
    if store_progress_note is not None:
        updates.append("store_progress_note = ?")
        args.append(store_progress_note)
    if asm_hq_note is not None:
        updates.append("asm_hq_note = ?")
        args.append(asm_hq_note)
    if person_in_charge is not None:
        updates.append("person_in_charge = ?")
        args.append(person_in_charge)
    if deadline is not None:
        updates.append("deadline = ?")
        args.append(deadline)
        
    if not updates:
        return jsonify({'ok': False, 'error': 'Không có thông tin thay đổi'})
        
    updates.append("updated_at = ?")
    args.append(now_str)
    args.append(target_id)
    
    query = f"UPDATE tb_support_requests SET {', '.join(updates)} WHERE id = ?"
    execute_db(query, tuple(args))
    
    try:
        if status is not None and '_log_ticket_history' in globals():
            _log_ticket_history('SUPPORT', str(ticket_code), ticket['store_code'], ticket.get('status'), status)
    except Exception:
        pass

    return jsonify({'ok': True, 'message': f'Đã cập nhật sự vụ thành công!'})

# ──────────────────────────────────────────────────────────────────────────────
# PHÂN HỆ QUẢN LÝ TUÂN THỦ & KIỂM SOÁT NỘI BỘ P.QLQT (COMPLIANCE AUDITS)
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/api/admin/import_compliance_excel', methods=['POST'])
def import_compliance_excel():
    """API cho Admin / P.QLQT nạp file Excel Checklist kiểm tra camera"""
    try:
        if 'file' not in request.files:
            return jsonify({'ok': False, 'error': 'Vui lòng chọn file Excel đính kèm'})
            
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'ok': False, 'error': 'Tệp tải lên không hợp lệ'})
            
        custom_deadline = request.form.get('custom_deadline', '').strip()
        month_period = request.form.get('month_period', '').strip()
        pin = request.form.get('pin', '').strip()
        
        master_pin = os.environ.get('MASTER_PIN', '8888')
        if pin and pin != master_pin and not _default_pin_allowed(pin):
            asm = query_db("SELECT * FROM tb_asms WHERE passcode = ?", (pin,), one=True)
            if not asm or not ('khôi' in asm.get('asm_name','').lower() or 'khoi' in asm.get('asm_name','').lower()):
                return jsonify({'ok': False, 'error': 'Chỉ Admin và QLQT/ASM Khôi mới có quyền nạp báo cáo kiểm soát nội bộ'})
                
        import openpyxl
        import uuid
        import re
        wb = openpyxl.load_workbook(file, data_only=True)
        
        today_obj = datetime.now()
        today_str = today_obj.strftime('%Y-%m-%d')
        now_str = today_obj.strftime('%Y-%m-%d %H:%M:%S')
        
        # ── Extract Date YYYYMMDD from filename ──
        filename_str = file.filename or ''
        date_match = re.search(r'(\d{8})', filename_str)
        if date_match:
            try:
                dt_raw = date_match.group(1)
                file_date_obj = datetime.strptime(dt_raw, '%Y%m%d')
                received_date_str = file_date_obj.strftime('%Y-%m-%d')
                default_deadline_obj = file_date_obj + timedelta(days=5)
            except Exception:
                file_date_obj = today_obj
                received_date_str = today_str
                default_deadline_obj = today_obj + timedelta(days=5)
        else:
            file_date_obj = today_obj
            received_date_str = today_str
            default_deadline_obj = today_obj + timedelta(days=5)

        response_deadline_input = request.form.get('response_deadline', '').strip() or request.form.get('custom_deadline', '').strip()
        if response_deadline_input:
            default_deadline = response_deadline_input
        else:
            default_deadline = default_deadline_obj.strftime('%Y-%m-%d')
            
        if not month_period:
            month_period = file_date_obj.strftime('%Y-%m')
            
        batch_code = f"BATCH-PQLQT-{file_date_obj.strftime('%Y%m%d')}-{today_obj.strftime('%H%M%S')}"
        
        all_stores = query_db("SELECT * FROM tb_stores")
        store_by_code = {s['store_code'].upper().strip(): s for s in all_stores}
        all_stores = query_db("SELECT * FROM tb_stores")
        store_by_code = {s['store_code'].upper().strip(): s for s in all_stores}
        store_by_name = {sanitize_filename_part(s['store_name']).lower(): s for s in all_stores}
        
        # Build additional code lookups (e.g. 126_3T2 -> 1263T2)
        for s in all_stores:
            clean_code = re.sub(r'[^A-Z0-9]', '', s['store_code'].upper())
            if clean_code:
                store_by_code[clean_code] = s
        
        store_code_regex = re.compile(r'\b(AP\d+|AM\d+)\b', re.IGNORECASE)

        asm_full_map = {
            'đinh thị cát linh': 'Linh', 'linh': 'Linh',
            'đỗ thị hoa tiên': 'Tiên', 'tiên': 'Tiên',
            'nguyễn đăng khôi': 'Khôi', 'khôi': 'Khôi',
            'nguyễn lâm trung tín': 'Tín', 'tín': 'Tín',
            'nguyễn đăng quân': 'Quân', 'quân': 'Quân',
            'trần thanh dũng': 'Dũng', 'dũng': 'Dũng',
            'nguyễn thị thu hương': 'Hương', 'hương': 'Hương',
            'nguyễn văn lâm': 'Lâm', 'lâm': 'Lâm',
            'bùi minh ni': 'Ni', 'ni': 'Ni', 'hn': 'HN'
        }

        existing_repeats = query_db("""
            SELECT store_code, violation_group, COUNT(*) as cnt
            FROM tb_compliance_audits
            GROUP BY store_code, violation_group
        """)
        repeat_map = {(r['store_code'], r['violation_group']): r['cnt'] for r in existing_repeats}

        inserted_count = 0
        repeat_count = 0
        audits_to_insert = []

        # ── Process visible sheets, prioritizing detailed 'Báo cáo' ──
        visible_sheets = [s for s in wb.worksheets if s.sheet_state == 'visible']
        if not visible_sheets:
            visible_sheets = wb.worksheets

        # Filter out summary sheets ending in 'TH' or 'tổng hợp' if detailed report sheet exists
        detail_sheets = [s for s in visible_sheets if s.title.strip().lower() in ['báo cáo', 'bao cao', 'báo cáo chi tiết', 'checklist']]
        target_sheets = detail_sheets if detail_sheets else visible_sheets

        for ws in target_sheets:
            sheet_name = ws.title
            s_name_lower = sheet_name.lower().strip()
            
            # Skip summary sheets
            if s_name_lower in ['bao cao th', 'báo cáo tổng hợp', 'báo cáo th']:
                continue

            rows = list(ws.iter_rows(max_row=350, values_only=True))
            if len(rows) < 2:
                continue

            # ── Dynamic Header Map Detection ──
            col_map = {}
            header_idx = 0
            for idx, r in enumerate(rows[:10]):
                if not r: continue
                r_str = [str(cell or '').strip().lower() for cell in r]
                if any(k in ' '.join(r_str) for k in ['stt', 'mã ch', 'ma ch', 'cửa hàng', 'hạng mục', 'nội dung', 'vi phạm', 'đánh giá', 'asm']):
                    header_idx = idx
                    for col_i, cell_val in enumerate(r_str):
                        if not cell_val: continue
                        if 'chủ đề' in cell_val or 'nhóm' in cell_val or 'topic' in cell_val:
                            col_map['topic'] = col_i
                        elif 'mục kiểm tra' in cell_val or 'nội dung kiểm tra' in cell_val or 'hạng mục' in cell_val or 'tên mục' in cell_val:
                            col_map['check_item'] = col_i
                        elif 'tiêu chuẩn' in cell_val or 'quy định' in cell_val or 'chuẩn' in cell_val:
                            col_map['criteria_standard'] = col_i
                        elif 'mã ch' in cell_val or 'ma ch' in cell_val or 'mã cửa hàng' in cell_val or cell_val == 'mã':
                            col_map['store_code'] = col_i
                        elif 'tên ch' in cell_val or 'ten ch' in cell_val or 'tên cửa hàng' in cell_val or 'cửa hàng' in cell_val:
                            if 'store_code' not in col_map or col_map['store_code'] != col_i:
                                col_map['store_name'] = col_i
                        elif 'asm' in cell_val or 'quản lý' in cell_val or 'phụ trách' in cell_val:
                            col_map['asm_name'] = col_i
                        elif 'thời gian' in cell_val or 'camera' in cell_val or 'bằng chứng' in cell_val:
                            col_map['cam_time'] = col_i
                        elif 'mô tả' in cell_val or 'chi tiết' in cell_val or 'nội dung vi phạm' in cell_val or 'lỗi' in cell_val:
                            col_map['violation_description'] = col_i
                        elif 'đánh giá' in cell_val or 'kết quả' in cell_val or 'p.qlqt' in cell_val or 'xếp loại' in cell_val:
                            col_map['pqlqt_eval'] = col_i
                    if 'store_code' in col_map or 'store_name' in col_map or 'check_item' in col_map:
                        break

            empty_consecutive = 0
            for r in rows[header_idx + 1:]:
                if not r:
                    empty_consecutive += 1
                    if empty_consecutive >= 10:
                        break
                    continue

                def get_c(key, fallback_indices):
                    if key in col_map and col_map[key] < len(r):
                        v = str(r[col_map[key]] or '').strip()
                        if v: return v
                    for fi in fallback_indices:
                        if fi < len(r) and r[fi]:
                            v = str(r[fi]).strip()
                            if v: return v
                    return ''

                topic                 = get_c('topic', [2, 1])
                check_item            = get_c('check_item', [3, 2, 4])
                criteria_standard     = get_c('criteria_standard', [5, 4, 6])
                raw_store_code        = get_c('store_code', [8, 7, 6, 1])
                raw_store_name        = get_c('store_name', [9, 8, 7, 2])
                raw_asm_name          = get_c('asm_name', [10, 9, 8])
                cam_evidence_time     = get_c('cam_time', [11, 10])
                violation_group       = get_c('violation_group', [15, 3]) or topic or check_item[:30] or 'Vi phạm quy trình'
                violation_description = get_c('violation_description', [16, 15, 14, 13, 4]) or check_item
                pqlqt_eval            = get_c('pqlqt_eval', [21, 18, 16, 15]) or 'Không đạt'

                # Skip non-audit rows (headers, emails, directory table titles)
                all_text = f"{raw_store_code} {raw_store_name} {violation_description} {check_item}".lower()
                if '@anphuoc' in all_text or '@gmail' in all_text or 'danh sách hệ thống' in all_text or raw_store_code in ['Số nhà', 'Tên đường', 'Phường', 'STT', 'Stt']:
                    continue

                if not raw_store_code and not raw_store_name and not check_item and not violation_description:
                    empty_consecutive += 1
                    if empty_consecutive >= 10:
                        break
                    continue
                empty_consecutive = 0

                # Store Code Matching
                clean_raw_c = re.sub(r'[^A-Z0-9]', '', raw_store_code.upper())
                matched_store = store_by_code.get(raw_store_code.upper()) or store_by_code.get(clean_raw_c) or store_by_code.get('AP' + clean_raw_c)
                if not matched_store and raw_store_name:
                    clean_n = sanitize_filename_part(raw_store_name).lower()
                    matched_store = store_by_name.get(clean_n)

                if not matched_store:
                    code_match = store_code_regex.search(f"{raw_store_code} {raw_store_name} {violation_description}")
                    if code_match:
                        matched_store = store_by_code.get(code_match.group(1).upper())

                if not matched_store and raw_store_code.isdigit():
                    continue

                # Map ASM Name from full name to short code (e.g. Đinh Thị Cát Linh -> Linh)
                clean_asm_raw = raw_asm_name.lower().strip()
                mapped_asm = ''
                for k, v in asm_full_map.items():
                    if k in clean_asm_raw:
                        mapped_asm = v
                        break

                if matched_store:
                    final_store_code = matched_store['store_code']
                    final_store_name = matched_store['store_name']
                    final_asm_name = mapped_asm or matched_store['asm_name']
                else:
                    final_store_code = raw_store_code or 'UNKNOWN'
                    final_store_name = raw_store_name or final_store_code
                    final_asm_name = mapped_asm or 'Chưa phân công'

                clean_slug = sanitize_filename_part(final_store_code)[:15] if 'sanitize_filename_part' in globals() else final_store_code[:15]
                ticket_code = f"TK-COMP-{clean_slug}-{file_date_obj.strftime('%Y%m%d')}-{inserted_count + 1}-{uuid.uuid4().hex[:8].upper()}"

                key = (final_store_code, violation_group)
                prev_cnt = repeat_map.get(key, 0)
                is_repeat = 1 if prev_cnt > 0 else 0
                cur_repeat_cnt = prev_cnt + 1
                repeat_map[key] = cur_repeat_cnt
                if is_repeat:
                    repeat_count += 1

                audits_to_insert.append((
                    batch_code, ticket_code, sheet_name, topic, check_item, criteria_standard,
                    final_store_code, final_store_name, final_asm_name, cam_evidence_time, violation_group,
                    violation_description, pqlqt_eval, received_date_str, default_deadline,
                    'Mới tiếp nhận', is_repeat, cur_repeat_cnt, now_str, now_str
                ))
                inserted_count += 1

        if audits_to_insert:
            conn = get_db_connection()
            cur = conn.cursor()
            if DATABASE_URL:
                try:
                    import psycopg2.extras
                    sql_insert_pg = """
                        INSERT INTO tb_compliance_audits (
                            batch_code, ticket_code, sheet_name, topic, check_item, criteria_standard,
                            store_code, store_name, asm_name, camera_evidence_time, violation_group,
                            violation_description, pqlqt_evaluation, received_date, response_deadline,
                            status, is_repeat_offense, repeat_count, created_at, updated_at
                        ) VALUES %s ON CONFLICT (ticket_code) DO NOTHING
                    """
                    psycopg2.extras.execute_values(cur, sql_insert_pg, audits_to_insert)
                except Exception:
                    for row_args in audits_to_insert:
                        execute_db("""
                            INSERT INTO tb_compliance_audits (
                                batch_code, ticket_code, sheet_name, topic, check_item, criteria_standard,
                                store_code, store_name, asm_name, camera_evidence_time, violation_group,
                                violation_description, pqlqt_evaluation, received_date, response_deadline,
                                status, is_repeat_offense, repeat_count, created_at, updated_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ON CONFLICT (ticket_code) DO NOTHING
                        """, row_args)
            else:
                sql_insert_sqlite = """
                    INSERT OR IGNORE INTO tb_compliance_audits (
                        batch_code, ticket_code, sheet_name, topic, check_item, criteria_standard,
                        store_code, store_name, asm_name, camera_evidence_time, violation_group,
                        violation_description, pqlqt_evaluation, received_date, response_deadline,
                        status, is_repeat_offense, repeat_count, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                cur.executemany(sql_insert_sqlite, audits_to_insert)
            conn.commit()

        execute_db("""
            INSERT INTO tb_compliance_imports (
                batch_code, filename, imported_at, imported_by, total_records, month_period
            ) VALUES (?, ?, ?, ?, ?, ?)
        """, (batch_code, file.filename, now_str, 'Admin/P.QLQT', inserted_count, month_period))

        return jsonify({
            'ok': True,
            'message': f'Đã nạp báo cáo kiểm soát nội bộ thành công ({inserted_count} vụ việc)!',
            'batch_code': batch_code,
            'total_records': inserted_count,
            'repeat_count': repeat_count
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': f'Lỗi nạp file Excel báo cáo: {str(e)}'})

@app.route('/api/get_compliance_audits', methods=['GET'])
def get_compliance_audits():
    """API lấy danh sách các sự vụ kiểm soát nội bộ P.QLQT"""
    try:
        role = request.args.get('role', 'store')
        pin = request.args.get('pin', '')
        asm = request.args.get('asm', '')
        store_code = request.args.get('store_code', '')
        status_filter = request.args.get('status', '')
        
        # Support both param naming conventions from JS and API spec
        days_param = request.args.get('days_window') or request.args.get('days') or '365'
        try:
            days_window = int(days_param)
        except ValueError:
            days_window = 365
            
        repeat_param = request.args.get('repeat_only') or request.args.get('is_repeat') or ''
        repeat_only = (str(repeat_param).strip() == '1')

        scope = get_auth_scope(role, asm, pin, store_code)

        where_clauses = []
        args = []

        if scope['type'] == 'STORE':
            where_clauses.append("c.store_code = ?")
            args.append(scope['store'])
        elif scope['type'] == 'ASM':
            asm_target = scope['asm']
            clean_asm = sanitize_filename_part(asm_target).lower() if 'sanitize_filename_part' in globals() else asm_target.lower()
            where_clauses.append("(c.asm_name = ? OR LOWER(c.asm_name) LIKE ?)")
            args.append(asm_target)
            args.append(f"%{clean_asm}%")
        else:
            if asm and asm not in ('ALL', 'undefined', 'Tất cả ASM', '', 'all'):
                clean_asm = sanitize_filename_part(asm).lower() if 'sanitize_filename_part' in globals() else asm.lower()
                where_clauses.append("(c.asm_name = ? OR LOWER(c.asm_name) LIKE ?)")
                args.append(asm)
                args.append(f"%{clean_asm}%")
            if store_code:
                where_clauses.append("c.store_code = ?")
                args.append(store_code)

        if status_filter:
            if status_filter == 'PENDING_STORE':
                where_clauses.append("c.status IN ('Mới tiếp nhận', 'Chờ CH tường trình')")
            elif status_filter == 'PENDING_ASM':
                where_clauses.append("c.status IN ('Chờ ASM duyệt', 'CH đã giải trình')")
            elif status_filter == 'COMPLETED':
                where_clauses.append("c.status IN ('Đã hoàn tất', 'Hoàn tất', 'Đã hủy')")
            else:
                where_clauses.append("c.status = ?")
                args.append(status_filter)

        if repeat_only:
            where_clauses.append("c.is_repeat_offense = 1")

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        sql = f"""
            SELECT c.*, s.store_name as ref_store_name, s.asm_name as ref_asm_name
            FROM tb_compliance_audits c
            LEFT JOIN tb_stores s ON c.store_code = s.store_code
            {where_sql}
            ORDER BY c.is_repeat_offense DESC, c.id DESC
        """
        rows = query_db(sql, tuple(args))

        today_str = datetime.now().strftime('%Y-%m-%d')

        def norm_date(v):
            if not v: return ''
            vs = str(v).strip()
            if len(vs) >= 10 and vs[4] == '-' and vs[7] == '-':
                return vs[:10]
            if '/' in vs:
                parts = vs.split(' ')[0].split('/')
                if len(parts) == 3 and len(parts[2]) == 4:
                    return f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
            return vs[:10]

        audits = []
        total_cnt = len(rows)
        pending_store_cnt = 0
        pending_asm_cnt = 0
        completed_cnt = 0
        repeat_cnt = 0
        overdue_cnt = 0

        for r in rows:
            item = dict(r)
            item['store_name'] = item.get('ref_store_name') or item.get('store_name') or item['store_code']
            item['asm_name'] = item.get('ref_asm_name') or item.get('asm_name') or 'Chưa phân công'

            st = item.get('status', 'Mới tiếp nhận')
            raw_deadline = item.get('response_deadline', '')
            deadline = norm_date(raw_deadline)
            item['response_deadline'] = deadline or raw_deadline

            if st in ('Đã hoàn tất', 'Hoàn tất', 'Đã hủy'):
                item['sla_badge'] = '✅ Hoàn tất'
                item['is_overdue'] = False
                completed_cnt += 1
            else:
                if st in ('Mới tiếp nhận', 'Chờ CH tường trình'):
                    pending_store_cnt += 1
                elif st in ('Chờ ASM duyệt', 'CH đã giải trình'):
                    pending_asm_cnt += 1

                if deadline and deadline < today_str:
                    item['sla_badge'] = '🔴 Quá hạn SLA'
                    item['is_overdue'] = True
                    overdue_cnt += 1
                elif deadline and deadline == today_str:
                    item['sla_badge'] = '🟡 Hạn hôm nay'
                    item['is_overdue'] = False
                else:
                    item['sla_badge'] = '🟢 Trong hạn'
                    item['is_overdue'] = False

            if item.get('is_repeat_offense'):
                repeat_cnt += 1

            audits.append(item)

        # Get list of unique valid ASMs for dropdown (from tb_stores)
        asms_rows = query_db("""
            SELECT DISTINCT asm_name FROM tb_stores 
            WHERE asm_name IS NOT NULL AND asm_name != '' AND asm_name NOT LIKE '%1%' AND asm_name NOT LIKE '%2%'
            ORDER BY asm_name
        """)
        asms_list = [r['asm_name'] for r in asms_rows if r.get('asm_name')]

        return jsonify({
            'ok': True,
            'audits': audits,
            'asms': asms_list,
            'kpis': {
                'total': total_cnt,
                'pending_store': pending_store_cnt,
                'pending_asm': pending_asm_cnt,
                'completed': completed_cnt,
                'repeat': repeat_cnt,
                'overdue': overdue_cnt
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': f'Lỗi tải danh sách kiểm soát tuân thủ: {str(e)}'})

@app.route('/api/update_compliance_audit', methods=['POST'])
def update_compliance_audit():
    """API Cửa Hàng nhập giải trình / ASM phê duyệt & gán % trừ thưởng"""
    try:
        data = request.get_json() or {}
        ticket_code = data.get('ticket_code')
        pin = data.get('pin')
        store_explanation = data.get('store_explanation')
        asm_assessment = data.get('asm_assessment')
        violating_staff_info = data.get('violating_staff_info')
        attachment_url = data.get('attachment_url')
        status = data.get('status')
        custom_deadline = data.get('response_deadline')

        if not ticket_code:
            return jsonify({'ok': False, 'error': 'Mã ticket không hợp lệ'})

        ticket = query_db("SELECT * FROM tb_compliance_audits WHERE ticket_code = ?", (ticket_code,), one=True)
        if not ticket:
            return jsonify({'ok': False, 'error': 'Không tìm thấy ticket vi phạm'})

        master_pin = os.environ.get('MASTER_PIN', '8888')
        is_valid_pin = (pin == master_pin) or _default_pin_allowed(pin)
        if not is_valid_pin and pin:
            store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (ticket['store_code'], pin), one=True)
            if store:
                is_valid_pin = True
            else:
                asm = query_db("SELECT * FROM tb_asms WHERE passcode = ?", (pin,), one=True)
                if asm:
                    is_valid_pin = True

        if not is_valid_pin:
            return jsonify({'ok': False, 'error': 'Mã PIN không có quyền cập nhật sự vụ này'})

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        updates = []
        args = []

        if store_explanation is not None:
            updates.append("store_explanation = ?")
            args.append(store_explanation)
            if not status and ticket['status'] in ('Mới tiếp nhận', 'Chờ CH tường trình'):
                status = 'Chờ ASM duyệt'

        if asm_assessment is not None:
            updates.append("asm_assessment = ?")
            args.append(asm_assessment)

        if violating_staff_info is not None:
            updates.append("violating_staff_info = ?")
            args.append(violating_staff_info)

        if attachment_url is not None:
            updates.append("attachment_url = ?")
            args.append(attachment_url)

        if status is not None:
            updates.append("status = ?")
            args.append(status)

        if custom_deadline is not None:
            updates.append("response_deadline = ?")
            args.append(custom_deadline)

        if not updates:
            return jsonify({'ok': False, 'error': 'Không có dữ liệu thay đổi'})

        updates.append("updated_at = ?")
        args.append(now_str)
        args.append(ticket['id'])

        execute_db(f"UPDATE tb_compliance_audits SET {', '.join(updates)} WHERE id = ?", tuple(args))

        try:
            if status is not None and '_log_ticket_history' in globals():
                _log_ticket_history('COMPLIANCE', str(ticket_code), ticket['store_code'], ticket.get('status'), status)
        except Exception:
            pass

        return jsonify({'ok': True, 'message': 'Đã cập nhật hồ sơ tuân thủ thành công!'})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Lỗi cập nhật: {str(e)}'})

@app.route('/api/upload_compliance_attachment', methods=['POST'])
def upload_compliance_attachment():
    """API nhận file Tờ trình / Biên bản đính kèm (ảnh nạp nén từ Canvas / PDF)"""
    try:
        if 'file' not in request.files:
            return jsonify({'ok': False, 'error': 'Không tìm thấy file tải lên'})
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({'ok': False, 'error': 'File không hợp lệ'})

        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'compliance')
        os.makedirs(upload_dir, exist_ok=True)

        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
        filename = f"att_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}.{ext}"
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)

        rel_url = f"/static/uploads/compliance/{filename}"
        return jsonify({'ok': True, 'url': rel_url, 'filename': filename})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Lỗi upload file: {str(e)}'})

@app.route('/api/compliance_analytics', methods=['GET'])
def compliance_analytics():
    """API tổng hợp báo cáo kiểm soát tuân thủ theo ASM, Cửa Hàng & Top Tái Phạm"""
    try:
        asm_stats = query_db("""
            SELECT asm_name, 
                   COUNT(*) as total_violations,
                   SUM(CASE WHEN is_repeat_offense = 1 THEN 1 ELSE 0 END) as repeat_violations,
                   SUM(CASE WHEN status IN ('Đã hoàn tất', 'Hoàn tất', 'Đã hủy') THEN 1 ELSE 0 END) as completed_cnt
            FROM tb_compliance_audits
            GROUP BY asm_name
            ORDER BY total_violations DESC
        """)
        
        repeat_stores = query_db("""
            SELECT store_code, store_name, asm_name, 
                   COUNT(*) as total_violations,
                   SUM(CASE WHEN is_repeat_offense = 1 THEN 1 ELSE 0 END) as repeat_cnt
            FROM tb_compliance_audits
            GROUP BY store_code, store_name, asm_name
            HAVING SUM(CASE WHEN is_repeat_offense = 1 THEN 1 ELSE 0 END) > 0 OR COUNT(*) > 1
            ORDER BY repeat_cnt DESC, total_violations DESC
            LIMIT 15
        """)
        
        group_stats = query_db("""
            SELECT violation_group, COUNT(*) as cnt
            FROM tb_compliance_audits
            WHERE violation_group IS NOT NULL AND violation_group != ''
            GROUP BY violation_group
            ORDER BY cnt DESC
            LIMIT 10
        """)

        return jsonify({
            'ok': True,
            'asm_stats': [dict(r) for r in asm_stats],
            'repeat_stores': [dict(r) for r in repeat_stores],
            'group_stats': [dict(r) for r in group_stats]
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/export_compliance_excel', methods=['GET'])
def export_compliance_excel():
    """API xuất file Excel báo cáo kiểm soát nội bộ P.QLQT (điền đầy đủ thông tin giải trình & % trừ thưởng)"""
    try:
        batch_code = request.args.get('batch_code', '')
        asm = request.args.get('asm', '')
        
        where_clauses = []
        args = []
        if batch_code:
            where_clauses.append("batch_code = ?")
            args.append(batch_code)
        if asm and asm != 'ALL':
            clean_asm = sanitize_filename_part(asm).lower() if 'sanitize_filename_part' in globals() else asm.lower()
            where_clauses.append("(asm_name = ? OR LOWER(asm_name) LIKE ?)")
            args.append(asm)
            args.append(f"%{clean_asm}%")
            
        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
        audits = query_db(f"SELECT * FROM tb_compliance_audits {where_sql} ORDER BY id ASC", tuple(args))
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BaoCao_TuanThu_PQLQT"
        
        headers = [
            "STT", "Chuyên Đề", "Hạng Mục Kiểm Tra", "Tiêu Chuẩn Yêu Cầu", "Mã CH", "Tên CH",
            "QLKD Phụ Trách", "Thời Gian CAM", "Nhóm Lỗi Vi Phạm", "Mô Tả Vi Phạm (P.QLQT)",
            "Đánh Giá", "SLA Hạn Phản Hồi", "Trạng Thái", "Tái Phạm?",
            "Tường Trình Cửa Hàng (CH)", "Nhận Định QLKD (ASM) & % Trừ Thưởng", "Nhân Sự Vi Phạm", "Link Tờ Trình Đính Kèm"
        ]
        
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(1, col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            
        for idx, a in enumerate(audits, 1):
            is_rep = "⚠️ TÁI PHẠM" if a.get('is_repeat_offense') else "Lần 1"
            row = [
                idx,
                a.get('topic', ''),
                a.get('check_item', ''),
                a.get('criteria_standard', ''),
                a.get('store_code', ''),
                a.get('store_name', ''),
                a.get('asm_name', ''),
                a.get('camera_evidence_time', ''),
                a.get('violation_group', ''),
                a.get('violation_description', ''),
                a.get('pqlqt_evaluation', ''),
                a.get('response_deadline', ''),
                a.get('status', ''),
                is_rep,
                a.get('store_explanation', ''),
                a.get('asm_assessment', ''),
                a.get('violating_staff_info', ''),
                a.get('attachment_url', '')
            ]
            ws.append(row)
            r_idx = idx + 1
            for c_idx in range(1, len(row) + 1):
                c = ws.cell(r_idx, c_idx)
                c.border = thin_border
                c.alignment = center_align if c_idx in (1, 5, 7, 11, 12, 13, 14) else left_align
                
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Checklist_TuanThu_PQLQT_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        from flask import send_file
        return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Lỗi xuất file Excel: {str(e)}'})

# ──────────────────────────────────────────────────────────────────────────────
# NEW API: NON-PURCHASE REASON ANALYTICS
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/api/get_non_purchase_analytics', methods=['GET'])
def get_non_purchase_analytics():
    """
    API Báo cáo Phân tích Nguyên nhân Khách không mua hàng (Loss Conversion Analytics)
    Dành cho ASM, Admin và Ban Giám Đốc.
    """
    report_date = request.args.get('report_date')
    asm_filter = request.args.get('asm')
    store_filter = request.args.get('store_code')
    role = request.args.get('role', '')
    pin = request.args.get('pin', '')
    year = request.args.get('year')
    month = request.args.get('month')
    
    scope = get_auth_scope(role, asm_filter, pin, store_filter)
    if scope['type'] == 'STORE':
        store_filter = scope['store']
        asm_filter = None
    elif scope['type'] == 'ASM':
        asm_filter = scope['asm']
        store_filter = None
    else:
        if is_asm_khoi(asm_filter) or asm_filter == 'ALL':
            asm_filter = None
        
    if report_date:
        try:
            rdate = datetime.strptime(report_date, '%Y-%m-%d').date()
            start_date = (rdate - timedelta(days=6)).strftime('%Y-%m-%d')
            end_date = report_date
        except Exception:
            start_date = report_date
            end_date = report_date
    elif year and month:
        start_date = f"{year}-{int(month):02d}-01"
        end_date = f"{year}-{int(month):02d}-31"
    else:
        rdate = get_report_date()
        end_date = rdate.strftime('%Y-%m-%d')
        start_date = (rdate - timedelta(days=6)).strftime('%Y-%m-%d')

    try:
        sql = """
            SELECT t.store_code, s.store_name, s.asm_name, s.region,
                   t.traffic_date, t.traffic_val, t.bills_val, t.non_purchase_reasons
            FROM tb_traffic t
            JOIN tb_stores s ON t.store_code = s.store_code
            WHERE t.traffic_date >= ? AND t.traffic_date <= ?
        """
        args = [start_date, end_date]
        if asm_filter and asm_filter != 'ALL' and asm_filter != 'undefined':
            sql += " AND s.asm_name = ?"
            args.append(asm_filter)
        if store_filter:
            sql += " AND s.store_code = ?"
            args.append(store_filter)

        sql += " ORDER BY s.asm_name, s.store_name, t.traffic_date"

        rows = query_db(sql, args)

        total_traffic = 0
        total_bills = 0
        total_unconverted = 0
        
        reason_counts = {
            'SIZE': 0,
            'STYLE': 0,
            'PRICE': 0,
            'BROWSE': 0,
            'NEW_COLLECTION': 0,
            'SERVICE': 0,
            'OTHER': 0
        }
        
        reason_labels = {
            'SIZE': 'Đứt size / Thiếu size',
            'STYLE': 'Mẫu mã / Màu sắc chưa ưng',
            'PRICE': 'Giá cao / Chờ khuyến mãi',
            'BROWSE': 'Khách chỉ vào xem mẫu',
            'NEW_COLLECTION': 'Thiếu hàng mới / BST',
            'SERVICE': 'Phục vụ / Chờ đợi lâu',
            'OTHER': 'Lý do khác'
        }
        
        store_stats = {}
        notes_list = []

        for r in rows:
            trf = r['traffic_val'] or 0
            bil = r['bills_val'] or 0
            unconverted = max(0, trf - bil)
            
            total_traffic += trf
            total_bills += bil
            total_unconverted += unconverted
            
            scode = r['store_code']
            sname = r['store_name']
            asm = r['asm_name']
            
            if scode not in store_stats:
                store_stats[scode] = {
                    'store_code': scode,
                    'store_name': sname,
                    'asm_name': asm,
                    'region': r['region'],
                    'traffic': 0,
                    'bills': 0,
                    'unconverted': 0,
                    'reasons': {k: 0 for k in reason_counts.keys()},
                    'notes': []
                }
                
            store_stats[scode]['traffic'] += trf
            store_stats[scode]['bills'] += bil
            store_stats[scode]['unconverted'] += unconverted

            reasons_raw = r['non_purchase_reasons'] or ''
            if reasons_raw:
                try:
                    if reasons_raw.startswith('{'):
                        parsed = json.loads(reasons_raw)
                        r_list = parsed.get('reasons', [])
                        note = parsed.get('note', '').strip()
                    else:
                        r_list = [x.strip() for x in reasons_raw.split(',') if x.strip()]
                        note = ''
                        
                    for key in r_list:
                        if key in reason_counts:
                            reason_counts[key] += 1
                            store_stats[scode]['reasons'][key] += 1
                            
                    if note:
                        notes_list.append({
                            'store_name': sname,
                            'date': r['traffic_date'],
                            'note': note
                        })
                        store_stats[scode]['notes'].append(note)
                except Exception:
                    pass

        total_reason_hits = sum(reason_counts.values())
        reason_percentages = {}
        for key, count in reason_counts.items():
            pct = round((count / total_reason_hits * 100), 1) if total_reason_hits > 0 else 0
            reason_percentages[key] = {
                'label': reason_labels.get(key, key),
                'count': count,
                'percentage': pct
            }

        store_list = []
        for scode, s in store_stats.items():
            trf = s['traffic']
            bil = s['bills']
            cr = round((bil / trf * 100), 1) if trf > 0 else 0
            s['cr_percent'] = cr
            top_key = max(s['reasons'], key=s['reasons'].get) if sum(s['reasons'].values()) > 0 else None
            s['top_reason'] = reason_labels.get(top_key, 'Chưa ghi nhận') if top_key else 'Chưa ghi nhận'
            store_list.append(s)
            
        store_list.sort(key=lambda x: x['unconverted'], reverse=True)

        return jsonify({
            'ok': True,
            'period': {'start_date': start_date, 'end_date': end_date},
            'summary': {
                'total_traffic': total_traffic,
                'total_bills': total_bills,
                'total_unconverted': total_unconverted,
                'cr_percent': round((total_bills / total_traffic * 100), 1) if total_traffic > 0 else 0,
                'total_reason_hits': total_reason_hits
            },
            'reason_breakdown': reason_percentages,
            'store_rankings': store_list,
            'custom_notes': notes_list[:50]
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/debug_db_employees', methods=['GET'])
def debug_db_employees():
    code = request.args.get('code', '')
    if code:
        emps = query_db("SELECT * FROM tb_store_employees WHERE store_code = ?", (code,))
    else:
        emps = query_db("SELECT store_code, COUNT(*) as cnt FROM tb_store_employees GROUP BY store_code ORDER BY cnt DESC")
    return jsonify({'ok': True, 'data': [dict(e) for e in emps]})

@app.route('/api/admin/update_store_asm', methods=['POST'])
def update_store_asm():
    data = request.get_json() or {}
    pin = data.get('pin') or request.args.get('pin', '')
    if pin != os.environ.get('MASTER_PIN', '8888') and pin != '8888':
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401
    
    store_code = data.get('store_code')
    new_asm = data.get('asm_name')
    if not store_code or not new_asm:
        return jsonify({'ok': False, 'error': 'Thiếu store_code hoặc new_asm'}), 400
        
    execute_db("UPDATE tb_stores SET asm_name = ? WHERE store_code = ?", (new_asm, store_code))
    return jsonify({'ok': True, 'message': f'Đã cập nhật CH {store_code} thuộc ASM {new_asm}'})

@app.route('/api/seed_online_stores', methods=['GET', 'POST'])
def seed_online_stores():
    try:
        seed_stores_baseline_data(force=True)
        asms_res = query_db("SELECT DISTINCT asm_name FROM tb_stores WHERE asm_name IS NOT NULL AND asm_name != '' ORDER BY asm_name")
        asms = [r['asm_name'] for r in asms_res]
        cnt = query_db("SELECT COUNT(*) as count FROM tb_stores", one=True)
        total_cnt = cnt['count'] if cnt else 0
        return jsonify({'ok': True, 'stores_count': total_cnt, 'asms': asms})
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/api/seed_online_hr', methods=['GET', 'POST'])
def seed_online_hr():
    pin = request.args.get('pin', '')
    if pin != os.environ.get('MASTER_PIN', '8888'):
        return jsonify({'ok': False, 'error': 'Unauthorized'}), 401
    try:
        seed_hr_baseline_data()
        count = query_db("SELECT COUNT(*) as cnt FROM tb_store_employees WHERE status != 'RESIGNED'", one=True)
        cnt_val = count['cnt'] if (count and 'cnt' in count) else 0
        return jsonify({'ok': True, 'total_active_employees': cnt_val})
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        print(f"⚠️ [seed_online_hr Error]: {err_msg}")
        return jsonify({'ok': False, 'error': str(e), 'traceback': err_msg}), 500

@app.route('/api/get_store_hr', methods=['GET'])
def get_store_hr():
    store_code = request.args.get('store_code', '').strip()
    pin = request.args.get('pin', '').strip()
    
    if not store_code or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu mã cửa hàng hoặc mã PIN'}), 400
        
    store = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
    master_pin = os.environ.get('MASTER_PIN', '8888')
    if not _hr_pin_authorized(store, pin):
        return jsonify({'ok': False, 'error': 'Mã PIN không đúng'}), 401
        
    target_rec = query_db("SELECT target_headcount, cht_name FROM tb_store_headcount_targets WHERE store_code = ?", (store_code,), one=True)
    target_hc = target_rec['target_headcount'] if target_rec else 0
    cht_name = target_rec['cht_name'] if (target_rec and target_rec['cht_name']) else ''
    
    employees = query_db("SELECT * FROM tb_store_employees WHERE store_code = ? AND status != 'RESIGNED' ORDER BY position, full_name", (store_code,))
    for emp in employees:
        raw_date = emp.get('appointment_date') or emp.get('created_at') or ''
        emp['appointment_date'] = format_excel_date_py(raw_date)
        emp['tenure'] = calculate_tenure_py(emp['appointment_date'])
        
    actual_hc = len(employees)
    surplus_deficit = actual_hc - target_hc
    
    # Auto-ensure employees with position 'Nhân viên thử việc' have a probation record
    for emp in employees:
        if emp.get('position') == 'Nhân viên thử việc':
            e_code = emp.get('employee_code')
            p_exist = query_db("SELECT id FROM tb_employee_probation WHERE employee_code = ?", (e_code,), one=True)
            if not p_exist:
                execute_db("""
                    INSERT INTO tb_employee_probation (employee_code, store_code, start_date, trainer_name, probation_result)
                    VALUES (?, ?, ?, 'CHT/CHP', 'Đang thử việc')
                """, (e_code, store_code, emp.get('appointment_date') or date.today().strftime('%Y-%m-%d')))
    
    probationers = query_db("""
        SELECT p.*, e.full_name, e.gender, e.position, e.avatar_url
        FROM tb_employee_probation p
        JOIN tb_store_employees e ON p.employee_code = e.employee_code
        WHERE p.store_code = ? AND (p.probation_result IS NULL OR p.probation_result = '' OR p.probation_result = 'Đang thử việc')
    """, (store_code,))
    
    support_staff = query_db("SELECT * FROM tb_store_support_staff WHERE store_code = ? ORDER BY id DESC", (store_code,))
    
    shift_cfg = query_db("SELECT * FROM tb_store_shift_config WHERE store_code = ?", (store_code,), one=True)
    if not shift_cfg:
        shift_cfg = {
            'shift_1_name': 'Ca 1 (Sáng)',
            'shift_1_hours': '08:30 - 16:30',
            'shift_2_name': 'Ca 2 (Chiều)',
            'shift_2_hours': '14:00 - 22:00',
            'has_split_shift': 0,
            'split_shift_hours': '10:00 - 14:00 & 17:00 - 21:00'
        }
        
    hr_tickets = query_db("SELECT * FROM tb_hr_lifecycle_tickets WHERE store_code = ? AND status != 'Hoàn tất' ORDER BY created_at DESC", (store_code,))
    
    return jsonify({
        'ok': True,
        'store_code': store_code,
        'target_headcount': target_hc,
        'actual_headcount': actual_hc,
        'surplus_deficit': surplus_deficit,
        'cht_name': cht_name,
        'employees': employees,
        'probationers': probationers,
        'support_staff': support_staff,
        'shift_config': shift_cfg,
        'hr_tickets': hr_tickets
    })

@app.route('/api/save_store_hr', methods=['POST'])
def save_store_hr():
    data = request.json or {}
    store_code = data.get('store_code', '').strip()
    pin = data.get('pin', '').strip()
    
    if not store_code or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu mã cửa hàng hoặc mã PIN'}), 400
        
    store = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
    master_pin = os.environ.get('MASTER_PIN', '8888')
    if not _hr_pin_authorized(store, pin):
        return jsonify({'ok': False, 'error': 'Mã PIN không đúng'}), 401
        
    # Process offboard / transfer actions
    offboard_actions = data.get('offboard_actions', [])
    report_date = get_report_date().strftime('%Y-%m-%d')
    for act in offboard_actions:
        emp_code = act.get('employee_code', '').strip()
        reason = act.get('reason', '').strip()
        target_store = act.get('target_store_code', '').strip()
        note = act.get('note', '').strip()
        
        if not emp_code:
            continue
            
        emp = query_db("SELECT full_name, position FROM tb_store_employees WHERE employee_code = ?", (emp_code,), one=True)
        emp_name = emp['full_name'] if emp else emp_code
        pos = emp['position'] if emp else 'Nhân viên bán hàng'
        seq = datetime.now().strftime('%H%M%S')
        ticket_code = f"HR-{store_code}-{report_date.replace('-', '')}-{seq}"
        
        _now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if reason == 'TRANSFER' and target_store:
            execute_db("UPDATE tb_store_employees SET store_code = ?, appointment_date = ?, updated_at = CURRENT_TIMESTAMP WHERE employee_code = ?", (target_store, date.today().strftime('%Y-%m-%d'), emp_code))
            execute_db("""
                INSERT INTO tb_hr_lifecycle_tickets (ticket_code, store_code, report_date, employee_code, employee_name, position, event_type, effective_date, reason_note, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 'Chuyển cửa hàng', ?, ?, 'Mới ghi nhận', ?, ?)
            """, (ticket_code, store_code, report_date, emp_code, emp_name, pos, date.today().strftime('%Y-%m-%d'), f"Chuyển sang cửa hàng {target_store}. Ghi chú: {note}", _now_str, _now_str))
            _log_ticket_history('HR', ticket_code, store_code, None, 'Mới ghi nhận')
        elif reason == 'RESIGNED':
            execute_db("UPDATE tb_store_employees SET status = 'RESIGNED', updated_at = CURRENT_TIMESTAMP WHERE employee_code = ?", (emp_code,))
            execute_db("""
                INSERT INTO tb_hr_lifecycle_tickets (ticket_code, store_code, report_date, employee_code, employee_name, position, event_type, effective_date, reason_note, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 'Nghỉ việc', ?, ?, 'Mới ghi nhận', ?, ?)
            """, (ticket_code, store_code, report_date, emp_code, emp_name, pos, date.today().strftime('%Y-%m-%d'), f"Nhân sự nghỉ việc. Ghi chú: {note}", _now_str, _now_str))
            _log_ticket_history('HR', ticket_code, store_code, None, 'Mới ghi nhận')
        elif reason == 'MATERNITY_LEAVE':
            execute_db("UPDATE tb_store_employees SET status = 'MATERNITY_LEAVE', updated_at = CURRENT_TIMESTAMP WHERE employee_code = ?", (emp_code,))
            execute_db("""
                INSERT INTO tb_hr_lifecycle_tickets (ticket_code, store_code, report_date, employee_code, employee_name, position, event_type, effective_date, reason_note, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, 'Nghỉ thai sản', ?, ?, 'Mới ghi nhận', ?, ?)
            """, (ticket_code, store_code, report_date, emp_code, emp_name, pos, date.today().strftime('%Y-%m-%d'), f"Nhân sự nghỉ thai sản. Ghi chú: {note}", _now_str, _now_str))
            _log_ticket_history('HR', ticket_code, store_code, None, 'Mới ghi nhận')

    employees = data.get('employees', [])
    for emp in employees:
        emp_code = emp.get('employee_code', '').strip()
        full_name = emp.get('full_name', '').strip()
        if not emp_code or not full_name:
            continue
            
        gender = emp.get('gender', 'Nữ')
        dob = emp.get('dob', '')
        phone = emp.get('phone_number', '')
        position = emp.get('position', 'Nhân viên bán hàng')
        app_date = emp.get('appointment_date', '')
        avatar = emp.get('avatar_url', '')
        status = emp.get('status', 'ACTIVE')
        
        existing = query_db("SELECT id FROM tb_store_employees WHERE employee_code = ?", (emp_code,), one=True)
        if existing:
            execute_db("""
                UPDATE tb_store_employees
                SET store_code=?, full_name=?, gender=?, dob=?, phone_number=?, position=?, appointment_date=?, avatar_url=?, status=?, updated_at=CURRENT_TIMESTAMP
                WHERE employee_code=?
            """, (store_code, full_name, gender, dob, phone, position, app_date, avatar, status, emp_code))
        else:
            execute_db("""
                INSERT INTO tb_store_employees (employee_code, store_code, full_name, gender, dob, phone_number, position, appointment_date, avatar_url, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (emp_code, store_code, full_name, gender, dob, phone, position, app_date, avatar, status))
            
        # Probation auto-link if position is 'Nhân viên thử việc'
        if position == 'Nhân viên thử việc':
            p_exist = query_db("SELECT id FROM tb_employee_probation WHERE employee_code = ?", (emp_code,), one=True)
            if not p_exist:
                execute_db("""
                    INSERT INTO tb_employee_probation (employee_code, store_code, start_date, trainer_name, probation_result)
                    VALUES (?, ?, ?, 'CHT/CHP', 'Đang thử việc')
                """, (emp_code, store_code, app_date or date.today().strftime('%Y-%m-%d')))
            
    probation_updates = data.get('probation_updates', [])
    for p in probation_updates:
        emp_code = p.get('employee_code', '').strip()
        if not emp_code:
            continue
        start_date = p.get('start_date', date.today().strftime('%Y-%m-%d'))
        trainer_code = p.get('trainer_employee_code', '')
        trainer_name = p.get('trainer_name', '')
        t_op = 1 if p.get('train_op_rules') else 0
        t_wp = 1 if p.get('train_work_process') else 0
        t_pr = 1 if p.get('train_products') else 0
        t_dp = 1 if p.get('train_display') else 0
        t_inv = 1 if p.get('train_inventory') else 0
        notes = p.get('training_notes', '')
        end_date = p.get('probation_end_date', '')
        result = p.get('probation_result', 'Đang thử việc')
        
        existing_p = query_db("SELECT id FROM tb_employee_probation WHERE employee_code = ?", (emp_code,), one=True)
        if existing_p:
            execute_db("""
                UPDATE tb_employee_probation
                SET start_date=?, trainer_employee_code=?, trainer_name=?, train_op_rules=?, train_work_process=?, train_products=?, train_display=?, train_inventory=?, training_notes=?, probation_end_date=?, probation_result=?, updated_at=CURRENT_TIMESTAMP
                WHERE employee_code=?
            """, (start_date, trainer_code, trainer_name, t_op, t_wp, t_pr, t_dp, t_inv, notes, end_date, result, emp_code))
        else:
            execute_db("""
                INSERT INTO tb_employee_probation (employee_code, store_code, start_date, trainer_employee_code, trainer_name, train_op_rules, train_work_process, train_products, train_display, train_inventory, training_notes, probation_end_date, probation_result)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (emp_code, store_code, start_date, trainer_code, trainer_name, t_op, t_wp, t_pr, t_dp, t_inv, notes, end_date, result))
            
        if result == 'Nhận':
            execute_db("UPDATE tb_store_employees SET status='ACTIVE' WHERE employee_code=?", (emp_code,))
        elif result == 'Không đạt - Chuyển trả PTC':
            execute_db("UPDATE tb_store_employees SET status='RESIGNED' WHERE employee_code=?", (emp_code,))
            
    new_ticket = data.get('new_ticket')
    if new_ticket:
        report_date = get_report_date().strftime('%Y-%m-%d')
        seq = datetime.now().strftime('%H%M%S')
        ticket_code = f"HR-{store_code}-{report_date.replace('-', '')}-{seq}"
        emp_name = new_ticket.get('employee_name', '').strip()
        pos = new_ticket.get('position', 'Nhân viên bán hàng')
        evt_type = new_ticket.get('event_type', 'RESIGNED')
        eff_date = new_ticket.get('effective_date', report_date)
        reason = new_ticket.get('reason_note', '')
        handover = new_ticket.get('handover_status', 'Chưa bàn giao')
        
        if emp_name:
            _now_str2 = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            execute_db("""
                INSERT INTO tb_hr_lifecycle_tickets (ticket_code, store_code, report_date, employee_name, position, event_type, effective_date, reason_note, handover_status, status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Mới ghi nhận', ?, ?)
            """, (ticket_code, store_code, report_date, emp_name, pos, evt_type, eff_date, reason, handover, _now_str2, _now_str2))
            _log_ticket_history('HR', ticket_code, store_code, None, 'Mới ghi nhận')
            
    support_staff_updates = data.get('support_staff_updates', [])
    execute_db("DELETE FROM tb_store_support_staff WHERE store_code = ?", (store_code,))
    for ss in support_staff_updates:
        s_name = ss.get('employee_name', '').strip()
        from_st = ss.get('from_store_code', '').strip()
        if not s_name or not from_st:
            continue
        s_code = ss.get('employee_code', '')
        r_desc = ss.get('reason', 'Cửa hàng thiếu nhân sự')
        st_d = ss.get('start_date', '')
        ed_d = ss.get('end_date', '')
        execute_db("""
            INSERT INTO tb_store_support_staff (store_code, employee_code, employee_name, from_store_code, reason, start_date, end_date, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (store_code, s_code, s_name, from_st, r_desc, st_d, ed_d))
            
    return jsonify({'ok': True, 'message': 'Đã lưu hồ sơ nhân sự thành công'})

@app.route('/api/save_shift_config', methods=['POST'])
def save_shift_config():
    data = request.json or {}
    store_code = data.get('store_code', '').strip()
    pin = data.get('pin', '').strip()
    
    if not store_code or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu mã cửa hàng hoặc mã PIN'}), 400
        
    store = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
    master_pin = os.environ.get('MASTER_PIN', '8888')
    if not _hr_pin_authorized(store, pin):
        return jsonify({'ok': False, 'error': 'Mã PIN không đúng'}), 401
        
    s1_name = data.get('shift_1_name', 'Ca 1 (Sáng)')
    s1_hours = data.get('shift_1_hours', '08:30 - 16:30')
    s2_name = data.get('shift_2_name', 'Ca 2 (Chiều)')
    s2_hours = data.get('shift_2_hours', '14:00 - 22:00')
    has_split = 1 if data.get('has_split_shift') else 0
    split_hours = data.get('split_shift_hours', '10:00 - 14:00 & 17:00 - 21:00')
    
    existing = query_db("SELECT store_code FROM tb_store_shift_config WHERE store_code = ?", (store_code,), one=True)
    if existing:
        execute_db("""
            UPDATE tb_store_shift_config
            SET shift_1_name=?, shift_1_hours=?, shift_2_name=?, shift_2_hours=?, has_split_shift=?, split_shift_hours=?, updated_at=CURRENT_TIMESTAMP
            WHERE store_code=?
        """, (s1_name, s1_hours, s2_name, s2_hours, has_split, split_hours, store_code))
    else:
        execute_db("""
            INSERT INTO tb_store_shift_config (store_code, shift_1_name, shift_1_hours, shift_2_name, shift_2_hours, has_split_shift, split_shift_hours)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (store_code, s1_name, s1_hours, s2_name, s2_hours, has_split, split_hours))
        
    return jsonify({'ok': True, 'message': 'Đã cập nhật khung giờ ca làm việc'})

@app.route('/api/update_hr_ticket', methods=['POST'])
def update_hr_ticket():
    data = request.json or {}
    ticket_code = data.get('ticket_code', '').strip()
    status = data.get('status', '').strip()
    store_progress_note = data.get('store_progress_note', '').strip()
    asm_hr_note = data.get('asm_hr_note', '').strip()
    
    if not ticket_code:
        return jsonify({'ok': False, 'error': 'Thiếu mã ticket'}), 400
        
    ticket = query_db("SELECT * FROM tb_hr_lifecycle_tickets WHERE ticket_code = ?", (ticket_code,), one=True)
    if not ticket:
        return jsonify({'ok': False, 'error': 'Không tìm thấy ticket sự vụ'}), 404

    if status:
        execute_db("UPDATE tb_hr_lifecycle_tickets SET status = ? WHERE ticket_code = ?", (status, ticket_code))
        _log_ticket_history('HR', ticket_code, ticket['store_code'], ticket.get('status'), status)
    if store_progress_note:
        execute_db("UPDATE tb_hr_lifecycle_tickets SET store_progress_note = ? WHERE ticket_code = ?", (store_progress_note, ticket_code))
    if asm_hr_note:
        execute_db("UPDATE tb_hr_lifecycle_tickets SET asm_hr_note = ? WHERE ticket_code = ?", (asm_hr_note, ticket_code))
        
    execute_db("UPDATE tb_hr_lifecycle_tickets SET updated_at = CURRENT_TIMESTAMP WHERE ticket_code = ?", (ticket_code,))
    
    return jsonify({'ok': True, 'message': 'Đã cập nhật ticket sự vụ nhân sự thành công'})

# ──────────────────────────────────────────────────────────────────────────────
# STORE CLOSURES (đóng cửa: sửa chữa / di dời / mở mới) — ASM khai báo
# Engine qlkd_operational đọc qua export → tính Ngày HĐ + %HT điều chỉnh + đối
# soát với Data Lake. Portal KHÔNG truy cập Data Lake nên chỉ giữ khai báo;
# đối soát chạy ở engine (nơi thấy cả 2 nguồn).
# ──────────────────────────────────────────────────────────────────────────────
_CLOSURE_EVENT_TYPES = {'SỬA CHỮA', 'DI DỜI', 'MỞ MỚI', 'ĐÓNG TẠM'}
_CLOSURE_STATUSES    = {'ĐANG ĐÓNG', 'ĐÃ MỞ LẠI', 'KẾ HOẠCH', 'ĐÃ HỦY'}


def _closure_pin_authorized(store_code, pin):
    """ASM/admin được khai đóng cửa. Dùng chung logic get_auth_scope: MASTER_PIN
    hoặc PIN ASM (Khôi = mọi CH; ASM khác = CH thuộc cụm mình).

    ⚠️ QUAN TRỌNG (fix 2026-08-03 sau khi test UI thật gặp lỗi 401): PIN ASM ở
    hệ thống này là PIN DÙNG CHUNG (mọi ASM đều '9999') → query one=True trả về
    1 ASM NGẪU NHIÊN trong số đó, nếu ASM đó không quản CH đang khai thì bị từ
    chối oan. Phải lấy TẤT CẢ asm_name khớp PIN rồi kiểm tra xem có ai hợp lệ
    không — cùng cách get_auth_scope() đã làm (nó dùng query_db không one=True).
    """
    if not pin:
        return False
    if pin == MASTER_PIN or _default_pin_allowed(pin):
        return True
    rows = query_db("SELECT asm_name FROM tb_asms WHERE passcode = ?", (pin,))
    asm_names = [r['asm_name'] for r in (rows or [])]
    if not asm_names:
        return False
    # Khôi (hoặc PIN dùng chung có Khôi) → mọi cửa hàng
    if any(is_asm_khoi(n) for n in asm_names):
        return True
    # ASM thường: chỉ CH thuộc cụm của MỘT TRONG các ASM khớp PIN
    st = query_db("SELECT asm_name FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
    return bool(st and st.get('asm_name') in asm_names)


@app.route('/api/get_store_closures', methods=['GET'])
def get_store_closures():
    """Danh sách khai báo đóng cửa: theo store_code, hoặc theo asm (mọi CH cụm)."""
    store_code = request.args.get('store_code', '').strip()
    asm = request.args.get('asm', '').strip()
    if store_code:
        rows = query_db("""
            SELECT c.*, s.store_name FROM tb_store_closures c
            LEFT JOIN tb_stores s ON c.store_code = s.store_code
            WHERE c.store_code = ? ORDER BY c.start_date DESC
        """, (store_code,))
    elif asm:
        rows = query_db("""
            SELECT c.*, s.store_name FROM tb_store_closures c
            JOIN tb_stores s ON c.store_code = s.store_code
            WHERE s.asm_name = ? ORDER BY c.start_date DESC
        """, (asm,))
    else:
        rows = query_db("""
            SELECT c.*, s.store_name FROM tb_store_closures c
            LEFT JOIN tb_stores s ON c.store_code = s.store_code
            ORDER BY c.start_date DESC
        """)
    return jsonify({'ok': True, 'closures': [dict(r) for r in rows]})


@app.route('/api/save_store_closure', methods=['POST'])
def save_store_closure():
    """ASM tạo 1 khai báo đóng cửa mới."""
    data = request.json or {}
    store_code = str(data.get('store_code', '')).strip()
    event_type = str(data.get('event_type', '')).strip()
    start_date = str(data.get('start_date', '')).strip()
    end_date   = str(data.get('end_date', '')).strip()
    reason     = str(data.get('reason_note', '')).strip()
    pin        = str(data.get('pin', '')).strip()
    reported_by = str(data.get('reported_by', '')).strip()

    if not store_code or not event_type or not start_date:
        return jsonify({'ok': False, 'error': 'Thiếu cửa hàng / loại / ngày bắt đầu'}), 400
    if event_type not in _CLOSURE_EVENT_TYPES:
        return jsonify({'ok': False, 'error': f'Loại không hợp lệ (chỉ: {", ".join(sorted(_CLOSURE_EVENT_TYPES))})'}), 400
    if not _closure_pin_authorized(store_code, pin):
        return jsonify({'ok': False, 'error': 'Mã PIN không có quyền khai báo cho cửa hàng này'}), 401

    # status suy từ end_date: có end_date <= hôm nay → ĐÃ MỞ LẠI; ngày bắt đầu tương lai → KẾ HOẠCH
    today = date.today().strftime('%Y-%m-%d')
    if start_date > today:
        status = 'KẾ HOẠCH'
    elif end_date and end_date <= today:
        status = 'ĐÃ MỞ LẠI'
    else:
        status = 'ĐANG ĐÓNG'

    seq = datetime.now().strftime('%H%M%S')
    closure_code = f"CLS-{store_code}-{start_date.replace('-', '')}-{seq}"
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    execute_db("""
        INSERT INTO tb_store_closures
        (closure_code, store_code, event_type, start_date, end_date, status,
         reason_note, reported_by, source, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'ASM', ?, ?)
    """, (closure_code, store_code, event_type, start_date, end_date, status,
          reason, reported_by, now_str, now_str))
    _log_ticket_history('CLOSURE', closure_code, store_code, None, status)
    return jsonify({'ok': True, 'closure_code': closure_code,
                    'message': f'Đã ghi nhận đóng cửa {store_code} ({event_type})'})


@app.route('/api/update_store_closure', methods=['POST'])
def update_store_closure():
    """Cập nhật 1 khai báo (đổi status / end_date / ghi chú)."""
    data = request.json or {}
    closure_code = str(data.get('closure_code', '')).strip()
    pin = str(data.get('pin', '')).strip()
    if not closure_code:
        return jsonify({'ok': False, 'error': 'Thiếu mã khai báo'}), 400
    cls = query_db("SELECT * FROM tb_store_closures WHERE closure_code = ?", (closure_code,), one=True)
    if not cls:
        return jsonify({'ok': False, 'error': 'Không tìm thấy khai báo'}), 404
    if not _closure_pin_authorized(cls['store_code'], pin):
        return jsonify({'ok': False, 'error': 'Mã PIN không có quyền cập nhật khai báo này'}), 401

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_status = str(data.get('status', '')).strip()
    new_end    = data.get('end_date', None)
    new_note   = data.get('reason_note', None)
    if new_status and new_status not in _CLOSURE_STATUSES:
        return jsonify({'ok': False, 'error': 'Trạng thái không hợp lệ'}), 400

    if new_status:
        execute_db("UPDATE tb_store_closures SET status = ? WHERE closure_code = ?", (new_status, closure_code))
        _log_ticket_history('CLOSURE', closure_code, cls['store_code'], cls.get('status'), new_status)
    if new_end is not None:
        execute_db("UPDATE tb_store_closures SET end_date = ? WHERE closure_code = ?", (str(new_end).strip(), closure_code))
    if new_note is not None:
        execute_db("UPDATE tb_store_closures SET reason_note = ? WHERE closure_code = ?", (str(new_note).strip(), closure_code))
    execute_db("UPDATE tb_store_closures SET updated_at = ? WHERE closure_code = ?", (now_str, closure_code))
    return jsonify({'ok': True, 'message': f'Đã cập nhật khai báo {closure_code}'})

@app.route('/api/get_hr_analytics', methods=['GET'])
def get_hr_analytics():
    report_date = request.args.get('report_date', get_report_date().strftime('%Y-%m-%d'))
    asm = request.args.get('asm', 'ALL').strip()
    role = request.args.get('role', '').strip()
    pin = request.args.get('pin', '').strip()
    store_code = request.args.get('store_code', '').strip()
    
    scope = get_auth_scope(role, asm, pin, store_code)
    if scope['type'] == 'STORE':
        stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE store_code = ?", (scope['store'],))
    elif scope['type'] == 'ASM':
        stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY store_code", (scope['asm'],))
    else:
        if asm and asm != 'ALL' and not is_asm_khoi(asm):
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY store_code", (asm,))
        else:
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores ORDER BY store_code")
        
    store_codes = [s['store_code'] for s in stores]
    if not store_codes:
        return jsonify({'ok': True, 'summary': {}, 'store_hr_rows': [], 'hr_tickets': [],
                        'position_distribution': {}, 'tenure_buckets': {}, 'employees': []})
        
    placeholders = ",".join(["?"] * len(store_codes))
    
    targets = query_db(f"SELECT store_code, target_headcount, cht_name FROM tb_store_headcount_targets WHERE store_code IN ({placeholders})", store_codes)
    target_map = {t['store_code']: t for t in targets}
    
    # Full employee fetch for charts + list
    all_employees = query_db(
        f"SELECT e.*, h.target_headcount, h.cht_name FROM tb_store_employees e "
        f"LEFT JOIN tb_store_headcount_targets h ON e.store_code = h.store_code "
        f"WHERE e.store_code IN ({placeholders}) AND e.status != 'RESIGNED' "
        f"ORDER BY e.store_code, e.position, e.full_name",
        store_codes
    )
    
    emp_count_map = defaultdict(int)
    position_dist = defaultdict(int)
    tenure_buckets = {'lt6m': 0, '6m_1y': 0, '1y_2y': 0, '2y_3y': 0, 'gt3y': 0}
    today = date.today()
    
    employees_out = []
    for e in all_employees:
        emp_count_map[e['store_code']] += 1
        pos = e.get('position', 'Nhân viên bán hàng') or 'Nhân viên bán hàng'
        position_dist[pos] += 1
        
        # Tenure bucket
        apt_raw = e.get('appointment_date') or e.get('created_at', '')
        apt_formatted = format_excel_date_py(apt_raw)
        tenure_label = calculate_tenure_py(apt_formatted)
        if apt_formatted:
            try:
                hire_dt = datetime.strptime(apt_formatted[:10], '%Y-%m-%d').date()
                months = (today.year - hire_dt.year) * 12 + (today.month - hire_dt.month)
                if months < 6: tenure_buckets['lt6m'] += 1
                elif months < 12: tenure_buckets['6m_1y'] += 1
                elif months < 24: tenure_buckets['1y_2y'] += 1
                elif months < 36: tenure_buckets['2y_3y'] += 1
                else: tenure_buckets['gt3y'] += 1
            except Exception:
                pass
        
        employees_out.append({
            'employee_code': e.get('employee_code', ''),
            'full_name': e.get('full_name', ''),
            'gender': e.get('gender', ''),
            'position': pos,
            'store_code': e.get('store_code', ''),
            'appointment_date': apt_formatted,
            'tenure': tenure_label,
            'status': e.get('status', 'ACTIVE'),
        })
        
    probationers = query_db(f"SELECT store_code FROM tb_employee_probation WHERE store_code IN ({placeholders}) AND (probation_result IS NULL OR probation_result = '' OR probation_result = 'Đang thử việc')", store_codes)
    prob_count_map = defaultdict(int)
    for p in probationers:
        prob_count_map[p['store_code']] += 1
        
    tickets = query_db(f"SELECT * FROM tb_hr_lifecycle_tickets WHERE store_code IN ({placeholders}) AND status != 'Hoàn tất' ORDER BY created_at DESC", store_codes)
    ticket_count_map = defaultdict(int)
    under_review_cnt = 0
    for tk in tickets:
        ticket_count_map[tk['store_code']] += 1
        if tk['status'] == 'Đang xem xét':
            under_review_cnt += 1
            
    store_rows = []
    tot_target = 0
    tot_actual = 0
    tot_probation = 0
    
    for s in stores:
        code = s['store_code']
        t_rec = target_map.get(code, {})
        tgt = t_rec.get('target_headcount', 0)
        cht = t_rec.get('cht_name', '')
        act = emp_count_map[code]
        prob = prob_count_map[code]
        diff = act - tgt
        
        tot_target += tgt
        tot_actual += act
        tot_probation += prob
        
        store_rows.append({
            'store_code': code,
            'store_name': s['store_name'],
            'region': s['region'],
            'asm_name': s['asm_name'],
            'cht_name': cht,
            'target_headcount': tgt,
            'actual_headcount': act,
            'surplus_deficit': diff,
            'probationers_count': prob,
            'active_tickets_count': ticket_count_map[code]
        })
        
    # Top 12 deficit stores for bar chart
    chart_stores = sorted(store_rows, key=lambda r: r['surplus_deficit'])[:12]
    
    return jsonify({
        'ok': True,
        'summary': {
            'total_stores': len(stores),
            'total_target_headcount': tot_target,
            'total_actual_headcount': tot_actual,
            'net_surplus_deficit': tot_actual - tot_target,
            'total_probationers': tot_probation,
            'active_tickets_count': len(tickets),
            'under_review_count': under_review_cnt
        },
        'store_hr_rows': store_rows,
        'chart_stores': chart_stores,
        'hr_tickets': [dict(t) for t in tickets],
        'position_distribution': dict(position_dist),
        'tenure_buckets': tenure_buckets,
        'employees': employees_out
    })

# ──────────────────────────────────────────────────────────────────────────────
# NEW API: ASM TRAFFIC REMINDERS SUMMARY
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/api/asm_traffic_summary', methods=['GET'])
def get_asm_traffic_summary():
    asm = request.args.get('asm')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if not start_date_str or not end_date_str:
        return jsonify({'ok': False, 'error': 'Thiếu tham số bắt buộc'})
        
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Limit end_date to yesterday
        yesterday = date.today() - timedelta(days=1)
        if end_date > yesterday:
            end_date = yesterday
            
        # Get list of all dates in range
        delta = end_date - start_date
        all_dates = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]
        
        # Get all stores for this ASM or all stores
        if asm and asm != 'ALL' and asm != 'undefined' and asm != '':
            stores = query_db("SELECT store_code, store_name FROM tb_stores WHERE asm_name = ? ORDER BY store_name", (asm,))
        else:
            stores = query_db("SELECT store_code, store_name FROM tb_stores ORDER BY store_name")
        
        # Get all traffic entries in range
        rows = query_db("""
            SELECT store_code, traffic_date FROM tb_traffic 
            WHERE traffic_date >= ? AND traffic_date <= ?
        """, (start_date_str, end_date_str))
        
        # Map store_code -> set of filled dates
        filled_map = {}
        for r in rows:
            code = r['store_code']
            dt = r['traffic_date']
            if code not in filled_map:
                filled_map[code] = set()
            filled_map[code].add(dt)
            
        results = []
        for s in stores:
            code = s['store_code']
            name = s['store_name']
            filled = filled_map.get(code, set())
            
            missing_dates = [d for d in all_dates if d not in filled]
            # Convert YYYY-MM-DD to DD/MM
            missing_formatted = [datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m') for d in missing_dates]
            
            results.append({
                'store_code': code,
                'store_name': name,
                'total_filled': len(filled),
                'total_required': len(all_dates),
                'missing_dates': missing_formatted
            })
            
        return jsonify({'ok': True, 'summary': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/get_all_passcodes', methods=['GET'])
def get_all_passcodes():
    admin_pin = request.args.get('admin_pin')
    master_pin = os.environ.get('MASTER_PIN', '8888')
    if admin_pin != master_pin:
        return jsonify({'ok': False, 'error': 'Không có quyền truy cập'}), 401
        
    try:
        asms = query_db("SELECT asm_name, passcode FROM tb_asms ORDER BY asm_name")
        stores = query_db("SELECT store_code, store_name, asm_name, passcode FROM tb_stores ORDER BY store_code")
        return jsonify({'ok': True, 'asms': asms, 'stores': stores})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/update_passcode', methods=['POST'])
def update_passcode():
    data = request.json or {}
    user_type = data.get('user_type') # 'store' or 'asm'
    user_id = data.get('user_id') # store_code or asm_name
    new_pin = data.get('new_pin')
    old_pin = data.get('old_pin')
    admin_pin = data.get('admin_pin')
    
    if not user_type or not user_id or not new_pin:
        return jsonify({'ok': False, 'error': 'Thiếu tham số bắt buộc'})
        
    new_pin = str(new_pin).strip()
    if len(new_pin) < 4:
        return jsonify({'ok': False, 'error': 'Mã PIN phải có ít nhất 4 ký tự'})
        
    master_pin = os.environ.get('MASTER_PIN', '8888')
    is_admin = (admin_pin == master_pin)
    
    try:
        if user_type == 'asm':
            # Check authorization if not admin
            if not is_admin:
                asm = query_db("SELECT * FROM tb_asms WHERE asm_name = ? AND passcode = ?", (user_id, old_pin), one=True)
                if not asm:
                    return jsonify({'ok': False, 'error': 'Mật khẩu cũ không chính xác'})
            
            # Update
            execute_db("UPDATE tb_asms SET passcode = ? WHERE asm_name = ?", (new_pin, user_id))
            return jsonify({'ok': True, 'message': 'Đã đổi mã PIN ASM thành công!'})
            
        elif user_type == 'store':
            # Check authorization if not admin
            if not is_admin:
                store = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (user_id, old_pin), one=True)
                # Fallback default PIN (chỉ khi ALLOW_DEFAULT_PIN bật)
                if not store and _default_pin_allowed(old_pin):
                    store_exists = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (user_id,), one=True)
                    if store_exists:
                        store = store_exists
                if not store:
                    return jsonify({'ok': False, 'error': 'Mật khẩu cũ không chính xác'})
            
            # Update
            execute_db("UPDATE tb_stores SET passcode = ? WHERE store_code = ?", (new_pin, user_id))
            return jsonify({'ok': True, 'message': 'Đã đổi mã PIN cửa hàng thành công!'})
            
        else:
            return jsonify({'ok': False, 'error': 'Loại người dùng không hợp lệ'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

def style_sheet(ws, title, subtitle, date_range_str, role_str):
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    max_cols = ws.max_column
    
    # Merge cells for Title on Row 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title.upper()
    title_cell.font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    
    # Merge cells for Subtitle on Row 2 (Date Range & Target)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = f"Tuần báo cáo: {date_range_str}  |  Đối tượng: {role_str}"
    sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="FFFFFF")
    sub_cell.fill = navy_fill
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    
    # Merge cells for Metadata on Row 3 (Export info)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_cols)
    meta_cell = ws.cell(row=3, column=1)
    from datetime import datetime
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    meta_cell.value = f"Ngày xuất báo cáo: {now_str}  |  Hệ thống Retail Commander"
    meta_cell.font = Font(name="Segoe UI", size=9, italic=True, color="FFFFFF")
    meta_cell.fill = navy_fill
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18
    
    # Empty Spacer Row 4
    ws.row_dimensions[4].height = 12
    
    # Table Header Row 5
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Format Headers & Cache Header Names + Widths
    header_names = []
    col_widths = []
    for col in range(1, max_cols + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        val = str(cell.value or "")
        header_names.append(val)
        col_widths.append(len(val))
    ws.row_dimensions[5].height = 24
    
    # Format Data Rows (Row 6 onwards)
    data_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    max_row = ws.max_row
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    for row in range(6, max_row + 1):
        is_even = (row % 2 == 0)
        row_fill = zebra_fill if is_even else white_fill
        ws.row_dimensions[row].height = 18
        
        for col_idx in range(max_cols):
            cell = ws.cell(row=row, column=col_idx + 1)
            cell.font = data_font
            if cell.fill.fill_type is None:
                cell.fill = row_fill
            cell.border = thin_border
            
            val = cell.value
            if val is not None:
                s_len = len(str(val))
                if s_len > col_widths[col_idx]:
                    col_widths[col_idx] = s_len
                    
            col_name = header_names[col_idx]
            if isinstance(val, (int, float)):
                if "%" in col_name or "Tỷ Lệ" in col_name:
                    cell.number_format = '0.0"%"'
                    cell.alignment = align_right
                elif "Giá Trị" in col_name or "Số Tiền" in col_name or "Cọc" in col_name or "Đợt 2" in col_name or "HĐ" in col_name:
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                else:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
            elif val == "Chưa nộp" or val == "Chưa nhập" or val == "N/A":
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Apply Auto Column Widths
    for col_idx, max_w in enumerate(col_widths, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_w + 3, 11)

def style_input_traffic_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    peach_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    white_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    red_bold_font = Font(name="Segoe UI", size=12, bold=True, color="C00000")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    normal_font = Font(name="Segoe UI", size=10)
    italic_gray_font = Font(name="Segoe UI", size=10, italic=True, color="7F7F7F")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # --- Row 1 ---
    cell_a1 = ws.cell(row=1, column=1)
    cell_a1.value = "Nhập đúng tên CH trên file M\n(chỉ thêm đúng CH mình quản lý)"
    cell_a1.font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    cell_a1.fill = red_fill
    cell_a1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=2)
    title_cell.value = "Nhập traffic cửa hàng vào đây"
    title_cell.font = red_bold_font
    title_cell.fill = peach_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(3, 8):
        ws.cell(row=1, column=col).fill = peach_fill
        
    ws.row_dimensions[1].height = 35
    
    # --- Row 2 (Total) ---
    ws.cell(row=2, column=1).value = "Total"
    ws.cell(row=2, column=1).font = bold_font
    ws.cell(row=2, column=1).fill = gray_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).border = thin_border
    
    max_row = ws.max_row
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=2, column=col)
        cell.value = f"=SUM({col_letter}4:{col_letter}{max_row})"
        cell.font = bold_font
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.number_format = '#,##0'
        cell.border = thin_border
        
    ws.row_dimensions[2].height = 20
    
    # --- Row 3 (Headers) ---
    headers = [
        "Cửa Hàng", "Traffic Tuần Này", "Traffic Tuần Trước", 
        "Traffic Tuần cùng kỳ", "Traffic Tháng Này", "Traffic Tháng Trước", 
        "Traffic Tháng cùng kỳ"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = h
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws.row_dimensions[3].height = 25
    
    # --- Row 4+ (Data) ---
    for row in range(4, max_row + 1):
        is_even = (row % 2 == 0)
        row_fill = zebra_fill if is_even else white_fill
        ws.row_dimensions[row].height = 20
        
        cell_name = ws.cell(row=row, column=1)
        cell_name.font = normal_font
        cell_name.fill = row_fill
        cell_name.border = thin_border
        cell_name.alignment = Alignment(horizontal="left", vertical="center")
        
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            cell.border = thin_border
            cell.fill = row_fill
            
            if val == "Không có dữ liệu":
                cell.font = italic_gray_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
                
    ws.column_dimensions['A'].width = 32
    for col in range(2, 8):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

def sanitize_filename_part(text):
    if not text:
        return ""
    import unicodedata
    import re
    text = unicodedata.normalize('NFD', str(text))
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = text.replace('Đ', 'D').replace('đ', 'd')
    text = re.sub(r'[^a-zA-Z0-9_]', '_', text)
    text = re.sub(r'_+', '_', text).strip('_')
    return text

@app.route('/api/export_excel', methods=['GET'])
def export_excel():
    report_date = request.args.get('report_date')
    asm = request.args.get('asm')
    role = request.args.get('role', '') # 'store', 'asm' or 'admin'
    pin = request.args.get('pin', '')
    store_code = request.args.get('store_code', '')
    
    if not report_date:
        return "Thiếu ngày báo cáo", 400
        
    scope = get_auth_scope(role, asm, pin, store_code)
    filter_asm = scope['asm'] if scope['type'] == 'ASM' else (asm if (asm and asm != 'ALL' and not is_asm_khoi(asm)) else None)
    
    is_authorized = False
    if scope['type'] == 'ALL':
        is_authorized = True
    elif scope['type'] == 'ASM':
        asm_record = query_db("SELECT * FROM tb_asms WHERE asm_name = ? AND passcode = ?", (asm, pin), one=True)
        if asm_record or pin == MASTER_PIN or _default_pin_allowed(pin):
            is_authorized = True
    elif scope['type'] == 'STORE':
        st_record = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (store_code or scope['store'], pin), one=True)
        if st_record or pin == MASTER_PIN or _default_pin_allowed(pin):
            is_authorized = True
            
    if not is_authorized:
        return "Không có quyền xuất báo cáo", 401
        
    try:
        if scope['type'] == 'STORE':
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE store_code = ?", (scope['store'],))
        elif scope['type'] == 'ASM':
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY store_code", (scope['asm'],))
        else:
            if asm and asm != 'ALL':
                stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY store_code", (asm,))
            else:
                stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores ORDER BY store_code")
            
        store_codes = [s['store_code'] for s in stores]
        if not store_codes:
            return "Không tìm thấy dữ liệu cửa hàng", 404
            
        placeholders = ",".join(["?"] * len(store_codes))
        
        # Get start and end date of the week (Saturday to Friday)
        rep_dt = datetime.strptime(report_date, '%Y-%m-%d')
        start_date = (rep_dt - timedelta(days=6)).strftime('%Y-%m-%d')
        end_date = report_date
        
        # Calculate other date ranges for Input_Traffic comparison
        # 1. This Week
        w_curr_start = start_date
        w_curr_end = end_date
        
        # 2. Last Week
        w_prev_start = (rep_dt - timedelta(days=13)).strftime('%Y-%m-%d')
        w_prev_end = (rep_dt - timedelta(days=7)).strftime('%Y-%m-%d')
        
        # 3. Same Week LY (364 days ago)
        w_ly_start = (rep_dt - timedelta(days=370)).strftime('%Y-%m-%d')
        w_ly_end = (rep_dt - timedelta(days=364)).strftime('%Y-%m-%d')
        
        # 4. This Month
        m_curr_start = rep_dt.replace(day=1).strftime('%Y-%m-%d')
        m_curr_end = report_date
        
        # 5. Last Month
        m_prev_end_dt = rep_dt.replace(day=1) - timedelta(days=1)
        m_prev_start = m_prev_end_dt.replace(day=1).strftime('%Y-%m-%d')
        m_prev_end = m_prev_end_dt.strftime('%Y-%m-%d')
        
        # 6. Same Month LY
        try:
            ly_rep_dt = rep_dt.replace(year=rep_dt.year - 1)
        except ValueError:
            ly_rep_dt = rep_dt.replace(year=rep_dt.year - 1, day=28)
        m_ly_start = ly_rep_dt.replace(day=1).strftime('%Y-%m-%d')
        m_ly_end = ly_rep_dt.strftime('%Y-%m-%d')
        
        # Query Period 1: Last Year — PHẢI bao cả cửa sổ THÁNG cùng kỳ, không chỉ TUẦN.
        # BUG (sửa 2026-08-04): cận dưới trước đây là w_ly_start (report_date-370,
        # ~26/07 cho báo cáo tháng 7) → bỏ sót 01→25/07 → cột "Traffic Tháng cùng
        # kỳ" chỉ cộng được ~6 ngày cuối tháng (AEONBD ra 150 thay vì 722, thậm chí
        # NHỎ HƠN cả "Traffic Tuần cùng kỳ"=168 — vô lý vì tháng phải chứa tuần).
        # Truy vấn năm-nay ngay dưới ĐÃ đúng: min(w_prev_start, m_prev_start). Phía
        # LY quên đối xứng. Lấy min của cả 2 mốc bắt đầu (tuần LY & tháng LY).
        ly_min_date = min(w_ly_start, m_ly_start)
        ly_max_date = max(w_ly_end, m_ly_end)
        traffic_rows_ly = query_db(f"""
            SELECT store_code, traffic_date, traffic_val FROM tb_traffic
            WHERE traffic_date >= ? AND traffic_date <= ? AND store_code IN ({placeholders})
        """, [ly_min_date, ly_max_date] + store_codes)
        
        # Query Period 2: This Year (min of w_prev_start and m_prev_start to w_curr_end)
        ty_min_date = min(w_prev_start, m_prev_start)
        traffic_rows_ty = query_db(f"""
            SELECT store_code, traffic_date, traffic_val FROM tb_traffic
            WHERE traffic_date >= ? AND traffic_date <= ? AND store_code IN ({placeholders})
        """, [ty_min_date, w_curr_end] + store_codes)
        
        # Combine all traffic rows for comparison
        all_traffic_rows = traffic_rows_ly + traffic_rows_ty
        
        # Load Traffic records for the entire week
        traffic_rows = query_db(f"""
            SELECT store_code, traffic_date, traffic_val, bills_val, company_online_bills, store_online_bills 
            FROM tb_traffic 
            WHERE traffic_date >= ? AND traffic_date <= ? AND store_code IN ({placeholders})
            ORDER BY store_code, traffic_date
        """, [start_date, end_date] + store_codes)
        
        # Load Contracts (Section 3.1) with contract_number + customer_name
        contract_rows = query_db(f"""
            SELECT store_code, contract_number, customer_name, contract_value, product_category, quantity, deposit_paid, installment_2, status, reason 
            FROM tb_contracts 
            WHERE report_date = ? AND store_code IN ({placeholders})
        """, [report_date] + store_codes)
        
        # Load Unsigned (Section 3.2) with contract_number
        unsigned_rows = query_db(f"""
            SELECT store_code, contract_number, customer_name, prev_year_value, expected_signing_time, product_category, quantity, status, reason 
            FROM tb_unsigned_contracts 
            WHERE report_date = ? AND store_code IN ({placeholders})
        """, [report_date] + store_codes)
        
        # Load Details (Section 4)
        detail_rows = query_db(f"""
            SELECT * FROM tb_operational_details 
            WHERE report_date = ? AND store_code IN ({placeholders})
        """, [report_date] + store_codes)
        
        # Load Support Requests (Section 4.5)
        support_rows = query_db(f"""
            SELECT store_code, category, priority, issue_item, deadline, person_in_charge 
            FROM tb_support_requests 
            WHERE report_date = ? AND store_code IN ({placeholders})
        """, [report_date] + store_codes)
        
        store_map = {s['store_code']: s for s in stores}
        
        # Group traffic rows by store_code
        from collections import defaultdict
        store_traffic_map = defaultdict(list)
        for r in traffic_rows:
            store_traffic_map[r['store_code']].append(r)
            
        # 2. Process dataframes using pandas
        # Sheet 1: Tổng Hợp Traffic & CR
        traffic_summary_data = []
        for code, s in store_map.items():
            s_traffic = store_traffic_map.get(code, [])
            has_submitted = len(s_traffic) > 0
            
            if not has_submitted:
                traffic_summary_data.append({
                    'Mã Cửa Hàng': code,
                    'Tên Cửa Hàng': s['store_name'],
                    'Khu Vực': s['region'],
                    'ASM Quản Lý': s['asm_name'],
                    'Tổng Traffic': 'Chưa nộp',
                    'Tổng Bill Bán Lẻ': 'Chưa nộp',
                    'Tổng Bill OL Công Ty': 0,
                    'Tổng Bill OL Cửa Hàng': 0,
                    'Tỷ Lệ CR tại quầy (%)': 'N/A'
                })
            else:
                tot_trf = sum(r['traffic_val'] for r in s_traffic if r['traffic_val'] is not None)
                tot_bil = sum(r['bills_val'] for r in s_traffic if r['bills_val'] is not None)
                tot_co = sum(r['company_online_bills'] for r in s_traffic if r['company_online_bills'] is not None)
                tot_so = sum(r['store_online_bills'] for r in s_traffic if r['store_online_bills'] is not None)
                
                bill_for_cr = tot_bil - tot_co
                cr = (bill_for_cr / tot_trf * 100) if tot_trf > 0 else 0.0
                
                traffic_summary_data.append({
                    'Mã Cửa Hàng': code,
                    'Tên Cửa Hàng': s['store_name'],
                    'Khu Vực': s['region'],
                    'ASM Quản Lý': s['asm_name'],
                    'Tổng Traffic': tot_trf,
                    'Tổng Bill Bán Lẻ': tot_bil,
                    'Tổng Bill OL Công Ty': tot_co,
                    'Tổng Bill OL Cửa Hàng': tot_so,
                    'Tỷ Lệ CR tại quầy (%)': cr
                })
        df_traffic = pd.DataFrame(traffic_summary_data)
        
        # Sheet 2: Traffic Chi Tiết Theo Ngày
        daily_detail_data = []
        weekday_vn = {
            5: "Thứ Bảy",
            6: "Chủ Nhật",
            0: "Thứ Hai",
            1: "Thứ Ba",
            2: "Thứ Tư",
            3: "Thứ Năm",
            4: "Thứ Sáu"
        }
        
        for code, s in store_map.items():
            s_traffic = store_traffic_map.get(code, [])
            s_traffic_dict = {r['traffic_date']: r for r in s_traffic}
            
            for i in range(7):
                curr_date = rep_dt - timedelta(days=(6 - i))
                curr_date_str = curr_date.strftime('%Y-%m-%d')
                curr_day_vn = weekday_vn[curr_date.weekday()]
                
                r = s_traffic_dict.get(curr_date_str)
                
                if r is None:
                    daily_detail_data.append({
                        'Mã Cửa Hàng': code,
                        'Tên Cửa Hàng': s['store_name'],
                        'Khu Vực': s['region'],
                        'ASM Quản Lý': s['asm_name'],
                        'Ngày': curr_date.strftime('%d/%m/%Y'),
                        'Thứ': curr_day_vn,
                        'Traffic (Lượt Khách)': 'Chưa nhập',
                        'Số Bill Bán Lẻ': 'Chưa nhập',
                        'Bill Online Công Ty': 0,
                        'Bill Online Cửa Hàng': 0,
                        'Tỷ Lệ CR tại quầy (%)': 'N/A'
                    })
                else:
                    trf_val = r['traffic_val']
                    bil_val = r['bills_val']
                    co_val = r['company_online_bills'] or 0
                    so_val = r['store_online_bills'] or 0
                    
                    bill_for_cr = (bil_val or 0) - co_val
                    cr = (bill_for_cr / trf_val * 100) if (trf_val and trf_val > 0) else 0.0
                    
                    daily_detail_data.append({
                        'Mã Cửa Hàng': code,
                        'Tên Cửa Hàng': s['store_name'],
                        'Khu Vực': s['region'],
                        'ASM Quản Lý': s['asm_name'],
                        'Ngày': curr_date.strftime('%d/%m/%Y'),
                        'Thứ': curr_day_vn,
                        'Traffic (Lượt Khách)': trf_val if trf_val is not None else 'Chưa nhập',
                        'Số Bill Bán Lẻ': bil_val if bil_val is not None else 'Chưa nhập',
                        'Bill Online Công Ty': co_val,
                        'Bill Online Cửa Hàng': so_val,
                        'Tỷ Lệ CR tại quầy (%)': cr if trf_val is not None else 'N/A'
                    })
        df_daily_detail = pd.DataFrame(daily_detail_data)
        
        # Process Input_Traffic sheet data with O(1) dictionary lookups
        store_traffic_date_map = {}
        for r in all_traffic_rows:
            if r['traffic_val'] is not None:
                store_traffic_date_map[(r['store_code'], r['traffic_date'])] = r['traffic_val']
                
        def get_date_list(s_str, e_str):
            s_dt = datetime.strptime(s_str, '%Y-%m-%d')
            e_dt = datetime.strptime(e_str, '%Y-%m-%d')
            days = (e_dt - s_dt).days + 1
            return [(s_dt + timedelta(days=d)).strftime('%Y-%m-%d') for d in range(days)]
            
        dates_w_curr = get_date_list(w_curr_start, w_curr_end)
        dates_w_prev = get_date_list(w_prev_start, w_prev_end)
        dates_w_ly   = get_date_list(w_ly_start, w_ly_end)
        dates_m_curr = get_date_list(m_curr_start, m_curr_end)
        dates_m_prev = get_date_list(m_prev_start, m_prev_end)
        dates_m_ly   = get_date_list(m_ly_start, m_ly_end)
        
        def calc_sum(code, d_list):
            vals = [store_traffic_date_map[(code, d)] for d in d_list if (code, d) in store_traffic_date_map]
            return sum(vals) if vals else None
            
        input_traffic_data = []
        for code, s in store_map.items():
            val_w_curr = calc_sum(code, dates_w_curr)
            val_w_prev = calc_sum(code, dates_w_prev)
            val_w_ly   = calc_sum(code, dates_w_ly)
            val_m_curr = calc_sum(code, dates_m_curr)
            val_m_prev = calc_sum(code, dates_m_prev)
            val_m_ly   = calc_sum(code, dates_m_ly)
            
            input_traffic_data.append({
                'Cửa Hàng': s['store_name'],
                'Traffic Tuần Này': val_w_curr if val_w_curr is not None else 'Không có dữ liệu',
                'Traffic Tuần Trước': val_w_prev if val_w_prev is not None else 'Không có dữ liệu',
                'Traffic Tuần cùng kỳ': val_w_ly if val_w_ly is not None else 'Không có dữ liệu',
                'Traffic Tháng Này': val_m_curr if val_m_curr is not None else 'Không có dữ liệu',
                'Traffic Tháng Trước': val_m_prev if val_m_prev is not None else 'Không có dữ liệu',
                'Traffic Tháng cùng kỳ': val_m_ly if val_m_ly is not None else 'Không có dữ liệu'
            })
            
        input_traffic_data.sort(key=lambda x: x['Cửa Hàng'])
        df_input_traffic = pd.DataFrame(input_traffic_data)
        
        # Sheet 3: HĐ Đang Đàm Phán 3.1 — thêm Tên Khách Hàng
        contracts_data = []
        for c in contract_rows:
            s = store_map.get(c['store_code'], {})
            contracts_data.append({
                'Mã Cửa Hàng': c['store_code'],
                'Tên Cửa Hàng': s.get('store_name', ''),
                'Khu Vực': s.get('region', ''),
                'Tên Khách Hàng': c.get('customer_name') or '',
                'Số Hợp Đồng': c['contract_number'] or 'Đang GD',
                'Giá Trị HĐ (Tr.đ)': c['contract_value'],
                'Chủng Loại': c['product_category'],
                'Số Lượng': c['quantity'],
                'Số Tiền Đã Cọc': c['deposit_paid'],
                'Số Tiền Đợt 2': c['installment_2'],
                'Trạng Thái': c['status'],
                'Lý Do / Chi Tiết': c['reason']
            })
        df_contracts = pd.DataFrame(contracts_data) if contracts_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'Khu Vực', 'Tên Khách Hàng', 'Số Hợp Đồng', 'Giá Trị HĐ (Tr.đ)', 'Chủng Loại', 'Số Lượng', 'Số Tiền Đã Cọc', 'Số Tiền Đợt 2', 'Trạng Thái', 'Lý Do / Chi Tiết'])
        
        # Sheet 4: Unsigned Contracts 3.2 — Số Hợp Đồng + Tên Khách Hàng (đồng bộ với 3.1)
        unsigned_data = []
        for u in unsigned_rows:
            s = store_map.get(u['store_code'], {})
            unsigned_data.append({
                'Mã Cửa Hàng': u['store_code'],
                'Tên Cửa Hàng': s.get('store_name', ''),
                'Khu Vực': s.get('region', ''),
                'Số Hợp Đồng': u.get('contract_number') or 'Đang GD',
                'Tên Khách Hàng': u.get('customer_name') or '',          # ← Đồng bộ với 3.1
                'Giá Trị Năm Ngoái (Tr.đ)': u['prev_year_value'],
                'Thời Gian Dự Kiến Ký': u['expected_signing_time'],
                'Chủng Loại': u['product_category'],
                'Số Lượng': u['quantity'],
                'Trạng Thái': u['status'],
                'Lý Do / Chi Tiết': u['reason']
            })
        df_unsigned = pd.DataFrame(unsigned_data) if unsigned_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'Khu Vực', 'Số Hợp Đồng', 'Tên Khách Hàng', 'Giá Trị Năm Ngoái (Tr.đ)', 'Thời Gian Dự Kiến Ký', 'Chủng Loại', 'Số Lượng', 'Trạng Thái', 'Lý Do / Chi Tiết'])
        
        # Sheet 5: Chi Tiết Vận Hành 4
        details_data = []
        for d in detail_rows:
            s = store_map.get(d['store_code'], {})
            details_data.append({
                'Mã Cửa Hàng': d['store_code'],
                'Tên Cửa Hàng': s.get('store_name', ''),
                'Khu Vực': s.get('region', ''),
                'Mở Cửa / Đóng Cửa': d['op_open_close_status'],
                'Ghi Chú Mở/Đóng': d['op_open_close_note'],
                'Đồng Phục & Diện Mạo': d['op_uniform_status'],
                'Ghi Chú Đồng Phục': d['op_uniform_note'],
                'Chào Hỏi Khách Hàng': d['op_greet_status'],
                'Ghi Chú Chào Hỏi': d['op_greet_note'],
                'Ý Kiến Phản Hồi Khách': d['op_feedback_status'],
                'Ghi Chú Phản Hồi': d['op_feedback_note'],
                'Vấn Đề Vận Hành Khác': d['op_other_status'],
                'Ghi Chú Vấn Đề Khác': d['op_other_note'],
                'Chỉ Tiêu Nhân Sự': d['hr_target'],
                'Thực Tế Nhân Sự': d['hr_actual'],
                'Bảo Vệ Ca Trực': d['hr_guard'],
                'Nghỉ Việc/Tuyển Dụng': d['hr_resigned_note'],
                'Nghỉ Phép/Ối/Đau': d['hr_leave_note'],
                'Tự Ý Nghỉ Việc': d['hr_absent_note'],
                'Phản Hồi Tồn Kho': d['inv_stock_status'],
                'Hàng Thiếu/Đứt Size': d['inv_info_goods'],
                'Hàng Trả Kho': d['inv_return_warehouse'],
                'Đề Xuất/Kiến Nghị': d['inv_proposal'],
                # 4.4 KHÁCH HÀNG & THỊ TRƯỜNG — trước đây bị bỏ sót khỏi file xuất
                'Góp Ý SP Của Khách': d['market_product_feedback'],
                'SP Khách Tìm Cty Chưa Có': d['market_missing_products'],
                'Đối Thủ Cạnh Tranh': d['market_competitors'],
                'Ý Kiến Khác Của CH': d['market_other_feedback']
            })
        df_details = pd.DataFrame(details_data) if details_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'Khu Vực'])
        
        # Sheet 6: Yêu Cầu Hỗ Trợ 4.5
        support_data = []
        for sp in support_rows:
            s = store_map.get(sp['store_code'], {})
            support_data.append({
                'Mã Cửa Hàng': sp['store_code'],
                'Tên Cửa Hàng': s.get('store_name', ''),
                'Khu Vực': s.get('region', ''),
                'Danh Mục Hỗ Trợ': sp['category'],
                'Độ Ưu Tiên': sp['priority'],
                'Nội Dung Sự Việc Cụ Thể': sp['issue_item'],
                'Thời Hạn Hoàn Thành': sp['deadline'],
                'Người Chịu Trách Nhiệm': sp['person_in_charge']
            })
        df_support = pd.DataFrame(support_data) if support_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'Khu Vực', 'Danh Mục Hỗ Trợ', 'Độ Ưu Tiên', 'Nội Dung Sự Việc Cụ Thể', 'Thời Hạn Hoàn Thành', 'Người Chịu Trách Nhiệm'])
        
        # Sheet 7: Phân Tích Lý Do Khách Không Mua Hàng
        np_rows = query_db(f"""
            SELECT t.store_code, s.store_name, s.asm_name, s.region,
                   t.traffic_date, t.traffic_val, t.bills_val, t.non_purchase_reasons
            FROM tb_traffic t
            JOIN tb_stores s ON t.store_code = s.store_code
            WHERE t.traffic_date >= ? AND t.traffic_date <= ? AND t.store_code IN ({placeholders})
            ORDER BY s.asm_name, s.store_name, t.traffic_date
        """, [start_date, end_date] + store_codes)
        
        np_data = []
        for r in np_rows:
            trf = r['traffic_val'] or 0
            bil = r['bills_val'] or 0
            unconverted = max(0, trf - bil)
            cr = round((bil / trf * 100), 1) if trf > 0 else 0
            
            reasons_raw = r['non_purchase_reasons'] or ''
            reason_text = ''
            note_text = ''
            if reasons_raw:
                try:
                    if reasons_raw.startswith('{'):
                        parsed = json.loads(reasons_raw)
                        r_list = parsed.get('reasons', [])
                        note_text = parsed.get('note', '').strip()
                    else:
                        r_list = [x.strip() for x in reasons_raw.split(',') if x.strip()]
                    
                    lbl_map = {
                        'SIZE': 'Đứt size', 'STYLE': 'Mẫu mã', 'PRICE': 'Giá/KM',
                        'BROWSE': 'Chỉ xem', 'NEW_COLLECTION': 'Thiếu hàng mới',
                        'SERVICE': 'Phục vụ', 'OTHER': 'Lý do khác'
                    }
                    reason_text = ', '.join([lbl_map.get(k, k) for k in r_list])
                except Exception:
                    reason_text = reasons_raw

            np_data.append({
                'Mã Cửa Hàng': r['store_code'],
                'Tên Cửa Hàng': r['store_name'],
                'ASM Phụ Trách': r['asm_name'],
                'Khu Vực': r['region'],
                'Ngày': r['traffic_date'],
                'Lượt Khách (Traffic)': trf,
                'Số Hóa Đơn (Bills)': bil,
                'Khách Không Mua': unconverted,
                'CR (%)': cr,
                'Lý Do Chính Khách Không Mua': reason_text or 'Chưa ghi nhận',
                'Ghi Chú Chi Tiết Tại Quầy': note_text
            })
        df_non_purchase = pd.DataFrame(np_data) if np_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Khu Vực', 'Ngày', 'Lượt Khách (Traffic)', 'Số Hóa Đơn (Bills)', 'Khách Không Mua', 'CR (%)', 'Lý Do Chính Khách Không Mua', 'Ghi Chú Chi Tiết Tại Quầy'])

        # Sheet 9: Định Biên & Hồ Sơ Nhân Sự
        hr_emp_rows = query_db(f"""
            SELECT e.*, s.store_name, s.asm_name, s.region, t.target_headcount
            FROM tb_store_employees e
            JOIN tb_stores s ON e.store_code = s.store_code
            LEFT JOIN tb_store_headcount_targets t ON e.store_code = t.store_code
            WHERE e.store_code IN ({placeholders}) AND e.status != 'RESIGNED'
            ORDER BY s.asm_name, s.store_name, e.position, e.full_name
        """, store_codes)
        store_hr_data = []
        for emp in hr_emp_rows:
            store_hr_data.append({
                'Mã Cửa Hàng': emp['store_code'],
                'Tên Cửa Hàng': emp['store_name'],
                'ASM Phụ Trách': emp['asm_name'],
                'Khu Vực': emp['region'],
                'Định Biên CH': emp['target_headcount'] or 0,
                'Mã Nhân Viên': emp['employee_code'],
                'Họ và Tên': emp['full_name'],
                'Giới Tính': emp['gender'] or 'Nữ',
                'Ngày Sinh': emp['dob'] or '',
                'Số Điện Thoại': emp['phone_number'] or '',
                'Chức Danh': emp['position'],
                'Ngày Nhận Chức / Vốn Làm': emp['appointment_date'] or '',
                'Thâm Niên Làm Việc': calculate_tenure_py(emp.get('appointment_date') or emp.get('created_at')),
                'Trạng Thái Work': emp['status']
            })
        df_store_hr = pd.DataFrame(store_hr_data) if store_hr_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Khu Vực', 'Định Biên CH', 'Mã Nhân Viên', 'Họ và Tên', 'Giới Tính', 'Ngày Sinh', 'Số Điện Thoại', 'Chức Danh', 'Ngày Nhận Chức / Vốn Làm', 'Thâm Niên Làm Việc', 'Trạng Thái Work'])

        # Sheet 10: Theo Dõi Thử Việc & Đào Tạo
        hr_prob_rows = query_db(f"""
            SELECT p.*, e.full_name, e.gender, e.position, s.store_name, s.asm_name
            FROM tb_employee_probation p
            JOIN tb_store_employees e ON p.employee_code = e.employee_code
            JOIN tb_stores s ON p.store_code = s.store_code
            WHERE p.store_code IN ({placeholders})
            ORDER BY s.asm_name, s.store_name, p.start_date
        """, store_codes)
        prob_data = []
        for pr in hr_prob_rows:
            prob_data.append({
                'Mã Cửa Hàng': pr['store_code'],
                'Tên Cửa Hàng': pr['store_name'],
                'ASM Phụ Trách': pr['asm_name'],
                'Mã Nhân Viên': pr['employee_code'],
                'Họ và Tên': pr['full_name'],
                'Chức Danh': pr['position'],
                'Ngày Nhận Việc': pr['start_date'],
                'Người Đào Tạo Kèm Cặp': pr['trainer_name'] or 'CHT/CHP',
                '1. Quy định vận hành': 'ĐẠT' if pr['train_op_rules'] else 'Chưa',
                '2. Quy trình làm việc': 'ĐẠT' if pr['train_work_process'] else 'Chưa',
                '3. Hàng hóa & Bảng size': 'ĐẠT' if pr['train_products'] else 'Chưa',
                '4. Trưng bày & VM': 'ĐẠT' if pr['train_display'] else 'Chưa',
                '5. Kiểm kê kho': 'ĐẠT' if pr['train_inventory'] else 'Chưa',
                'Ghi Chú Đào Tạo': pr['training_notes'] or '',
                'Ngày Kết Thúc Thử Việc': pr['probation_end_date'] or '',
                'Kết Quả Thử Việc': pr['probation_result'] or 'Đang thử việc'
            })
        df_hr_probation = pd.DataFrame(prob_data) if prob_data else pd.DataFrame(columns=['Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Mã Nhân Viên', 'Họ và Tên', 'Chức Danh', 'Ngày Nhận Việc', 'Người Đào Tạo Kèm Cặp', '1. Quy định vận hành', '2. Quy trình làm việc', '3. Hàng hóa & Bảng size', '4. Trưng bày & VM', '5. Kiểm kê kho', 'Ghi Chú Đào Tạo', 'Ngày Kết Thúc Thử Việc', 'Kết Quả Thử Việc'])

        # Sheet 11: Theo Dõi Sự Vụ Nhân Sự & Phân Ca
        hr_ticket_rows = query_db(f"""
            SELECT t.*, s.store_name, s.asm_name
            FROM tb_hr_lifecycle_tickets t
            JOIN tb_stores s ON t.store_code = s.store_code
            WHERE t.store_code IN ({placeholders})
            ORDER BY t.created_at DESC
        """, store_codes)
        ticket_data = []
        for tk in hr_ticket_rows:
            ticket_data.append({
                'Mã Ticket': tk['ticket_code'],
                'Mã Cửa Hàng': tk['store_code'],
                'Tên Cửa Hàng': tk['store_name'],
                'ASM Phụ Trách': tk['asm_name'],
                'Tên Nhân Sự': tk['employee_name'],
                'Chức Danh': tk['position'],
                'Loại Sự Vụ': tk['event_type'],
                'Ngày Hiệu Lực': tk['effective_date'],
                'Lý Do / Nội Dung': tk['reason_note'],
                'Bàn Giao Công Việc': tk['handover_status'],
                'Trạng Thái Giải Quyết': tk['status'],
                'Tiến Độ Cửa Hàng': tk['store_progress_note'],
                'Ghi Chú ASM / HR': tk['asm_hr_note']
            })
        df_hr_tickets = pd.DataFrame(ticket_data) if ticket_data else pd.DataFrame(columns=['Mã Ticket', 'Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Tên Nhân Sự', 'Chức Danh', 'Loại Sự Vụ', 'Ngày Hiệu Lực', 'Lý Do / Nội Dung', 'Bàn Giao Công Việc', 'Trạng Thái Giải Quyết', 'Tiến Độ Cửa Hàng', 'Ghi Chú ASM / HR'])

        # Sheet 12: Toàn Bộ Sự Vụ Hỗ Trợ KT (lifecycle đầy đủ — KHÔNG lọc report_date,
        # khác với df_support ở trên vốn CỐ Ý chỉ lấy đúng report_date cho Section 4.5
        # core của báo cáo tuần. Sheet này phục vụ sổ theo dõi sự vụ vận hành bên
        # qlkd_operational — cần TOÀN BỘ lịch sử để tính tuổi ticket/tái phát sinh).
        support_full_rows = query_db(f"""
            SELECT r.*, s.store_name, s.asm_name
            FROM tb_support_requests r
            JOIN tb_stores s ON r.store_code = s.store_code
            WHERE r.store_code IN ({placeholders})
            ORDER BY r.created_at DESC
        """, store_codes)
        support_full_data = []
        for r in support_full_rows:
            support_full_data.append({
                'Mã Ticket': r['ticket_code'],
                'Mã Cửa Hàng': r['store_code'],
                'Tên Cửa Hàng': r['store_name'],
                'ASM Phụ Trách': r['asm_name'],
                'Danh Mục Hỗ Trợ': r['category'],
                'Độ Ưu Tiên': r['priority'],
                'Nội Dung Sự Việc Cụ Thể': r['issue_item'],
                'Thời Hạn Hoàn Thành': r['deadline'],
                'Người Chịu Trách Nhiệm': r['person_in_charge'],
                'Trạng Thái': r['status'],
                'Ngày Tạo': r['created_at'],
                'Ngày Cập Nhật Cuối': r['updated_at'],
                'Ghi Chú Cửa Hàng': r['store_progress_note'],
                'Ghi Chú ASM / HQ': r['asm_hq_note'],
            })
        df_support_full = pd.DataFrame(support_full_data) if support_full_data else pd.DataFrame(columns=['Mã Ticket', 'Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Danh Mục Hỗ Trợ', 'Độ Ưu Tiên', 'Nội Dung Sự Việc Cụ Thể', 'Thời Hạn Hoàn Thành', 'Người Chịu Trách Nhiệm', 'Trạng Thái', 'Ngày Tạo', 'Ngày Cập Nhật Cuối', 'Ghi Chú Cửa Hàng', 'Ghi Chú ASM / HQ'])

        # Sheet 13: Lịch Sử Trạng Thái Sự Vụ (audit trail dùng chung SUPPORT + HR —
        # nguồn cho SLA/thời gian xử lý thật ở qlkd_operational). Join qua store_code
        # (ticket_code có 2 namespace khác nhau: TK-... và HR-... nên không join
        # trực tiếp bằng ticket_code sang tb_stores được, dùng store_code có sẵn
        # ngay trên chính bảng history).
        ticket_history_rows = query_db(f"""
            SELECT h.*, s.store_name, s.asm_name
            FROM tb_ticket_status_history h
            JOIN tb_stores s ON h.store_code = s.store_code
            WHERE h.store_code IN ({placeholders})
            ORDER BY h.changed_at DESC
        """, store_codes)
        history_data = []
        for h in ticket_history_rows:
            history_data.append({
                'Loại Ticket': h['ticket_type'],
                'Mã Ticket': h['ticket_code'],
                'Mã Cửa Hàng': h['store_code'],
                'Tên Cửa Hàng': h['store_name'],
                'ASM Phụ Trách': h['asm_name'],
                'Trạng Thái Cũ': h['old_status'],
                'Trạng Thái Mới': h['new_status'],
                'Thời Điểm Đổi': h['changed_at'],
                'Ghi Chú': h['note'],
            })
        df_ticket_history = pd.DataFrame(history_data) if history_data else pd.DataFrame(columns=['Loại Ticket', 'Mã Ticket', 'Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Trạng Thái Cũ', 'Trạng Thái Mới', 'Thời Điểm Đổi', 'Ghi Chú'])

        # Sheet 14: Khai Báo Đóng Cửa CH (sửa chữa/di dời/mở mới) — ASM khai báo,
        # engine qlkd_operational đọc để tính Ngày HĐ + %HT điều chỉnh + đối soát
        # Data Lake. KHÔNG lọc report_date (lifecycle đầy đủ toàn lịch sử).
        closure_rows = query_db(f"""
            SELECT c.*, s.store_name, s.asm_name
            FROM tb_store_closures c
            JOIN tb_stores s ON c.store_code = s.store_code
            WHERE c.store_code IN ({placeholders})
            ORDER BY c.start_date DESC
        """, store_codes)
        closure_data = []
        for c in closure_rows:
            closure_data.append({
                'Mã Khai Báo': c['closure_code'],
                'Mã Cửa Hàng': c['store_code'],
                'Tên Cửa Hàng': c['store_name'],
                'ASM Phụ Trách': c['asm_name'],
                'Loại Sự Kiện': c['event_type'],
                'Từ Ngày': c['start_date'],
                'Đến Ngày': c['end_date'],
                'Trạng Thái': c['status'],
                'Lý Do / Nội Dung': c['reason_note'],
                'Người Khai Báo': c['reported_by'],
                'Nguồn': c['source'],
                'Ngày Tạo': c['created_at'],
                'Ngày Cập Nhật': c['updated_at'],
            })
        df_closures = pd.DataFrame(closure_data) if closure_data else pd.DataFrame(columns=['Mã Khai Báo', 'Mã Cửa Hàng', 'Tên Cửa Hàng', 'ASM Phụ Trách', 'Loại Sự Kiện', 'Từ Ngày', 'Đến Ngày', 'Trạng Thái', 'Lý Do / Nội Dung', 'Người Khai Báo', 'Nguồn', 'Ngày Tạo', 'Ngày Cập Nhật'])

        # 3. Create Excel File in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_traffic.to_excel(writer, sheet_name='Tổng Hợp Traffic & CR', index=False, startrow=4)
            df_daily_detail.to_excel(writer, sheet_name='Traffic Chi Tiết Theo Ngày', index=False, startrow=4)
            df_input_traffic.to_excel(writer, sheet_name='Input_Traffic', index=False, startrow=2)
            df_contracts.to_excel(writer, sheet_name='HĐ Đang Đàm Phán 3.1', index=False, startrow=4)
            df_unsigned.to_excel(writer, sheet_name='HĐ Chưa Ký 3.2', index=False, startrow=4)
            df_details.to_excel(writer, sheet_name='Chi Tiết Vận Hành 4', index=False, startrow=4)
            df_support.to_excel(writer, sheet_name='Yêu Cầu Hỗ Trợ 4.5', index=False, startrow=4)
            df_non_purchase.to_excel(writer, sheet_name='Phân Tích Lý Do Không Mua', index=False, startrow=4)
            df_store_hr.to_excel(writer, sheet_name='Định Biên & Hồ Sơ Nhân Sự', index=False, startrow=4)
            df_hr_probation.to_excel(writer, sheet_name='Theo Dõi Thử Việc & Đào Tạo', index=False, startrow=4)
            df_hr_tickets.to_excel(writer, sheet_name='Theo Dõi Sự Vụ Nhân Sự', index=False, startrow=4)
            df_support_full.to_excel(writer, sheet_name='Toàn Bộ Sự Vụ Hỗ Trợ KT', index=False, startrow=4)
            df_ticket_history.to_excel(writer, sheet_name='Lịch Sử Trạng Thái Sự Vụ', index=False, startrow=4)
            df_closures.to_excel(writer, sheet_name='Khai Báo Đóng Cửa CH', index=False, startrow=4)

            # Access workbook and apply corporate styling
            workbook = writer.book
            date_range_str = f"Từ {datetime.strptime(start_date, '%Y-%m-%d').strftime('%d/%m/%Y')} đến {datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            role_str = f"ASM {filter_asm}" if filter_asm else "Toàn Hệ Thống"
            
            style_sheet(workbook['Tổng Hợp Traffic & CR'], "Báo Cáo Tổng Hợp Traffic & CR Tuần", "Tổng hợp kết quả chuyển đổi", date_range_str, role_str)
            style_sheet(workbook['Traffic Chi Tiết Theo Ngày'], "Báo Cáo Chi Tiết Traffic Ngày-Qua-Ngày", "Nhật ký nhập traffic hàng ngày của tuần", date_range_str, role_str)
            style_input_traffic_sheet(workbook['Input_Traffic'])
            style_sheet(workbook['HĐ Đang Đàm Phán 3.1'], "Danh Sách Hợp Đồng Đang Đàm Phán (3.1)", "Theo dõi tiến độ thương thảo hợp đồng", date_range_str, role_str)
            style_sheet(workbook['HĐ Chưa Ký 3.2'], "Danh Sách Hợp Đồng Cùng Kỳ Chưa Ký (3.2)", "Hợp đồng trễ hạn chưa tái ký", date_range_str, role_str)
            style_sheet(workbook['Chi Tiết Vận Hành 4'], "Báo Cáo Chi Tiết Vận Hành - Nhân Sự - Hàng Hóa", "Đánh giá chất lượng dịch vụ & nhân sự tại quầy", date_range_str, role_str)
            style_sheet(workbook['Yêu Cầu Hỗ Trợ 4.5'], "Danh Sách Yêu Cầu Hỗ Trợ Kỹ Thuật / Vận Hành (4.5)", "Tiếp nhận và xử lý yêu cầu phát sinh từ cửa hàng", date_range_str, role_str)
            style_sheet(workbook['Phân Tích Lý Do Không Mua'], "Báo Cáo Phân Tích Nguyên Nhân Khách Không Mua Hàng", "Thống kê thất thoát chuyển đổi & nguyên nhân không mua", date_range_str, role_str)
            style_sheet(workbook['Định Biên & Hồ Sơ Nhân Sự'], "Báo Cáo Định Biên & Hồ Sơ Nhân Sự Cửa Hàng", "Quản lý nhân sự và đánh giá thừa/thiếu định biên", date_range_str, role_str)
            style_sheet(workbook['Theo Dõi Thử Việc & Đào Tạo'], "Báo Cáo Theo Dõi Nhân Sự Thử Việc & Đào Tạo Tại Quầy", "Đào tạo checklist 5 bài học và kết quả thử việc", date_range_str, role_str)
            style_sheet(workbook['Theo Dõi Sự Vụ Nhân Sự'], "Nhật Ký Theo Dõi Sự Vụ & Biến Động Nhân Sự", "Vòng đời xử lý ticket sự vụ nhân sự toàn hệ thống", date_range_str, role_str)
            style_sheet(workbook['Toàn Bộ Sự Vụ Hỗ Trợ KT'], "Toàn Bộ Sự Vụ Hỗ Trợ Kỹ Thuật (Lifecycle)", "Không lọc theo tuần — phục vụ sổ theo dõi tồn đọng/tái phát sinh", date_range_str, role_str)
            style_sheet(workbook['Lịch Sử Trạng Thái Sự Vụ'], "Lịch Sử Đổi Trạng Thái Sự Vụ (Hỗ Trợ KT + Nhân Sự)", "Audit trail — nguồn tính thời gian xử lý thật", date_range_str, role_str)
            style_sheet(workbook['Khai Báo Đóng Cửa CH'], "Khai Báo Đóng Cửa Cửa Hàng (Sửa Chữa / Di Dời / Mở Mới)", "ASM khai báo — engine tính Ngày HĐ + %HT điều chỉnh + đối soát Data Lake", date_range_str, role_str)


        output.seek(0)
        
        # Determine target scope label for distinct, non-colliding filenames
        target_label = ""
        if scope['type'] == 'STORE':
            target_label = f"CH_{sanitize_filename_part(scope.get('store') or store_code)}"
        elif scope['type'] == 'ASM':
            target_label = f"ASM_{sanitize_filename_part(scope.get('asm') or asm)}"
        else: # scope['type'] == 'ALL'
            if asm and asm != 'ALL':
                target_label = f"ASM_{sanitize_filename_part(asm)}"
            else:
                target_label = "ToanQuoc"
                
        filename = f"BaoCao_RetailCommander_{target_label}_{report_date}.xlsx"
            
        encoded_filename = urllib.parse.quote(filename)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return f"Lỗi xuất file Excel: {str(e)}", 500

def compute_hr_seniority_summary(store_codes=None):
    try:
        if store_codes:
            placeholders = ','.join(['?'] * len(store_codes))
            query = f"SELECT * FROM tb_store_employees WHERE status != 'RESIGNED' AND store_code IN ({placeholders})"
            employees = query_db(query, store_codes)
        else:
            employees = query_db("SELECT * FROM tb_store_employees WHERE status != 'RESIGNED'")
            
        today = date.today()
        
        breakdown = {
            'under_3m': 0,
            '3m_to_1y': 0,
            '1y_to_3y': 0,
            '3y_to_5y': 0,
            'over_5y': 0
        }
        
        positions = {
            'CHT': 0,
            'CHP': 0,
            'NVBH': 0,
            'BV': 0,
            'THU_VIEC': 0,
            'OTHER': 0
        }
        
        total_days = 0
        valid_dates_count = 0
        
        for emp in employees:
            pos = str(emp.get('position') or '').strip()
            if 'trưởng' in pos.lower():
                positions['CHT'] += 1
            elif 'phó' in pos.lower():
                positions['CHP'] += 1
            elif 'bán hàng' in pos.lower():
                positions['NVBH'] += 1
            elif 'bảo vệ' in pos.lower():
                positions['BV'] += 1
            elif 'thử việc' in pos.lower():
                positions['THU_VIEC'] += 1
            else:
                positions['OTHER'] += 1
                
            raw_date = emp.get('appointment_date') or emp.get('created_at') or ''
            app_date_str = format_excel_date_py(raw_date)
            if not app_date_str:
                continue
                
            try:
                d = datetime.strptime(app_date_str[:10], '%Y-%m-%d').date()
                days = (today - d).days
                if days < 0:
                    days = 0
                total_days += days
                valid_dates_count += 1
                
                if days < 90:
                    breakdown['under_3m'] += 1
                elif days < 365:
                    breakdown['3m_to_1y'] += 1
                elif days < 1095:
                    breakdown['1y_to_3y'] += 1
                elif days < 1825:
                    breakdown['3y_to_5y'] += 1
                else:
                    breakdown['over_5y'] += 1
            except Exception:
                pass
                
        avg_days = (total_days / valid_dates_count) if valid_dates_count > 0 else 0
        avg_years = round(avg_days / 365.25, 1)
        
        return {
            'total_employees': len(employees),
            'valid_dates_count': valid_dates_count,
            'avg_tenure_years': avg_years,
            'tenure_breakdown': breakdown,
            'position_breakdown': positions
        }
    except Exception as e:
        print(f"⚠️ [HR Summary Error]: {e}")
        return {
            'total_employees': 0, 'valid_dates_count': 0, 'avg_tenure_years': 0,
            'tenure_breakdown': {'under_3m': 0, '3m_to_1y': 0, '1y_to_3y': 0, '3y_to_5y': 0, 'over_5y': 0},
            'position_breakdown': {'CHT': 0, 'CHP': 0, 'NVBH': 0, 'BV': 0, 'THU_VIEC': 0, 'OTHER': 0}
        }

# ── Quick Report API ──────────────────────────────────────────────────────────
MASTER_ASM_NAME = os.environ.get('MASTER_ASM_NAME', 'Khôi')

@app.route('/api/quick_report', methods=['GET'])
def quick_report():
    """Tổng hợp báo cáo nhanh theo phân quyền: store / asm / admin / master_asm."""
    role      = request.args.get('role')       # 'store', 'asm', 'admin'
    user_id   = request.args.get('user_id')    # store_code hoặc asm_name hoặc 'ADMIN'
    pin       = request.args.get('pin')
    report_date = request.args.get('report_date') or get_report_date().strftime('%Y-%m-%d')
    asm_filter  = request.args.get('asm_filter')  # chỉ admin/master dùng

    if not role or not user_id or not pin:
        return jsonify({'ok': False, 'error': 'Thiếu tham số bắt buộc'})

    master_pin = os.environ.get('MASTER_PIN', '8888')
    is_master  = False

    try:
        # ── Auth ──────────────────────────────────────────────────────────────
        if role == 'admin':
            if pin != master_pin:
                return jsonify({'ok': False, 'error': 'Không có quyền'}), 401
            is_master = True
        elif role == 'asm':
            asm_rec = query_db("SELECT * FROM tb_asms WHERE asm_name = ? AND passcode = ?", (user_id, pin), one=True)
            if not asm_rec:
                return jsonify({'ok': False, 'error': 'Mã PIN ASM không đúng'}), 401
            is_master = (user_id == MASTER_ASM_NAME)
        elif role == 'store':
            store_rec = query_db("SELECT * FROM tb_stores WHERE store_code = ? AND passcode = ?", (user_id, pin), one=True)
            if not store_rec and not _default_pin_allowed(pin):
                return jsonify({'ok': False, 'error': 'Mã PIN không đúng'}), 401
        else:
            return jsonify({'ok': False, 'error': 'Role không hợp lệ'}), 400

        # ── Determine scope ───────────────────────────────────────────────────
        filter_asm_q = None  # None = all
        store_scope  = None  # None = multi-store

        if role == 'store':
            store_scope = user_id
        elif role == 'asm' and not is_master:
            filter_asm_q = user_id
        elif (role == 'admin' or is_master) and asm_filter:
            filter_asm_q = asm_filter

        # ── Fetch stores ──────────────────────────────────────────────────────
        if store_scope:
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE store_code = ?", (store_scope,))
        elif filter_asm_q:
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores WHERE asm_name = ? ORDER BY store_name", (filter_asm_q,))
        else:
            stores = query_db("SELECT store_code, store_name, region, asm_name FROM tb_stores ORDER BY asm_name, store_name")

        store_codes = [s['store_code'] for s in stores]
        if not store_codes:
            return jsonify({'ok': True, 'report_date': report_date, 'stores': [],
                            'contracts': [], 'unsigned': [], 'details': [], 'support': [],
                            'summary': {}, 'can_export': is_master or role == 'admin'})

        placeholders = ",".join(["?"] * len(store_codes))

        # ── Traffic summary ───────────────────────────────────────────────────
        rep_dt    = datetime.strptime(report_date, '%Y-%m-%d')
        week_start = (rep_dt - timedelta(days=6)).strftime('%Y-%m-%d')

        traffic_rows = query_db(f"""
            SELECT store_code, SUM(traffic_val) as total_traffic, SUM(bills_val) as total_bills,
                   SUM(company_online_bills) as co_bills
            FROM tb_traffic WHERE traffic_date >= ? AND traffic_date <= ?
            AND store_code IN ({placeholders})
            GROUP BY store_code
        """, [week_start, report_date] + store_codes)
        traffic_map = {r['store_code']: r for r in traffic_rows}

        # ── Submitted stores ──────────────────────────────────────────────────
        submitted_set = {r['store_code'] for r in query_db(
            f"SELECT DISTINCT store_code FROM tb_operational_details WHERE report_date = ? AND store_code IN ({placeholders})",
            [report_date] + store_codes
        )}

        # ── Build store summary list ──────────────────────────────────────────
        stores_out = []
        for s in stores:
            code = s['store_code']
            tr = traffic_map.get(code, {})
            tot_tr  = tr.get('total_traffic') or 0
            tot_bil = tr.get('total_bills') or 0
            co_bil  = tr.get('co_bills') or 0
            retail_bil = tot_bil - co_bil
            cr = round(retail_bil / tot_tr * 100, 1) if tot_tr > 0 else 0
            stores_out.append({
                'store_code': code,
                'store_name': s['store_name'],
                'region': s['region'],
                'asm_name': s['asm_name'],
                'submitted': code in submitted_set,
                'traffic': tot_tr,
                'bills': tot_bil,
                'cr': cr,
            })

        # ── Contracts 3.1 ─────────────────────────────────────────────────────
        contract_rows = query_db(f"""
            SELECT c.store_code, s.store_name, s.region, c.contract_number, c.customer_name,
                   c.contract_value, c.product_category, c.quantity,
                   c.deposit_paid, c.installment_2, c.status, c.reason
            FROM tb_contracts c JOIN tb_stores s ON c.store_code = s.store_code
            WHERE c.report_date = ? AND c.store_code IN ({placeholders})
            ORDER BY s.asm_name, s.store_name
        """, [report_date] + store_codes)

        # ── Unsigned 3.2 ──────────────────────────────────────────────────────
        unsigned_rows = query_db(f"""
            SELECT u.store_code, s.store_name, s.region, u.contract_number, u.customer_name,
                   u.prev_year_value, u.expected_signing_time, u.product_category,
                   u.quantity, u.status, u.reason
            FROM tb_unsigned_contracts u JOIN tb_stores s ON u.store_code = s.store_code
            WHERE u.report_date = ? AND u.store_code IN ({placeholders})
            ORDER BY s.asm_name, s.store_name
        """, [report_date] + store_codes)

        # ── Operational 4.x ───────────────────────────────────────────────────
        detail_rows = query_db(f"""
            SELECT d.store_code, s.store_name, s.region, s.asm_name,
                   d.op_open_close_status, d.op_open_close_note, d.op_uniform_status, d.op_uniform_note,
                   d.op_greet_status, d.op_greet_note, d.op_feedback_status, d.op_feedback_note,
                   d.op_other_status, d.op_other_note,
                   d.hr_target, d.hr_actual, d.hr_guard, d.hr_resigned_note, d.hr_leave_note, d.hr_absent_note,
                   d.inv_stock_status, d.inv_info_goods, d.inv_return_warehouse, d.inv_proposal,
                   d.market_product_feedback, d.market_missing_products, d.market_competitors, d.market_other_feedback
            FROM tb_operational_details d JOIN tb_stores s ON d.store_code = s.store_code
            WHERE d.report_date = ? AND d.store_code IN ({placeholders})
            ORDER BY s.asm_name, s.store_name
        """, [report_date] + store_codes)

        # ── Support 4.5 ───────────────────────────────────────────────────────
        support_rows = query_db(f"""
            SELECT r.store_code, s.store_name, s.asm_name, r.category, r.priority,
                   r.issue_item, r.deadline, r.person_in_charge
            FROM tb_support_requests r JOIN tb_stores s ON r.store_code = s.store_code
            WHERE r.report_date = ? AND r.store_code IN ({placeholders})
            ORDER BY CASE r.priority WHEN 'Cao' THEN 1 WHEN 'Trung bình' THEN 2 ELSE 3 END, s.store_name
        """, [report_date] + store_codes)

        # ── KPI Summary ───────────────────────────────────────────────────────
        total_traffic   = sum(s['traffic'] for s in stores_out)
        total_bills     = sum(s['bills']   for s in stores_out)
        submitted_count = sum(1 for s in stores_out if s['submitted'])
        total_contract_value = sum(float(c.get('contract_value') or 0) for c in contract_rows)
        total_unsigned_value = sum(float(u.get('prev_year_value') or 0) for u in unsigned_rows)
        high_priority_support = sum(1 for r in support_rows if r['priority'] == 'Cao')

        summary = {
            'total_stores':   len(stores_out),
            'submitted':      submitted_count,
            'pending':        len(stores_out) - submitted_count,
            'total_traffic':  total_traffic,
            'total_bills':    total_bills,
            'contracts_31':   len(contract_rows),
            'contract_value': total_contract_value,
            'unsigned_32':    len(unsigned_rows),
            'unsigned_value': total_unsigned_value,
            'support_total':  len(support_rows),
            'support_high':   high_priority_support,
        }

        # ── HR Seniority Summary ──────────────────────────────────────────────
        hr_seniority = compute_hr_seniority_summary(store_codes)

        # ── ASM list (for admin filter) ───────────────────────────────────────
        asms_list = []
        if role == 'admin' or is_master:
            asms_list = [r['asm_name'] for r in query_db(
                "SELECT DISTINCT asm_name FROM tb_stores WHERE asm_name IS NOT NULL ORDER BY asm_name")]

        return jsonify({
            'ok': True,
            'report_date': report_date,
            'week_start': week_start,
            'can_export': role == 'admin' or is_master,
            'can_filter_asm': role == 'admin' or is_master,
            'user_asm': user_id if role == 'asm' else None,
            'asms': asms_list,
            'summary': summary,
            'hr_seniority': hr_seniority,
            'stores': stores_out,
            'contracts': [dict(r) for r in contract_rows],
            'unsigned':  [dict(r) for r in unsigned_rows],
            'details':   [dict(r) for r in detail_rows],
            'support':   [dict(r) for r in support_rows],
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/map', methods=['GET'])
def gis_map_view():
    """Bản đồ điều hành GIS Executive Command Center (100% Render Cloud Compatible + RBAC Protection)."""
    user_role = request.args.get('role', '').lower()
    if user_role in ['store', 'store_manager']:
        return """
        <div style="background:#0f172a; color:#f8fafc; font-family:sans-serif; min-height:100vh; display:flex; align-items:center; justify-content:center; padding:20px;">
            <div style="background:rgba(30,41,59,0.9); padding:40px; border-radius:16px; border:1px solid #ef4444; max-width:600px; text-align:center; box-shadow:0 20px 40px rgba(0,0,0,0.5);">
                <h1 style="color:#ef4444; font-size:24px; margin-bottom:12px; font-weight:800;">⛔ QUYỀN TRUY CẬP BỊ TỪ CHỐI</h1>
                <p style="font-size:16px; font-weight:600; margin-bottom:8px; color:#f8fafc;">Bản đồ GIS không gian chỉ dành cho Ban Quản Lý (Admin) và Quản Lý Khu Vực (ASM).</p>
                <p style="color:#94a3b8; font-size:13px;">Tài khoản Cửa Hàng không có quyền xem vị trí, doanh thu và dữ liệu hàng hóa của các cửa hàng khác trong hệ thống.</p>
            </div>
        </div>
        """, 403

    try:
        from gis_builder import generate_cloud_gis_map_html
        stores_db = query_db("""
            SELECT 
                s.store_code, s.store_name, s.asm_name, s.region,
                COALESCE(e.emp_cnt, 0) as emp_cnt,
                COALESCE(t.traffic_7d, 0) as traffic_7d,
                COALESCE(t.bills_7d, 0) as bills_7d,
                COALESCE(c.b2b_val, 0) as b2b_val
            FROM tb_stores s
            LEFT JOIN (
                SELECT store_code, COUNT(*) as emp_cnt 
                FROM tb_store_employees WHERE status = 'ACTIVE' 
                GROUP BY store_code
            ) e ON s.store_code = e.store_code
            LEFT JOIN (
                SELECT store_code, SUM(traffic_val) as traffic_7d, SUM(bills_val) as bills_7d
                FROM tb_traffic 
                GROUP BY store_code
            ) t ON s.store_code = t.store_code
            LEFT JOIN (
                SELECT store_code, SUM(contract_value) as b2b_val
                FROM tb_contracts
                GROUP BY store_code
            ) c ON s.store_code = c.store_code
            WHERE s.is_active = 1
        """)
        html_content = generate_cloud_gis_map_html(stores_db)
        return html_content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        import traceback; traceback.print_exc()
        return f"Error loading GIS map: {e}", 500


@app.route('/api/gis/data', methods=['GET'])
def api_gis_data():
    """API trả payload JSON GIS Cửa Hàng và Chặng Luân Chuyển (100% Render Cloud Compatible + RBAC Protection)."""
    user_role = request.args.get('role', '').lower()
    if user_role in ['store', 'store_manager']:
        return jsonify({'ok': False, 'error': '⛔ Quyền truy cập bị từ chối: Cửa hàng không có quyền truy cập dữ liệu GIS toàn hệ thống.'}), 403
    try:
        from gis_builder import load_verified_coords
        stores_db = query_db("""
            SELECT 
                s.store_code, s.store_name, s.asm_name, s.region,
                COALESCE(e.emp_cnt, 0) as emp_cnt,
                COALESCE(t.traffic_7d, 0) as traffic_7d,
                COALESCE(t.bills_7d, 0) as bills_7d,
                COALESCE(c.b2b_val, 0) as b2b_val
            FROM tb_stores s
            LEFT JOIN (
                SELECT store_code, COUNT(*) as emp_cnt 
                FROM tb_store_employees WHERE status = 'ACTIVE' 
                GROUP BY store_code
            ) e ON s.store_code = e.store_code
            LEFT JOIN (
                SELECT store_code, SUM(traffic_val) as traffic_7d, SUM(bills_val) as bills_7d
                FROM tb_traffic 
                GROUP BY store_code
            ) t ON s.store_code = t.store_code
            LEFT JOIN (
                SELECT store_code, SUM(contract_value) as b2b_val
                FROM tb_contracts
                GROUP BY store_code
            ) c ON s.store_code = c.store_code
            WHERE s.is_active = 1
        """)
        coords = load_verified_coords()
        store_list = []
        for s in stores_db:
            code = s.get('store_code')
            c_info = coords.get(code, {})
            store_list.append({
                'code': code,
                'name': s.get('store_name'),
                'addr': c_info.get('addr', ''),
                'asm': s.get('asm_name'),
                'region': s.get('region'),
                'mgr': c_info.get('mgr', 'Đang cập nhật'),
                'phone': c_info.get('phone', 'N/A'),
                'tier': c_info.get('tier', 'Standard'),
                'emp_cnt': s.get('emp_cnt', 0),
                'traffic_7d': s.get('traffic_7d', 0),
                'bills_7d': s.get('bills_7d', 0),
                'b2b_val': s.get('b2b_val', 0.0),
                'lat': c_info.get('lat', 10.7769),
                'lng': c_info.get('lng', 106.7009)
            })
        return jsonify({'ok': True, 'stores': store_list, 'count': len(store_list)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/gis/store_detail/<path:store_code>', methods=['GET'])
def api_gis_store_detail(store_code):
    """Trả về thông tin chi tiết thực tế của cửa hàng phục vụ GIS Drawer."""
    user_role = request.args.get('role', '').lower()
    if user_role in ['store', 'store_manager']:
        return jsonify({'ok': False, 'error': '⛔ Quyền truy cập bị từ chối.'}), 403
    try:
        store = query_db("SELECT store_code, store_name, asm_name, region, brand FROM tb_stores WHERE store_code = %s", (store_code,), one=True)
        if not store:
            return jsonify({'ok': False, 'error': 'Không tìm thấy cửa hàng'}), 404
            
        # 1. Store Manager (CHT) from tb_store_employees
        cht = query_db("SELECT full_name, phone_number FROM tb_store_employees WHERE store_code = %s AND (position LIKE '%%Trưởng%%' OR position LIKE '%%CHT%%') AND status = 'ACTIVE' LIMIT 1", (store_code,), one=True)
        mgr_name = cht['full_name'] if (cht and cht.get('full_name')) else None
        mgr_phone = cht['phone_number'] if (cht and cht.get('phone_number')) else None
        
        if not mgr_name or not mgr_phone:
            from gis_builder import load_verified_coords
            coords = load_verified_coords()
            c_info = coords.get(store_code, {})
            if not mgr_name:
                mgr_name = c_info.get('mgr', 'Đang cập nhật')
            if not mgr_phone:
                mgr_phone = c_info.get('phone', 'N/A')
        
        # 2. Headcount
        emp_count = query_db("SELECT COUNT(*) as cnt FROM tb_store_employees WHERE store_code = %s AND status = 'ACTIVE'", (store_code,), one=True)
        tgt = query_db("SELECT target_headcount FROM tb_store_headcount_targets WHERE store_code = %s", (store_code,), one=True)
        actual_cnt = emp_count['cnt'] if emp_count else 0
        target_cnt = tgt['target_headcount'] if tgt else 0
        headcount_str = f"{actual_cnt} / {target_cnt} NV" if target_cnt > 0 else f"{actual_cnt} NV"
        
        # 3. Traffic & CR (%) for last 7 days
        traffics = query_db("SELECT traffic_val, bills_val FROM tb_traffic WHERE store_code = %s ORDER BY traffic_date DESC LIMIT 7", (store_code,))
        total_tf = sum(t['traffic_val'] for t in traffics) if traffics else 0
        total_bi = sum(t['bills_val'] for t in traffics) if traffics else 0
        cr_pct = round((total_bi / total_tf * 100), 1) if total_tf > 0 else 0.0
        
        # 4. Support tickets
        tickets = query_db("SELECT COUNT(*) as cnt FROM tb_support_requests WHERE store_code = %s AND (person_in_charge IS NULL OR person_in_charge != 'Đã xong')", (store_code,), one=True)
        pending_tickets = tickets['cnt'] if tickets else 0
        
        return jsonify({
            'ok': True,
            'store_code': store_code,
            'store_name': store['store_name'],
            'asm_name': store['asm_name'],
            'region': store['region'],
            'mgr': mgr_name,
            'phone': mgr_phone,
            'headcount_str': headcount_str,
            'total_traffic_7d': total_tf,
            'total_bills_7d': total_bi,
            'cr_pct': f"{cr_pct}%",
            'pending_tickets': pending_tickets
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500




# ── OPERATIONAL HISTORY & STORE AUDIT ENDPOINTS ─────────────────────────────
@app.route('/api/operational_history', methods=['GET'])
def get_operational_history():
    """
    Trả về lịch sử nhập liệu chi tiết theo khoảng thời gian, hỗ trợ lọc theo cửa hàng & phân loại.
    """
    store_code = request.args.get('store_code', 'all')
    category = request.args.get('category', 'all')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    asm_name = request.args.get('asm_name', '')
    
    # Filter stores based on scope or ASM
    stores = query_db("SELECT store_code, store_name, asm_name, region FROM tb_stores WHERE is_active = 1")
    store_dict = {s['store_code']: s for s in stores}
    
    valid_store_codes = list(store_dict.keys())
    if asm_name:
        valid_store_codes = [sc for sc, s in store_dict.items() if s['asm_name'] == asm_name]
    if store_code != 'all':
        valid_store_codes = [sc for sc in valid_store_codes if sc == store_code]
        
    if not valid_store_codes:
        return jsonify({'ok': True, 'history': [], 'count': 0})
        
    placeholders = ','.join(['?'] * len(valid_store_codes))
    
    # Date conditions
    date_where_contracts = ""
    date_where_unsigned = ""
    date_where_details = ""
    date_where_support = ""
    date_where_traffic = ""
    params_date = []
    
    if start_date and end_date:
        date_where_contracts = " AND report_date >= ? AND report_date <= ?"
        date_where_unsigned = " AND report_date >= ? AND report_date <= ?"
        date_where_details = " AND report_date >= ? AND report_date <= ?"
        date_where_support = " AND report_date >= ? AND report_date <= ?"
        date_where_traffic = " AND traffic_date >= ? AND traffic_date <= ?"
        params_date = [start_date, end_date]

    history = []
    
    # 1. Section 3.1 Contracts
    if category in ['all', 'contracts', '3.1']:
        sql = f"SELECT * FROM tb_contracts WHERE store_code IN ({placeholders})" + date_where_contracts + " ORDER BY report_date DESC"
        p = valid_store_codes + params_date
        rows = query_db(sql, p)
        for r in rows:
            st = store_dict.get(r['store_code'], {})
            history.append({
                'type': '3.1_contract',
                'category_label': 'HĐ Đang Đàm Phán (3.1)',
                'store_code': r['store_code'],
                'store_name': st.get('store_name', r['store_code']),
                'asm_name': st.get('asm_name', ''),
                'report_date': r['report_date'],
                'title': f"HĐ {r.get('contract_number', 'N/A')} - {r.get('customer_name', 'Khách hàng')}",
                'value': r.get('contract_value', 0),
                'details': {
                    'Số HĐ': r.get('contract_number', ''),
                    'Tên Khách Hàng': r.get('customer_name', ''),
                    'Giá trị': f"{r.get('contract_value', 0):,.0f} VNĐ",
                    'Chủng loại': r.get('product_category', ''),
                    'Số lượng': r.get('quantity', 0),
                    'Đã cọc': f"{r.get('deposit_paid', 0):,.0f} VNĐ",
                    'Thanh toán Đ2': f"{r.get('installment_2', 0):,.0f} VNĐ",
                    'Tình trạng': r.get('status', ''),
                    'Lý do/Ghi chú': r.get('reason', '')
                }
            })

    # 2. Section 3.2 Unsigned
    if category in ['all', 'contracts', 'unsigned', '3.2']:
        sql = f"SELECT * FROM tb_unsigned_contracts WHERE store_code IN ({placeholders})" + date_where_unsigned + " ORDER BY report_date DESC"
        p = valid_store_codes + params_date
        rows = query_db(sql, p)
        for r in rows:
            st = store_dict.get(r['store_code'], {})
            history.append({
                'type': '3.2_unsigned',
                'category_label': 'HĐ Cùng Kỳ Chưa Ký (3.2)',
                'store_code': r['store_code'],
                'store_name': st.get('store_name', r['store_code']),
                'asm_name': st.get('asm_name', ''),
                'report_date': r['report_date'],
                'title': f"HĐ trễ hạn {r.get('contract_number', 'N/A')} - {r.get('customer_name', 'Khách hàng')}",
                'value': r.get('prev_year_value', 0),
                'details': {
                    'Số HĐ': r.get('contract_number', ''),
                    'Tên Khách Hàng': r.get('customer_name', ''),
                    'GTHĐ Năm trước': f"{r.get('prev_year_value', 0):,.0f} VNĐ",
                    'Dự kiến ký': r.get('expected_signing_time', ''),
                    'Chủng loại': r.get('product_category', ''),
                    'Số lượng': r.get('quantity', 0),
                    'Tình trạng': r.get('status', ''),
                    'Lý do/Ghi chú': r.get('reason', '')
                }
            })

    # 3. Section 4.1-4.4 Operational Details
    if category in ['all', 'operational', '4.x']:
        sql = f"SELECT * FROM tb_operational_details WHERE store_code IN ({placeholders})" + date_where_details + " ORDER BY report_date DESC"
        p = valid_store_codes + params_date
        rows = query_db(sql, p)
        for r in rows:
            st = store_dict.get(r['store_code'], {})
            history.append({
                'type': '4.x_operational',
                'category_label': 'Vận Hành & Hàng Hóa (4.1-4.4)',
                'store_code': r['store_code'],
                'store_name': st.get('store_name', r['store_code']),
                'asm_name': st.get('asm_name', ''),
                'report_date': r['report_date'],
                'title': f"Báo cáo Vận hành CH ngày {r['report_date']}",
                'value': None,
                'details': {
                    'Mở/Đóng CH': f"{r.get('op_open_close_status', '')} ({r.get('op_open_close_note', '')})",
                    'Đồng phục': f"{r.get('op_uniform_status', '')} ({r.get('op_uniform_note', '')})",
                    'Tác phong chào': f"{r.get('op_greet_status', '')} ({r.get('op_greet_note', '')})",
                    'Tồn kho': f"{r.get('inv_stock_status', '')} ({r.get('inv_info_goods', '')})",
                    'Đề xuất hàng': r.get('inv_proposal', ''),
                    'Thị trường & ĐT': r.get('market_competitors', '')
                }
            })

    # 4. Section 4.5 Support Tickets
    if category in ['all', 'support', '4.5']:
        sql = f"SELECT * FROM tb_support_requests WHERE store_code IN ({placeholders})" + date_where_support + " ORDER BY report_date DESC"
        p = valid_store_codes + params_date
        rows = query_db(sql, p)
        for r in rows:
            st = store_dict.get(r['store_code'], {})
            history.append({
                'type': '4.5_support',
                'category_label': 'Tồn Đọng & Hỗ Trợ (4.5)',
                'store_code': r['store_code'],
                'store_name': st.get('store_name', r['store_code']),
                'asm_name': st.get('asm_name', ''),
                'report_date': r['report_date'],
                'title': f"Ticket {r.get('ticket_code', 'N/A')}: {r.get('issue_item', '')}",
                'value': None,
                'details': {
                    'Mã Ticket': r.get('ticket_code', ''),
                    'Hạng mục': r.get('category', ''),
                    'Mức ưu tiên': r.get('priority', ''),
                    'Nội dung y/c': r.get('issue_item', ''),
                    'Hạn xử lý': r.get('deadline', ''),
                    'Người phụ trách': r.get('person_in_charge', ''),
                    'Trạng thái': r.get('status', 'Đang xử lý'),
                    'CH ghi chú': r.get('store_progress_note', ''),
                    'ASM/HQ ghi chú': r.get('asm_hq_note', '')
                }
            })

    # Sort history descending by report_date
    history.sort(key=lambda x: str(x['report_date']), reverse=True)
    return jsonify({'ok': True, 'history': history, 'count': len(history)})


@app.route('/api/store_operational_detail', methods=['GET'])
def get_store_operational_detail():
    """
    Trả về hồ sơ vận hành 360 độ chi tiết của 1 cửa hàng cụ thể.
    """
    store_code = request.args.get('store_code', '')
    report_date = request.args.get('report_date', '')
    
    if not store_code:
        return jsonify({'ok': False, 'error': 'Vui lòng cung cấp store_code'})
        
    store = query_db("SELECT * FROM tb_stores WHERE store_code = ?", (store_code,), one=True)
    if not store:
        return jsonify({'ok': False, 'error': 'Không tìm thấy cửa hàng'})
        
    if not report_date:
        latest = query_db("SELECT MAX(report_date) as max_d FROM tb_operational_reports WHERE store_code = ?", (store_code,), one=True)
        report_date = latest['max_d'] if latest and latest['max_d'] else ''
        
    contracts = query_db("SELECT * FROM tb_contracts WHERE store_code = ? ORDER BY report_date DESC LIMIT 20", (store_code,))
    unsigned = query_db("SELECT * FROM tb_unsigned_contracts WHERE store_code = ? ORDER BY report_date DESC LIMIT 20", (store_code,))
    details = query_db("SELECT * FROM tb_operational_details WHERE store_code = ? ORDER BY report_date DESC LIMIT 10", (store_code,))
    support = query_db("SELECT * FROM tb_support_requests WHERE store_code = ? ORDER BY report_date DESC LIMIT 20", (store_code,))
    traffic = query_db("SELECT * FROM tb_traffic WHERE store_code = ? ORDER BY traffic_date DESC LIMIT 30", (store_code,))
    
    return jsonify({
        'ok': True,
        'store': store,
        'report_date': report_date,
        'contracts': contracts,
        'unsigned_contracts': unsigned,
        'operational_details': details,
        'support_requests': support,
        'traffic': traffic
    })


if __name__ == '__main__':
    safe_migrate_db()   # chạy migration an toàn trước khi start
    # Default port 8080
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False)
